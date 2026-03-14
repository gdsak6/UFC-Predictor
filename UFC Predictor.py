"""
UFC PREDICTOR
===================================
XGBoost + LightGBM + CatBoost + RandomForest + MLP stacked with LogisticRegression meta-learner.

"""

# ─────────────────────────────────────────────────────────────────────────────
# STANDARD LIBRARY IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import os
os.environ["PYTHONWARNINGS"] = "ignore"
import sys
import random
import time
import math
import copy
import atexit
import shutil
import warnings
import threading
import traceback
import multiprocessing as mp
from datetime import datetime, date
from collections import defaultdict

# ─────────────────────────────────────────────────────────────────────────────
# THIRD-PARTY IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestClassifier, VotingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.preprocessing import StandardScaler, LabelEncoder
from sklearn.decomposition import TruncatedSVD
from sklearn.calibration import CalibratedClassifierCV
from sklearn.cluster import KMeans
from sklearn.metrics import accuracy_score, log_loss
from sklearn.model_selection import cross_val_score, TimeSeriesSplit, KFold
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

warnings.filterwarnings("ignore")

# Captured before any code can redirect them (used by _GUIConsole to mirror
# GUI output to the real terminal).
_REAL_STDOUT = sys.stdout
_REAL_STDERR = sys.stderr

# ─────────────────────────────────────────────────────────────────────────────
# OPTIONAL LIBRARY DETECTION
# ─────────────────────────────────────────────────────────────────────────────
try:
    import xgboost as xgb
    HAS_XGB = True
except ImportError:
    HAS_XGB = False

try:
    import lightgbm as lgb
    HAS_LGB = True
except ImportError:
    HAS_LGB = False

try:
    import catboost as cb
    HAS_CAT = True
except ImportError:
    HAS_CAT = False

try:
    import optuna
    optuna.logging.set_verbosity(optuna.logging.WARNING)
    HAS_OPTUNA = True
except ImportError:
    HAS_OPTUNA = False

# ─────────────────────────────────────────────────────────────────────────────
# GLOBAL CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
RANDOM_SEED = 42
# Seed every source of randomness that can affect training results.
# Together these make the full pipeline deterministic on CPU.
# NOTE: XGBoost with device='cuda' still has non-deterministic float-reduction
# order from parallel GPU threads; to eliminate that last source of variance
# set device='cpu' in the XGB params (at the cost of GPU speed).
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)
os.environ["PYTHONHASHSEED"]        = str(RANDOM_SEED)
os.environ["TF_DETERMINISTIC_OPS"]  = "1"   # no-op if TF absent, harmless

SAFE_N_JOBS = max(1, mp.cpu_count() // 2)

# Path setup
if getattr(sys, "frozen", False):
    SCRIPT_DIR = os.path.dirname(sys.executable)
else:
    SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))

DEFAULT_DATA_PATH = os.path.join(SCRIPT_DIR, "fight_data.csv")

# ─────────────────────────────────────────────────────────────────────────────
# CLEANUP
# ─────────────────────────────────────────────────────────────────────────────
def cleanup_temp_files():
    catboost_dir = os.path.join(SCRIPT_DIR, "catboost_info")
    if os.path.isdir(catboost_dir):
        try:
            shutil.rmtree(catboost_dir)
        except Exception:
            pass

atexit.register(cleanup_temp_files)

# ─────────────────────────────────────────────────────────────────────────────
# GPU DETECTION
# ─────────────────────────────────────────────────────────────────────────────
def detect_gpu():
    gpu_info = {"xgb": False, "lgb": False, "cat": False}
    X_dummy = np.random.rand(2, 4)
    y_dummy = np.array([0, 1])

    if HAS_XGB:
        try:
            m = xgb.XGBClassifier(device="cuda", n_estimators=1, verbosity=0,
                                   random_state=RANDOM_SEED, eval_metric="logloss")
            m.fit(X_dummy, y_dummy)
            gpu_info["xgb"] = True
        except Exception:
            pass

    # LightGBM always runs on CPU to avoid GPU out-of-resources errors
    gpu_info["lgb"] = False

    if HAS_CAT:
        try:
            m = cb.CatBoostClassifier(task_type="GPU", iterations=1, verbose=0,
                                       random_seed=RANDOM_SEED)
            m.fit(X_dummy, y_dummy)
            gpu_info["cat"] = True
        except Exception:
            pass

    return gpu_info

# ─────────────────────────────────────────────────────────────────────────────
# TERMINAL OUTPUT HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def print_section(title):
    width = 70
    print("\n" + "=" * width)
    print(f"  {title}")
    print("=" * width)

def print_step(msg):
    print(f"  >> {msg}")

def print_metric(label, value):
    print(f"  {label:<35} {value}")

def print_divider():
    print("  " + "-" * 66)

# ─────────────────────────────────────────────────────────────────────────────
# FEATURE ENGINEER
# ─────────────────────────────────────────────────────────────────────────────
class FeatureEngineer:
    """
    Handles advanced fighter statistics: ELO, Glicko-2, weight-class Z-scores,
    common opponent matrix, K-Means style clustering.
    """

    def __init__(self):
        # ELO
        self.elo_ratings = {}

        # Glicko-2
        self.glicko_ratings = {}   # fighter -> (rating, RD, vol)
        self.GLICKO_MU = 1500.0
        self.GLICKO_PHI = 200.0
        self.GLICKO_SIGMA = 0.06
        self.GLICKO_TAU = 0.5
        self.GLICKO_SCALE = 173.7178

        # Weight-class Z-score stats
        self.weight_class_stats = defaultdict(lambda: defaultdict(list))

        # Common opponents
        self.fighter_opponents = defaultdict(set)
        self.fight_outcomes = {}  # (winner, loser) -> result

        # K-Means clustering
        self.kmeans = None
        self.cluster_scaler = StandardScaler()
        self.n_clusters = 8
        self.style_features = ["SLpM", "SApM", "TD", "Sub", "Finish"]
        self.fighter_cluster = {}

        # Style performance tracking
        self.style_performance = defaultdict(lambda: defaultdict(list))

    # ── ELO ──────────────────────────────────────────────────────────────────
    def elo_initial(self, fighter):
        if fighter not in self.elo_ratings:
            self.elo_ratings[fighter] = 1500.0
        return self.elo_ratings[fighter]

    def elo_expected(self, rA, rB):
        return 1.0 / (1.0 + 10 ** ((rB - rA) / 400.0))

    def elo_k_factor(self, fighter_fights, is_title, method, result, finish_round=0, winner_streak=0, opponent_elo=1500, loser_streak=0):
        k = 32
        if fighter_fights < 5:
            k = 48
        elif fighter_fights > 20:
            k = 24
        if is_title:
            k *= 2
        if method in ("KO/TKO", "Submission") and result == "win":
            k *= 1.2
            # Round-specific dominance modifier
            if finish_round == 1:
                k *= 1.5
            elif finish_round == 2:
                k *= 1.3
            elif finish_round == 3:
                k *= 1.2
            # else round 4+ -> no extra modifier (×1.0)
        # Split/majority decisions are less informative — close fights carry a
        # smaller rating update in both directions (winner and loser).
        if ("Split" in method or "Majority" in method):
            k *= 0.8
        # Streak multiplier: 3+ fight win streak
        if winner_streak >= 3 and result == "win":
            k *= 1.1
        # Opponent strength bonus (existing)
        if opponent_elo > 1600 and result == "win":
            k *= 1.15
        # Phase 7: Upset bonus — beating a much higher-rated opponent
        if result == "win" and opponent_elo > 1600:
            upset_factor = min((opponent_elo - 1600) / 400.0, 0.5)
            k *= (1.0 + upset_factor)
        # Phase 7: Loss streak penalty
        if result == "loss" and loser_streak <= -3:
            k *= 1.25
        return k

    def elo_update(self, r_fighter, b_fighter, winner, is_title, method, r_fights, b_fights,
                   finish_round=0, winner_streak=0, opponent_elo=1500):
        rA = self.elo_initial(r_fighter)
        rB = self.elo_initial(b_fighter)
        eA = self.elo_expected(rA, rB)
        eB = 1.0 - eA

        if winner == "Red":
            sA, sB = 1.0, 0.0
        elif winner == "Blue":
            sA, sB = 0.0, 1.0
        else:
            sA, sB = 0.5, 0.5

        kA = self.elo_k_factor(r_fights, is_title, method, "win" if winner == "Red" else "loss",
                               finish_round=finish_round,
                               winner_streak=winner_streak if winner == "Red" else 0,
                               opponent_elo=opponent_elo if winner == "Red" else rA)
        kB = self.elo_k_factor(b_fights, is_title, method, "win" if winner == "Blue" else "loss",
                               finish_round=finish_round,
                               winner_streak=winner_streak if winner == "Blue" else 0,
                               opponent_elo=opponent_elo if winner == "Blue" else rB)

        pre_rA = rA
        pre_rB = rB
        self.elo_ratings[r_fighter] = rA + kA * (sA - eA)
        self.elo_ratings[b_fighter] = rB + kB * (sB - eB)
        return pre_rA, pre_rB

    def elo_get(self, fighter):
        return self.elo_ratings.get(fighter, 1500.0)

    # ── GLICKO-2 ─────────────────────────────────────────────────────────────
    def glicko2_initial(self, fighter):
        if fighter not in self.glicko_ratings:
            self.glicko_ratings[fighter] = (self.GLICKO_MU, self.GLICKO_PHI, self.GLICKO_SIGMA)
        return self.glicko_ratings[fighter]

    def _g(self, phi):
        return 1.0 / math.sqrt(1 + 3 * phi ** 2 / math.pi ** 2)

    def _E(self, mu, mu_j, phi_j):
        return 1.0 / (1.0 + math.exp(-self._g(phi_j) * (mu - mu_j)))

    def glicko2_update(self, fighter, opponents):
        """opponents: list of (opp_rating, opp_RD, score)"""
        mu, phi, sigma = self.glicko2_initial(fighter)
        mu_s = (mu - self.GLICKO_MU) / self.GLICKO_SCALE
        phi_s = phi / self.GLICKO_SCALE

        if not opponents:
            phi_star = math.sqrt(phi_s ** 2 + sigma ** 2)
            self.glicko_ratings[fighter] = (mu, phi_star * self.GLICKO_SCALE, sigma)
            return

        v_inv = 0.0
        delta_sum = 0.0
        for opp_r, opp_rd, score in opponents:
            mu_j = (opp_r - self.GLICKO_MU) / self.GLICKO_SCALE
            phi_j = opp_rd / self.GLICKO_SCALE
            g_j = self._g(phi_j)
            E_j = self._E(mu_s, mu_j, phi_j)
            v_inv += g_j ** 2 * E_j * (1 - E_j)
            delta_sum += g_j * (score - E_j)

        v = 1.0 / v_inv if v_inv > 0 else 1e6
        delta = v * delta_sum

        # Iterative sigma update
        a = math.log(sigma ** 2)
        tau = self.GLICKO_TAU

        def f(x):
            ex = math.exp(x)
            num = ex * (delta ** 2 - phi_s ** 2 - v - ex)
            den = 2 * (phi_s ** 2 + v + ex) ** 2
            return num / den - (x - a) / (tau ** 2)

        A = a
        if delta ** 2 > phi_s ** 2 + v:
            B = math.log(delta ** 2 - phi_s ** 2 - v)
        else:
            k = 1
            while f(a - k * tau) < 0:
                k += 1
            B = a - k * tau

        fA, fB = f(A), f(B)
        for _ in range(100):
            C = A + (A - B) * fA / (fB - fA)
            fC = f(C)
            if fC * fB < 0:
                A, fA = B, fB
            else:
                fA /= 2
            B, fB = C, fC
            if abs(B - A) < 1e-6:
                break
        new_sigma = math.exp(A / 2)

        phi_star = math.sqrt(phi_s ** 2 + new_sigma ** 2)
        new_phi_s = 1.0 / math.sqrt(1.0 / phi_star ** 2 + 1.0 / v)
        new_mu_s = mu_s + new_phi_s ** 2 * delta_sum

        self.glicko_ratings[fighter] = (
            new_mu_s * self.GLICKO_SCALE + self.GLICKO_MU,
            new_phi_s * self.GLICKO_SCALE,
            new_sigma
        )

    def glicko2_get(self, fighter):
        r, rd, vol = self.glicko2_initial(fighter)
        return r, rd, vol

    # ── WEIGHT-CLASS Z-SCORES ────────────────────────────────────────────────
    def update_weight_class_stats(self, weight_class, year, stats_dict):
        for feat, val in stats_dict.items():
            if val is not None and not math.isnan(float(val)):
                self.weight_class_stats[(weight_class, year)][feat].append(float(val))

    def get_z_score(self, weight_class, year, feat, value):
        key = (weight_class, year)
        if key not in self.weight_class_stats:
            return 0.0
        vals = self.weight_class_stats[key].get(feat, [])
        if len(vals) < 2:
            return 0.0
        mu = np.mean(vals)
        std = np.std(vals)
        if std < 1e-9:
            return 0.0
        return (value - mu) / std

    # ── COMMON OPPONENTS ─────────────────────────────────────────────────────
    def update_common_opponents(self, r_fighter, b_fighter, winner):
        self.fighter_opponents[r_fighter].add(b_fighter)
        self.fighter_opponents[b_fighter].add(r_fighter)
        if winner == "Red":
            self.fight_outcomes[(r_fighter, b_fighter)] = 1
            self.fight_outcomes[(b_fighter, r_fighter)] = 0
        elif winner == "Blue":
            self.fight_outcomes[(r_fighter, b_fighter)] = 0
            self.fight_outcomes[(b_fighter, r_fighter)] = 1
        else:
            self.fight_outcomes[(r_fighter, b_fighter)] = 0.5
            self.fight_outcomes[(b_fighter, r_fighter)] = 0.5

    def get_common_opponent_features(self, r_fighter, b_fighter):
        common = self.fighter_opponents[r_fighter] & self.fighter_opponents[b_fighter]
        n_common = len(common)
        r_wins_common = 0
        b_wins_common = 0
        for opp in common:
            r_wins_common += self.fight_outcomes.get((r_fighter, opp), 0.5)
            b_wins_common += self.fight_outcomes.get((b_fighter, opp), 0.5)
        return {
            "n_common_opponents": n_common,
            "r_wins_vs_common": r_wins_common,
            "b_wins_vs_common": b_wins_common,
            "common_opp_edge": r_wins_common - b_wins_common,
        }

    # ── K-MEANS CLUSTERING ───────────────────────────────────────────────────
    def fit_clusters(self, fighter_stats):
        """
        fighter_stats: dict of {fighter: {SLpM, SApM, TD, Sub, Finish}}
        """
        if len(fighter_stats) < self.n_clusters:
            return
        rows = []
        fighters = []
        for f, s in fighter_stats.items():
            row = [s.get(k, 0.0) for k in self.style_features]
            rows.append(row)
            fighters.append(f)
        X = np.array(rows, dtype=float)
        X = np.nan_to_num(X)
        X_scaled = self.cluster_scaler.fit_transform(X)
        self.kmeans = KMeans(n_clusters=self.n_clusters, random_state=RANDOM_SEED, n_init=10)
        labels = self.kmeans.fit_predict(X_scaled)
        for f, lbl in zip(fighters, labels):
            self.fighter_cluster[f] = int(lbl)

    def get_fighter_cluster(self, fighter):
        return self.fighter_cluster.get(fighter, -1)

    def update_style_performance(self, cluster, opponent_cluster, won):
        self.style_performance[cluster][opponent_cluster].append(1 if won else 0)

    def get_style_matchup_features(self, r_cluster, b_cluster):
        r_vs_b = self.style_performance[r_cluster].get(b_cluster, [])
        b_vs_r = self.style_performance[b_cluster].get(r_cluster, [])
        r_winrate = np.mean(r_vs_b) if r_vs_b else 0.5
        b_winrate = np.mean(b_vs_r) if b_vs_r else 0.5
        return {
            "r_style_win_vs_opp_cluster": r_winrate,
            "b_style_win_vs_opp_cluster": b_winrate,
            "style_matchup_edge": r_winrate - b_winrate,
            "r_cluster": r_cluster,
            "b_cluster": b_cluster,
        }


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 6: PURGED TIME SERIES SPLIT
# ─────────────────────────────────────────────────────────────────────────────
class PurgedTimeSeriesSplit:
    """TimeSeriesSplit with a purge gap to prevent leakage between train/test."""
    def __init__(self, n_splits=5, purge_days=30):
        self.n_splits = n_splits
        self.purge_days = purge_days

    def get_n_splits(self, X=None, y=None, groups=None):
        return self.n_splits

    def split(self, X, y=None, groups=None):
        n = len(X)
        fold_size = n // (self.n_splits + 1)
        for i in range(self.n_splits):
            train_end = fold_size * (i + 1)
            test_start = train_end + max(1, self.purge_days // 7)
            test_end = min(test_start + fold_size, n)
            if test_end <= test_start:
                continue
            train_idx = np.arange(0, train_end)
            test_idx  = np.arange(test_start, test_end)
            yield train_idx, test_idx


# ─────────────────────────────────────────────────────────────────────────────
# MANUAL OOF STACKING ENSEMBLE
# ─────────────────────────────────────────────────────────────────────────────
class _ManualStackingEnsemble:
    """KFold-3 out-of-fold stacking that avoids StackingClassifier's NaN bug.

    StackingClassifier's internal cross_val_predict produces NaN meta-features
    for some base models (rf, mlp, cat), causing their LR meta-learner
    coefficients to be NaN and making those models useless. This class does
    the same OOF stacking explicitly, catching NaN/exceptions per model per
    fold so every base model contributes valid meta-features to the LR meta-
    learner. Base models are passed already-fitted; clones are used for OOF
    folds, and the originals are used for test-time inference.
    """

    def __init__(self, estimators, meta_C=0.05, n_splits=3, random_state=42):
        # estimators: list of (name, already-fitted estimator)
        self.estimators  = estimators
        self.meta_C      = meta_C
        self.n_splits    = n_splits
        self.random_state = random_state
        self.classes_         = None
        self.final_estimator_ = None

    def fit(self, X, y, sample_weight=None):
        from sklearn.base import clone as _clone
        y = np.asarray(y)
        self.classes_ = np.unique(y)
        nc = len(self.classes_)          # 2 for binary
        n  = X.shape[0]
        nm = len(self.estimators)

        sw = np.asarray(sample_weight) if sample_weight is not None else None

        # OOF meta-feature matrix: n_samples × (n_models × n_classes)
        meta_X_oof = np.full((n, nm * nc), 1.0 / nc)

        kf = KFold(n_splits=self.n_splits, shuffle=False)
        for tr_idx, va_idx in kf.split(X):
            X_tr_f, X_va_f = X[tr_idx], X[va_idx]
            y_tr_f = y[tr_idx]
            # OOF fold training uses equal weights so predictions are calibrated
            # across all eras. Passing recency weights into fold training skews
            # predictions toward recent patterns, producing miscalibrated OOF
            # probabilities for older fights and confusing the meta-learner.
            for ei, (name, est) in enumerate(self.estimators):
                col = ei * nc
                try:
                    fold_est = _clone(est)
                    fold_est.fit(X_tr_f, y_tr_f)
                    proba = fold_est.predict_proba(X_va_f)
                    if proba.shape[1] != nc or np.any(np.isnan(proba)):
                        raise ValueError("bad proba shape or NaN")
                    meta_X_oof[va_idx, col:col + nc] = proba
                except Exception:
                    # Fall back to uniform for this model/fold — still contributes
                    meta_X_oof[va_idx, col:col + nc] = 1.0 / nc

        # Train LR meta-learner on OOF meta-features (with recency weights)
        self.final_estimator_ = LogisticRegression(
            C=self.meta_C, max_iter=1000,
            random_state=self.random_state, solver="lbfgs"
        )
        self.final_estimator_.fit(meta_X_oof, y, sample_weight=sw)
        return self

    def _meta_features(self, X):
        nc = len(self.classes_)
        nm = len(self.estimators)
        meta_X = np.full((X.shape[0], nm * nc), 1.0 / nc)
        for ei, (name, est) in enumerate(self.estimators):
            col = ei * nc
            try:
                proba = est.predict_proba(X)
                if proba.shape[1] != nc or np.any(np.isnan(proba)):
                    raise ValueError("bad proba")
                meta_X[:, col:col + nc] = proba
            except Exception:
                meta_X[:, col:col + nc] = 1.0 / nc
        return meta_X

    def predict(self, X):
        return self.final_estimator_.predict(self._meta_features(X))

    def predict_proba(self, X):
        return self.final_estimator_.predict_proba(self._meta_features(X))


# ─────────────────────────────────────────────────────────────────────────────
# UFC PREDICTOR
# ─────────────────────────────────────────────────────────────────────────────
class UFCPredictor:

    def __init__(self, data_path=None, status_callback=None):
        self.data_path = data_path or DEFAULT_DATA_PATH
        self.status_callback = status_callback or (lambda msg: None)
        self.df = None
        self.feature_engineer = FeatureEngineer()
        self.feature_cols = []
        self.scaler = StandardScaler()
        self.label_encoder = LabelEncoder()
        self.models = {}
        self.stacking_clf = None
        self.method_clf = None
        self.method_scaler = StandardScaler()
        self.is_trained = False
        self.gpu_info = {}

        # SVD decomposers per bucket
        self.svd_striking = TruncatedSVD(n_components=5, random_state=RANDOM_SEED)
        self.svd_grappling = TruncatedSVD(n_components=5, random_state=RANDOM_SEED)
        self.svd_physical = TruncatedSVD(n_components=3, random_state=RANDOM_SEED)
        self.svd_form = TruncatedSVD(n_components=5, random_state=RANDOM_SEED)
        self.svd_fitted = False
        # Exact column lists used at SVD fit time (saved so transform uses identical input shape)
        self.svd_striking_cols  = []
        self.svd_grappling_cols = []
        self.svd_physical_cols  = []
        self.svd_form_cols      = []
        self._selected_d_indices = []

        self.weight_classes = []
        self.all_fighters = []
        self.predictions = []

    def _log(self, msg):
        self.status_callback(msg)
        print_step(msg)

    # ── LOAD DATA ────────────────────────────────────────────────────────────
    def load_data(self):
        print_section("LOADING DATA")
        self._log(f"Reading: {self.data_path}")
        self.df = pd.read_csv(self.data_path, low_memory=False)
        self._log(f"Loaded {len(self.df):,} rows x {len(self.df.columns)} columns")

        # Parse dates
        self.df["event_date"] = pd.to_datetime(self.df["event_date"], format="%m/%d/%Y", errors="coerce")
        self.df = self.df.sort_values("event_date").reset_index(drop=True)

        # Normalize winner
        self.df["winner"] = self.df["winner"].str.strip()
        self.df["method"] = self.df["method"].str.strip()

        # Normalize method labels
        method_map = {
            'Decision - Unanimous': 'Decision',
            'Decision - Split': 'Decision',
            'Decision - Majority': 'Decision',
            'TKO - Doctor\'s Stoppage': 'KO/TKO',
            'Could Not Continue': None,  # drop
            'Overturned': None,          # drop
            'DQ': None,                  # drop
            'No Contest': None,          # drop
        }
        self.df['method'] = self.df['method'].apply(lambda m: method_map.get(str(m).strip(), str(m).strip()) if pd.notna(m) else m)
        self.df = self.df[self.df['method'].notna() & self.df['method'].isin(['Decision', 'KO/TKO', 'Submission'])]

        # Keep only decisive outcomes for training
        self.df = self.df[self.df["winner"].isin(["Red", "Blue", "Draw"])].reset_index(drop=True)

        self.weight_classes = sorted(self.df["weight_class"].dropna().unique().tolist())
        self.all_fighters = sorted(
            set(self.df["r_fighter"].dropna().tolist()) | set(self.df["b_fighter"].dropna().tolist())
        )
        print_metric("Total fights:", len(self.df))
        print_metric("Unique fighters:", len(self.all_fighters))
        print_metric("Weight classes:", len(self.weight_classes))

    # ── FIX DATA LEAKAGE ─────────────────────────────────────────────────────
    def fix_data_leakage(self):
        print_section("FIXING DATA LEAKAGE")
        self._log("Recalculating fighter stats chronologically...")

        stats_to_track = {
            "wins": 0, "losses": 0, "draws": 0,
            "ko_wins": 0, "sub_wins": 0, "dec_wins": 0,
            "ko_losses": 0, "sub_losses": 0, "dec_losses": 0,
            "total_fights": 0,
            "title_fights": 0, "title_wins": 0,
            "recent_wins": 0, "recent_losses": 0,
            "win_streak": 0, "loss_streak": 0,
            "finish_rate": 0.0,
            "recent_finish_rate": 0.0,
            "total_fight_time": 0.0,
            "avg_fight_time": 0.0,
            # Per-fight stats accumulated
            "total_sig_str": 0, "total_sig_str_att": 0,
            "total_str": 0, "total_str_att": 0,
            "total_td": 0, "total_td_att": 0,
            "total_sub_att": 0,
            "total_kd": 0,
            "total_head": 0, "total_body": 0, "total_leg": 0,
            "total_distance": 0, "total_clinch": 0, "total_ground": 0,
            "total_ctrl_sec": 0, "total_rev": 0,
            # Rolling window (last 3 fights)
            "rolling3_wins": 0, "rolling3_sig_str": 0.0,
            "rolling3_td": 0.0, "rolling3_kd": 0.0,
            "rolling3_sub_att": 0.0, "rolling3_ctrl": 0.0,
            # Rolling window (last 5 fights)
            "rolling5_wins": 0, "rolling5_sig_str": 0.0,
            "rolling5_td": 0.0, "rolling5_kd": 0.0,
            # Pro stats (cumulative averages)
            "pro_SLpM": 0.0, "pro_SApM": 0.0,
            "pro_sig_str_acc": 0.0, "pro_str_def": 0.0,
            "pro_td_avg": 0.0, "pro_td_acc": 0.0,
            "pro_td_def": 0.0, "pro_sub_avg": 0.0,
            # Opponent quality
            "avg_opp_wins": 0.0, "avg_opp_losses": 0.0,
            # Last fight result
            "last_result": None,
            # Fight history for rolling
            "_history": [],
            # KD absorbed by opponent landing on this fighter
            "kd_absorbed": 0,
            # Fight dates for days-since-last calculation
            "fight_dates": [],
            # Opponent ELO history for quality tracking
            "opp_elo_history": [],
            # Phase 3: Positional striking
            "distance_strikes": 0, "clinch_strikes": 0, "ground_strikes": 0,
            "head_strikes": 0, "body_strikes": 0, "leg_strikes": 0,
            "total_positional_strikes": 0,
            # Phase 3: Defense tracking
            "str_def_attempts": 0, "str_def_success": 0,
            "td_def_attempts": 0, "td_def_success": 0,
            # Phase 3: Finish timing
            "early_finishes": 0,
            "late_finishes": 0,
            "first_round_kos": 0,
            "total_rounds_fought": 0,
            "five_round_fights": 0,
            "championship_rounds": 0,
            # Phase 3: Rolling windows at depth 10
            "rolling10_history": [],
            # Phase 3: Career arc
            "career_elo_peak": 1500.0,
            "fights_since_peak": 0,
            "weight_class_history": [],
            "weight_class_tenure": 0,
            # Phase 3: Last fight details
            "last_fight_method": None,
            "last_fight_finish_round": 0,
            "last_fight_was_finish": False,
            "last_fight_was_win": False,
            # Phase 3: Opponent-adjusted
            "vs_elite_wins": 0, "vs_elite_fights": 0,
            "vs_striker_wins": 0, "vs_striker_fights": 0,
            "vs_grappler_wins": 0, "vs_grappler_fights": 0,
        }

        fighter_stats = defaultdict(lambda: copy.deepcopy(stats_to_track))
        fighter_fights_count = defaultdict(int)

        leakage_cols_r = [c for c in self.df.columns if c.startswith("r_")]
        leakage_cols_b = [c for c in self.df.columns if c.startswith("b_")]

        # New computed columns
        new_cols = [
            "r_pre_wins", "r_pre_losses", "r_pre_draws",
            "b_pre_wins", "b_pre_losses", "b_pre_draws",
            "r_pre_ko_wins", "r_pre_sub_wins", "r_pre_dec_wins",
            "b_pre_ko_wins", "b_pre_sub_wins", "b_pre_dec_wins",
            "r_pre_total_fights", "b_pre_total_fights",
            "r_pre_finish_rate", "b_pre_finish_rate",
            "r_pre_win_streak", "b_pre_win_streak",
            "r_pre_loss_streak", "b_pre_loss_streak",
            "r_pre_title_fights", "b_pre_title_fights",
            "r_pre_title_wins", "b_pre_title_wins",
            "r_pre_avg_fight_time", "b_pre_avg_fight_time",
            "r_pre_sig_str_acc", "b_pre_sig_str_acc",
            "r_pre_td_acc", "b_pre_td_acc",
            "r_pre_sub_att_rate", "b_pre_sub_att_rate",
            "r_pre_kd_rate", "b_pre_kd_rate",
            "r_pre_ctrl_avg", "b_pre_ctrl_avg",
            "r_pre_SLpM", "b_pre_SLpM",
            "r_pre_SApM", "b_pre_SApM",
            "r_pre_td_avg", "b_pre_td_avg",
            "r_rolling3_wins", "b_rolling3_wins",
            "r_rolling3_sig_str", "b_rolling3_sig_str",
            "r_rolling3_td", "b_rolling3_td",
            "r_rolling3_kd", "b_rolling3_kd",
            "r_rolling3_sub_att", "b_rolling3_sub_att",
            "r_rolling5_wins", "b_rolling5_wins",
            "r_rolling5_sig_str", "b_rolling5_sig_str",
            "r_rolling5_td", "b_rolling5_td",
            "r_rolling5_kd", "b_rolling5_kd",
            # Glicko-2 pre-fight snapshots (to prevent data leakage in Tier 4)
            "r_glicko_pre_r", "r_glicko_pre_rd", "r_glicko_pre_vol",
            "b_glicko_pre_r", "b_glicko_pre_rd", "b_glicko_pre_vol",
            # Days since last fight (ring rust)
            "r_days_since_last", "b_days_since_last",
            # KD absorbed over career (chin deterioration)
            "r_pre_kd_absorbed", "b_pre_kd_absorbed",
            # Opponent quality: avg ELO of last 5 opponents
            "r_avg_opp_elo_L5", "b_avg_opp_elo_L5",
            # Trajectory: linear regression slope of last 3 fight results
            "r_trajectory_3", "b_trajectory_3",
            # Phase 3: Positional striking percentages
            "r_pre_distance_pct", "b_pre_distance_pct",
            "r_pre_clinch_pct", "b_pre_clinch_pct",
            "r_pre_ground_pct", "b_pre_ground_pct",
            "r_pre_head_pct", "b_pre_head_pct",
            "r_pre_body_pct", "b_pre_body_pct",
            "r_pre_leg_pct", "b_pre_leg_pct",
            # Phase 3: Defense rates
            "r_pre_str_def", "b_pre_str_def",
            "r_pre_td_def", "b_pre_td_def",
            # Phase 3: Finish timing rates
            "r_pre_early_finish_rate", "b_pre_early_finish_rate",
            "r_pre_late_finish_rate", "b_pre_late_finish_rate",
            "r_pre_first_round_ko_rate", "b_pre_first_round_ko_rate",
            "r_pre_total_rounds_fought", "b_pre_total_rounds_fought",
            "r_pre_five_round_fights", "b_pre_five_round_fights",
            # Phase 3: Rolling10 aggregates
            "r_rolling10_wins", "b_rolling10_wins",
            "r_rolling10_sig_str", "b_rolling10_sig_str",
            "r_rolling10_td", "b_rolling10_td",
            "r_rolling10_kd", "b_rolling10_kd",
            "r_rolling10_finishes", "b_rolling10_finishes",
            # Phase 3: Career arc
            "r_career_elo_peak", "b_career_elo_peak",
            "r_fights_since_peak", "b_fights_since_peak",
            # Phase 3: Opponent-adjusted win rates
            "r_vs_elite_win_rate", "b_vs_elite_win_rate",
            "r_vs_striker_win_rate", "b_vs_striker_win_rate",
            "r_vs_grappler_win_rate", "b_vs_grappler_win_rate",
            # Phase 3: Last fight momentum
            "r_last_fight_was_win", "b_last_fight_was_win",
            "r_last_fight_was_finish", "b_last_fight_was_finish",
            # Volatility & career arc (computed from rolling10_history)
            "r_pre_finish_rate_l5",  "b_pre_finish_rate_l5",
            "r_pre_finish_rate_l10", "b_pre_finish_rate_l10",
            "r_pre_slpm_cv",         "b_pre_slpm_cv",
            "r_pre_mileage_adj_age", "b_pre_mileage_adj_age",
            # Rolling SLpM & volatility snapshots (for 7 missing features)
            "r_pre_rolling3_wins",        "b_pre_rolling3_wins",
            "r_pre_rolling5_slpm",        "b_pre_rolling5_slpm",
            "r_pre_slpm_std_l10",         "b_pre_slpm_std_l10",
            "r_pre_damage_ratio_std_l10", "b_pre_damage_ratio_std_l10",
            "r_pre_tactical_evolution",   "b_pre_tactical_evolution",
            # Exponentially-weighted rolling stats (Item 1)
            "r_ewm_wins",    "b_ewm_wins",
            "r_ewm_sig_str", "b_ewm_sig_str",
            "r_ewm_td",      "b_ewm_td",
            "r_ewm_kd",      "b_ewm_kd",
            "r_ewm_sub_att", "b_ewm_sub_att",
            # Loss method breakdown (Item 5)
            "r_pre_ko_losses",  "b_pre_ko_losses",
            "r_pre_sub_losses", "b_pre_sub_losses",
            "r_pre_dec_losses", "b_pre_dec_losses",
        ]
        for col in new_cols:
            self.df[col] = 0.0

        # Exponentially-decayed weighted average over history entries.
        # Most recent = weight 1.0; each fight further back decays by e^(-0.5).
        # Positional (not date-based) since _history entries don't store dates.
        def _ewm_hist(entries, key):
            if not entries:
                return 0.0
            vals = [h.get(key, 0) for h in entries]
            n = len(vals)
            w = np.array([np.exp(-0.5 * (n - 1 - i)) for i in range(n)])
            w /= w.sum()
            return float(np.dot(w, vals))

        for idx, row in self.df.iterrows():
            r = str(row.get("r_fighter", "")).strip()
            b = str(row.get("b_fighter", "")).strip()
            winner = str(row.get("winner", "")).strip()
            method = str(row.get("method", "")).strip()
            is_title = bool(row.get("is_title_bout", False))

            rs = fighter_stats[r]
            bs = fighter_stats[b]

            # ── Snapshot pre-fight stats ──────────────────────────────────
            def _safe(v, default=0.0):
                if v is None:
                    return default
                try:
                    f = float(v)
                    return default if math.isnan(f) else f
                except (TypeError, ValueError):
                    return default

            r_nf = max(rs["total_fights"], 1)
            b_nf = max(bs["total_fights"], 1)

            self.df.at[idx, "r_pre_wins"] = rs["wins"]
            self.df.at[idx, "r_pre_losses"] = rs["losses"]
            self.df.at[idx, "r_pre_draws"] = rs["draws"]
            self.df.at[idx, "b_pre_wins"] = bs["wins"]
            self.df.at[idx, "b_pre_losses"] = bs["losses"]
            self.df.at[idx, "b_pre_draws"] = bs["draws"]
            self.df.at[idx, "r_pre_ko_wins"] = rs["ko_wins"]
            self.df.at[idx, "r_pre_sub_wins"] = rs["sub_wins"]
            self.df.at[idx, "r_pre_dec_wins"] = rs["dec_wins"]
            self.df.at[idx, "b_pre_ko_wins"] = bs["ko_wins"]
            self.df.at[idx, "b_pre_sub_wins"] = bs["sub_wins"]
            self.df.at[idx, "b_pre_dec_wins"] = bs["dec_wins"]
            # Loss method breakdown — KO losses signal chin deterioration,
            # sub losses signal grappling vulnerability
            self.df.at[idx, "r_pre_ko_losses"]  = rs["ko_losses"]
            self.df.at[idx, "r_pre_sub_losses"] = rs["sub_losses"]
            self.df.at[idx, "r_pre_dec_losses"] = rs["dec_losses"]
            self.df.at[idx, "b_pre_ko_losses"]  = bs["ko_losses"]
            self.df.at[idx, "b_pre_sub_losses"] = bs["sub_losses"]
            self.df.at[idx, "b_pre_dec_losses"] = bs["dec_losses"]
            self.df.at[idx, "r_pre_total_fights"] = rs["total_fights"]
            self.df.at[idx, "b_pre_total_fights"] = bs["total_fights"]
            self.df.at[idx, "r_pre_finish_rate"] = rs["finish_rate"]
            self.df.at[idx, "b_pre_finish_rate"] = bs["finish_rate"]
            self.df.at[idx, "r_pre_win_streak"] = rs["win_streak"]
            self.df.at[idx, "b_pre_win_streak"] = bs["win_streak"]
            self.df.at[idx, "r_pre_loss_streak"] = rs["loss_streak"]
            self.df.at[idx, "b_pre_loss_streak"] = bs["loss_streak"]
            self.df.at[idx, "r_pre_title_fights"] = rs["title_fights"]
            self.df.at[idx, "b_pre_title_fights"] = bs["title_fights"]
            self.df.at[idx, "r_pre_title_wins"] = rs["title_wins"]
            self.df.at[idx, "b_pre_title_wins"] = bs["title_wins"]
            self.df.at[idx, "r_pre_avg_fight_time"] = rs["avg_fight_time"]
            self.df.at[idx, "b_pre_avg_fight_time"] = bs["avg_fight_time"]

            sig_str_att_r = max(rs["total_sig_str_att"], 1)
            sig_str_att_b = max(bs["total_sig_str_att"], 1)
            td_att_r = max(rs["total_td_att"], 1)
            td_att_b = max(bs["total_td_att"], 1)
            fight_time_r = max(rs["total_fight_time"], 1.0)
            fight_time_b = max(bs["total_fight_time"], 1.0)

            self.df.at[idx, "r_pre_sig_str_acc"] = rs["total_sig_str"] / sig_str_att_r
            self.df.at[idx, "b_pre_sig_str_acc"] = bs["total_sig_str"] / sig_str_att_b
            self.df.at[idx, "r_pre_td_acc"] = rs["total_td"] / td_att_r
            self.df.at[idx, "b_pre_td_acc"] = bs["total_td"] / td_att_b
            self.df.at[idx, "r_pre_sub_att_rate"] = rs["total_sub_att"] / r_nf
            self.df.at[idx, "b_pre_sub_att_rate"] = bs["total_sub_att"] / b_nf
            self.df.at[idx, "r_pre_kd_rate"] = rs["total_kd"] / r_nf
            self.df.at[idx, "b_pre_kd_rate"] = bs["total_kd"] / b_nf
            self.df.at[idx, "r_pre_ctrl_avg"] = rs["total_ctrl_sec"] / r_nf
            self.df.at[idx, "b_pre_ctrl_avg"] = bs["total_ctrl_sec"] / b_nf

            r_min = fight_time_r / 60.0
            b_min = fight_time_b / 60.0
            self.df.at[idx, "r_pre_SLpM"] = rs["total_sig_str"] / max(r_min, 1.0)
            self.df.at[idx, "b_pre_SLpM"] = bs["total_sig_str"] / max(b_min, 1.0)
            self.df.at[idx, "r_pre_SApM"] = rs["total_str"] / max(r_min, 1.0)
            self.df.at[idx, "b_pre_SApM"] = bs["total_str"] / max(b_min, 1.0)
            self.df.at[idx, "r_pre_td_avg"] = rs["total_td"] / max(r_min, 1.0)
            self.df.at[idx, "b_pre_td_avg"] = bs["total_td"] / max(b_min, 1.0)

            # Rolling 3
            hist_r3 = rs["_history"][-3:]
            hist_b3 = bs["_history"][-3:]
            self.df.at[idx, "r_rolling3_wins"] = sum(h["won"] for h in hist_r3)
            self.df.at[idx, "b_rolling3_wins"] = sum(h["won"] for h in hist_b3)
            self.df.at[idx, "r_rolling3_sig_str"] = np.mean([h["sig_str"] for h in hist_r3]) if hist_r3 else 0.0
            self.df.at[idx, "b_rolling3_sig_str"] = np.mean([h["sig_str"] for h in hist_b3]) if hist_b3 else 0.0
            self.df.at[idx, "r_rolling3_td"] = np.mean([h["td"] for h in hist_r3]) if hist_r3 else 0.0
            self.df.at[idx, "b_rolling3_td"] = np.mean([h["td"] for h in hist_b3]) if hist_b3 else 0.0
            self.df.at[idx, "r_rolling3_kd"] = np.mean([h["kd"] for h in hist_r3]) if hist_r3 else 0.0
            self.df.at[idx, "b_rolling3_kd"] = np.mean([h["kd"] for h in hist_b3]) if hist_b3 else 0.0
            self.df.at[idx, "r_rolling3_sub_att"] = np.mean([h["sub_att"] for h in hist_r3]) if hist_r3 else 0.0
            self.df.at[idx, "b_rolling3_sub_att"] = np.mean([h["sub_att"] for h in hist_b3]) if hist_b3 else 0.0

            # Rolling 5
            hist_r5 = rs["_history"][-5:]
            hist_b5 = bs["_history"][-5:]
            self.df.at[idx, "r_rolling5_wins"] = sum(h["won"] for h in hist_r5)
            self.df.at[idx, "b_rolling5_wins"] = sum(h["won"] for h in hist_b5)
            self.df.at[idx, "r_rolling5_sig_str"] = np.mean([h["sig_str"] for h in hist_r5]) if hist_r5 else 0.0
            self.df.at[idx, "b_rolling5_sig_str"] = np.mean([h["sig_str"] for h in hist_b5]) if hist_b5 else 0.0
            self.df.at[idx, "r_rolling5_td"] = np.mean([h["td"] for h in hist_r5]) if hist_r5 else 0.0
            self.df.at[idx, "b_rolling5_td"] = np.mean([h["td"] for h in hist_b5]) if hist_b5 else 0.0
            self.df.at[idx, "r_rolling5_kd"] = np.mean([h["kd"] for h in hist_r5]) if hist_r5 else 0.0
            self.df.at[idx, "b_rolling5_kd"] = np.mean([h["kd"] for h in hist_b5]) if hist_b5 else 0.0

            # Exponentially-weighted rolling stats (last 5 fights, positional decay)
            # Gives more weight to recent fights than simple unweighted means.
            for _corner, _hist in (("r", rs["_history"]), ("b", bs["_history"])):
                _h5 = _hist[-5:]
                self.df.at[idx, f"{_corner}_ewm_wins"]    = _ewm_hist(_h5, "won")
                self.df.at[idx, f"{_corner}_ewm_sig_str"] = _ewm_hist(_h5, "sig_str")
                self.df.at[idx, f"{_corner}_ewm_td"]      = _ewm_hist(_h5, "td")
                self.df.at[idx, f"{_corner}_ewm_kd"]      = _ewm_hist(_h5, "kd")
                self.df.at[idx, f"{_corner}_ewm_sub_att"] = _ewm_hist(_h5, "sub_att")

            # Glicko-2 pre-fight snapshot (BEFORE update — prevents data leakage in Tier 4)
            r_glicko_snap = self.feature_engineer.glicko2_get(r)
            b_glicko_snap = self.feature_engineer.glicko2_get(b)
            self.df.at[idx, "r_glicko_pre_r"]   = r_glicko_snap[0]
            self.df.at[idx, "r_glicko_pre_rd"]  = r_glicko_snap[1]
            self.df.at[idx, "r_glicko_pre_vol"] = r_glicko_snap[2]
            self.df.at[idx, "b_glicko_pre_r"]   = b_glicko_snap[0]
            self.df.at[idx, "b_glicko_pre_rd"]  = b_glicko_snap[1]
            self.df.at[idx, "b_glicko_pre_vol"] = b_glicko_snap[2]

            # Days since last fight
            current_date = row.get("event_date", None)
            for corner, fs in [("r", rs), ("b", bs)]:
                if fs["fight_dates"] and pd.notna(current_date):
                    last_date = fs["fight_dates"][-1]
                    try:
                        days_gap = (current_date - last_date).days
                    except Exception:
                        days_gap = 365
                else:
                    days_gap = 365  # debut default
                self.df.at[idx, f"{corner}_days_since_last"] = days_gap

            # KD absorbed
            self.df.at[idx, "r_pre_kd_absorbed"] = rs.get("kd_absorbed", 0)
            self.df.at[idx, "b_pre_kd_absorbed"] = bs.get("kd_absorbed", 0)

            # Opponent quality: avg ELO of last 5 opponents
            r_opp_elos = rs["opp_elo_history"][-5:]
            b_opp_elos = bs["opp_elo_history"][-5:]
            self.df.at[idx, "r_avg_opp_elo_L5"] = float(np.mean(r_opp_elos)) if r_opp_elos else 1500.0
            self.df.at[idx, "b_avg_opp_elo_L5"] = float(np.mean(b_opp_elos)) if b_opp_elos else 1500.0

            # Trajectory: linear regression slope of last 3 fight results (1=win, 0=loss)
            def _trajectory_slope(history):
                if len(history) < 2:
                    return 0.0
                results = [h["won"] for h in history[-3:]]
                x = np.arange(len(results), dtype=float)
                try:
                    slope = np.polyfit(x, results, 1)[0]
                except Exception:
                    slope = 0.0
                return float(slope)
            self.df.at[idx, "r_trajectory_3"] = _trajectory_slope(rs["_history"])
            self.df.at[idx, "b_trajectory_3"] = _trajectory_slope(bs["_history"])

            # ── Phase 3: Positional striking percentages ──────────────────
            for corner, fs in [("r", rs), ("b", bs)]:
                total_pos = max(fs["total_distance"] + fs["total_clinch"] + fs["total_ground"], 1)
                total_tgt = max(fs["total_head"] + fs["total_body"] + fs["total_leg"], 1)
                self.df.at[idx, f"{corner}_pre_distance_pct"] = fs["total_distance"] / total_pos
                self.df.at[idx, f"{corner}_pre_clinch_pct"]   = fs["total_clinch"]   / total_pos
                self.df.at[idx, f"{corner}_pre_ground_pct"]   = fs["total_ground"]   / total_pos
                self.df.at[idx, f"{corner}_pre_head_pct"]     = fs["total_head"]     / total_tgt
                self.df.at[idx, f"{corner}_pre_body_pct"]     = fs["total_body"]     / total_tgt
                self.df.at[idx, f"{corner}_pre_leg_pct"]      = fs["total_leg"]      / total_tgt
                # Defense rates
                str_def_att = max(fs["str_def_attempts"], 1)
                td_def_att  = max(fs["td_def_attempts"],  1)
                self.df.at[idx, f"{corner}_pre_str_def"] = fs["str_def_success"] / str_def_att
                self.df.at[idx, f"{corner}_pre_td_def"]  = fs["td_def_success"]  / td_def_att
                # Finish timing
                wins = max(fs["wins"], 1)
                self.df.at[idx, f"{corner}_pre_early_finish_rate"]    = fs["early_finishes"]   / wins
                self.df.at[idx, f"{corner}_pre_late_finish_rate"]     = fs["late_finishes"]    / wins
                self.df.at[idx, f"{corner}_pre_first_round_ko_rate"]  = fs["first_round_kos"]  / max(fs["total_fights"], 1)
                self.df.at[idx, f"{corner}_pre_total_rounds_fought"]  = fs["total_rounds_fought"]
                self.df.at[idx, f"{corner}_pre_five_round_fights"]    = fs["five_round_fights"]
                # Rolling10
                hist10 = fs["rolling10_history"][-10:]
                self.df.at[idx, f"{corner}_rolling10_wins"]     = sum(h["won"]     for h in hist10)
                self.df.at[idx, f"{corner}_rolling10_sig_str"]  = float(np.mean([h["sig_str"] for h in hist10])) if hist10 else 0.0
                self.df.at[idx, f"{corner}_rolling10_td"]       = float(np.mean([h["td"]      for h in hist10])) if hist10 else 0.0
                self.df.at[idx, f"{corner}_rolling10_kd"]       = float(np.mean([h["kd"]      for h in hist10])) if hist10 else 0.0
                self.df.at[idx, f"{corner}_rolling10_finishes"] = sum(h.get("finished", 0) for h in hist10)
                # Career arc
                self.df.at[idx, f"{corner}_career_elo_peak"]   = fs["career_elo_peak"]
                self.df.at[idx, f"{corner}_fights_since_peak"]  = fs["fights_since_peak"]
                # Opponent-adjusted win rates
                elite_fights   = max(fs["vs_elite_fights"],    1)
                striker_fights = max(fs["vs_striker_fights"],  1)
                grappler_fights= max(fs["vs_grappler_fights"], 1)
                self.df.at[idx, f"{corner}_vs_elite_win_rate"]    = fs["vs_elite_wins"]    / elite_fights
                self.df.at[idx, f"{corner}_vs_striker_win_rate"]  = fs["vs_striker_wins"]  / striker_fights
                self.df.at[idx, f"{corner}_vs_grappler_win_rate"] = fs["vs_grappler_wins"] / grappler_fights
                # Last fight momentum
                self.df.at[idx, f"{corner}_last_fight_was_win"]    = float(fs["last_fight_was_win"])
                self.df.at[idx, f"{corner}_last_fight_was_finish"] = float(fs["last_fight_was_finish"])
                # Volatility: finish rates and SLpM CV from rolling10_history
                _h10_v  = fs["rolling10_history"]
                _h5_v   = _h10_v[-5:]
                _h10_vl = _h10_v[-10:]
                _fin_l5  = sum(h.get("finished", 0) for h in _h5_v)  / max(len(_h5_v),  1)
                _fin_l10 = sum(h.get("finished", 0) for h in _h10_vl) / max(len(_h10_vl), 1)
                self.df.at[idx, f"{corner}_pre_finish_rate_l5"]  = _fin_l5
                self.df.at[idx, f"{corner}_pre_finish_rate_l10"] = _fin_l10
                _sig_vals = [h.get("sig_str", 0) for h in _h10_vl]
                if len(_sig_vals) >= 3:
                    _mu_v = float(np.mean(_sig_vals))
                    _cv_v = float(np.std(_sig_vals)) / (_mu_v + 0.1)
                else:
                    _cv_v = 0.0
                self.df.at[idx, f"{corner}_pre_slpm_cv"] = _cv_v
                _corner_age = _safe(row.get(f"{corner}_age_at_event", 28.0), 28.0)
                self.df.at[idx, f"{corner}_pre_mileage_adj_age"] = _corner_age * fs["total_rounds_fought"] / 100.0
                # Rolling3 wins (for recent_form_ratio_diff)
                self.df.at[idx, f"{corner}_pre_rolling3_wins"] = float(sum(h.get("won", 0) for h in _h10_v[-3:]))
                # Rolling SLpM snapshots (last-5 mean, last-10 std)
                _slpm_l5  = [h.get("slpm", 0.0) for h in _h10_v[-5:]]
                _slpm_l10 = [h.get("slpm", 0.0) for h in _h10_v[-10:]]
                self.df.at[idx, f"{corner}_pre_rolling5_slpm"] = float(np.mean(_slpm_l5))  if _slpm_l5  else 0.0
                self.df.at[idx, f"{corner}_pre_slpm_std_l10"]  = float(np.std(_slpm_l10))  if len(_slpm_l10) >= 2 else 0.0
                # Damage ratio std (last-10)
                _dr_l10 = [h.get("damage_ratio", 1.0) for h in _h10_v[-10:]]
                self.df.at[idx, f"{corner}_pre_damage_ratio_std_l10"] = float(np.std(_dr_l10)) if len(_dr_l10) >= 2 else 0.0
                # Tactical evolution: mean distance_pct last-5 minus career mean distance_pct
                _dp_l5   = [h.get("distance_pct", 0.5) for h in _h10_v[-5:]]
                _dp_all  = [h.get("distance_pct", 0.5) for h in _h10_v]
                _dp_l5m  = float(np.mean(_dp_l5))  if _dp_l5  else 0.5
                _dp_allm = float(np.mean(_dp_all))  if _dp_all else 0.5
                self.df.at[idx, f"{corner}_pre_tactical_evolution"] = _dp_l5m - _dp_allm

            # ── ELO pre-fight ─────────────────────────────────────────────
            r_elo_pre = self.feature_engineer.elo_get(r)
            b_elo_pre = self.feature_engineer.elo_get(b)

            # Update ELO
            _finish_round = int(_safe(row.get("finish_round", 0), 0))
            _r_streak = rs["win_streak"]
            _b_streak = bs["win_streak"]
            _r_elo_pre = self.feature_engineer.elo_get(r)
            _b_elo_pre = self.feature_engineer.elo_get(b)
            self.feature_engineer.elo_update(
                r, b, winner, is_title, method,
                fighter_fights_count[r], fighter_fights_count[b],
                finish_round=_finish_round,
                winner_streak=_r_streak if winner == "Red" else _b_streak,
                opponent_elo=_b_elo_pre if winner == "Red" else _r_elo_pre,
            )
            self.df.at[idx, "r_elo_pre_fight"] = r_elo_pre if "r_elo_pre_fight" in self.df.columns else None
            self.df.at[idx, "b_elo_pre_fight"] = b_elo_pre if "b_elo_pre_fight" in self.df.columns else None

            # ── Glicko-2 update ───────────────────────────────────────────
            r_glicko = self.feature_engineer.glicko2_get(r)
            b_glicko = self.feature_engineer.glicko2_get(b)
            r_score = 1.0 if winner == "Red" else (0.5 if winner == "Draw" else 0.0)
            b_score = 1.0 - r_score
            self.feature_engineer.glicko2_update(r, [(b_glicko[0], b_glicko[1], r_score)])
            self.feature_engineer.glicko2_update(b, [(r_glicko[0], r_glicko[1], b_score)])

            # ── Common opponents ──────────────────────────────────────────
            self.feature_engineer.update_common_opponents(r, b, winner)

            # ── Update stats post-fight ───────────────────────────────────
            r_won = (winner == "Red")
            b_won = (winner == "Blue")
            ft = _safe(row.get("total_fight_time_sec", 0), 0.0)

            def update_fighter(fs, won, fighter_name):
                fs["total_fights"] += 1
                if won:
                    fs["wins"] += 1
                    fs["win_streak"] += 1
                    fs["loss_streak"] = 0
                    if method in ("KO/TKO",):
                        fs["ko_wins"] += 1
                    elif method in ("Submission",):
                        fs["sub_wins"] += 1
                    else:
                        fs["dec_wins"] += 1
                    if is_title:
                        fs["title_wins"] += 1
                elif winner == "Draw":
                    fs["draws"] += 1
                else:
                    fs["losses"] += 1
                    fs["loss_streak"] += 1
                    fs["win_streak"] = 0
                    if method in ("KO/TKO",):
                        fs["ko_losses"] += 1
                    elif method in ("Submission",):
                        fs["sub_losses"] += 1
                    else:
                        fs["dec_losses"] += 1
                if is_title:
                    fs["title_fights"] += 1
                if fs["wins"] > 0:
                    fs["finish_rate"] = (fs["ko_wins"] + fs["sub_wins"]) / fs["wins"]
                fs["total_fight_time"] += ft
                fs["avg_fight_time"] = fs["total_fight_time"] / max(fs["total_fights"], 1)

            def update_fight_stats(fs, prefix, row_data):
                ss = _safe(row_data.get(f"{prefix}_sig_str", 0))
                ss_att = _safe(row_data.get(f"{prefix}_sig_str_att", 0))
                st = _safe(row_data.get(f"{prefix}_str", 0))
                st_att = _safe(row_data.get(f"{prefix}_str_att", 0))
                td = _safe(row_data.get(f"{prefix}_td", 0))
                td_att = _safe(row_data.get(f"{prefix}_td_att", 0))
                sub = _safe(row_data.get(f"{prefix}_sub_att", 0))
                kd = _safe(row_data.get(f"{prefix}_kd", 0))
                ctrl = _safe(row_data.get(f"{prefix}_ctrl_sec", 0))
                rev = _safe(row_data.get(f"{prefix}_rev", 0))
                head = _safe(row_data.get(f"{prefix}_head", 0))
                body = _safe(row_data.get(f"{prefix}_body", 0))
                leg = _safe(row_data.get(f"{prefix}_leg", 0))
                dist = _safe(row_data.get(f"{prefix}_distance", 0))
                clinch = _safe(row_data.get(f"{prefix}_clinch", 0))
                ground = _safe(row_data.get(f"{prefix}_ground", 0))
                fs["total_sig_str"] += ss
                fs["total_sig_str_att"] += ss_att
                fs["total_str"] += st
                fs["total_str_att"] += st_att
                fs["total_td"] += td
                fs["total_td_att"] += td_att
                fs["total_sub_att"] += sub
                fs["total_kd"] += kd
                fs["total_ctrl_sec"] += ctrl
                fs["total_rev"] += rev
                fs["total_head"] += head
                fs["total_body"] += body
                fs["total_leg"] += leg
                fs["total_distance"] += dist
                fs["total_clinch"] += clinch
                fs["total_ground"] += ground
                return {"won": 1 if (prefix == "r" and r_won) or (prefix == "b" and b_won) else 0,
                        "sig_str": ss, "td": td, "kd": kd, "sub_att": sub, "ctrl": ctrl}

            update_fighter(rs, r_won, r)
            update_fighter(bs, b_won, b)
            r_hist = update_fight_stats(rs, "r", row)
            b_hist = update_fight_stats(bs, "b", row)
            rs["_history"].append(r_hist)
            bs["_history"].append(b_hist)
            fighter_fights_count[r] += 1
            fighter_fights_count[b] += 1

            # Track fight dates for ring-rust calculation
            event_date = row.get("event_date", None)
            if pd.notna(event_date):
                rs["fight_dates"].append(event_date)
                bs["fight_dates"].append(event_date)

            # Track opponent ELO for quality-of-opposition feature
            rs["opp_elo_history"].append(b_elo_pre)
            bs["opp_elo_history"].append(r_elo_pre)

            # Track knockdowns absorbed (opponent's kd landing on this fighter)
            r_kd_absorbed = int(_safe(row.get("b_kd", 0), 0))  # blue's KDs landed = red absorbs
            b_kd_absorbed = int(_safe(row.get("r_kd", 0), 0))  # red's KDs landed = blue absorbs
            rs["kd_absorbed"] += r_kd_absorbed
            bs["kd_absorbed"] += b_kd_absorbed

            # ── Phase 3: Update extended tracking fields ───────────────────
            _finish_rnd = int(_safe(row.get("finish_round", 0), 0))
            _total_rds  = int(_safe(row.get("total_rounds", 3), 3))

            for corner, fs, opp_prefix, won in [("r", rs, "b", r_won), ("b", bs, "r", b_won)]:
                # Positional striking (already accumulated by update_fight_stats, just need head/body/leg/dist/clinch/ground)
                # (those are already tracked in total_head/body/leg/distance/clinch/ground above)

                # Defense: opponent's sig_str att and success
                opp_ss_att = int(_safe(row.get(f"{opp_prefix}_sig_str_att", 0), 0))
                opp_ss_lnd = int(_safe(row.get(f"{opp_prefix}_sig_str",     0), 0))
                opp_td_att = int(_safe(row.get(f"{opp_prefix}_td_att",      0), 0))
                opp_td_lnd = int(_safe(row.get(f"{opp_prefix}_td",          0), 0))
                fs["str_def_attempts"] += opp_ss_att
                fs["str_def_success"]  += (opp_ss_att - opp_ss_lnd)
                fs["td_def_attempts"]  += opp_td_att
                fs["td_def_success"]   += (opp_td_att - opp_td_lnd)

                # Finish timing
                if won and method in ("KO/TKO", "Submission"):
                    if _finish_rnd <= 2:
                        fs["early_finishes"] += 1
                    else:
                        fs["late_finishes"]  += 1
                    if _finish_rnd == 1 and method == "KO/TKO":
                        fs["first_round_kos"] += 1

                # Rounds tracking
                fs["total_rounds_fought"] += _finish_rnd if _finish_rnd > 0 else _total_rds
                if _total_rds >= 5:
                    fs["five_round_fights"] += 1

                # Rolling10 history (append finished flag)
                _h10_ss  = _safe(row.get(f"{corner}_sig_str", 0))
                _h10_ft  = max(ft / 60.0, 1.0 / 60.0)
                _h10_opp = max(opp_ss_lnd, 1)
                _h10_d   = _safe(row.get(f"{corner}_distance", 0))
                _h10_c   = _safe(row.get(f"{corner}_clinch", 0))
                _h10_g   = _safe(row.get(f"{corner}_ground", 0))
                h10 = {
                    "won": 1 if won else 0,
                    "sig_str": _h10_ss,
                    "td": _safe(row.get(f"{corner}_td", 0)),
                    "kd": _safe(row.get(f"{corner}_kd", 0)),
                    "sub_att": _safe(row.get(f"{corner}_sub_att", 0)),
                    "finished": 1 if (won and method in ("KO/TKO", "Submission")) else 0,
                    "slpm": _h10_ss / _h10_ft,
                    "damage_ratio": _h10_ss / _h10_opp,
                    "distance_pct": _h10_d / max(_h10_d + _h10_c + _h10_g, 1.0),
                }
                fs["rolling10_history"].append(h10)

                # Last fight details
                fs["last_fight_was_win"]    = bool(won)
                fs["last_fight_was_finish"] = bool(won and method in ("KO/TKO", "Submission"))
                fs["last_fight_method"]     = method
                fs["last_fight_finish_round"] = _finish_rnd

                # Career ELO peak (use current ELO after update)
                cur_elo = self.feature_engineer.elo_get(r if corner == "r" else b)
                if cur_elo > fs["career_elo_peak"]:
                    fs["career_elo_peak"]    = cur_elo
                    fs["fights_since_peak"]  = 0
                else:
                    fs["fights_since_peak"] += 1

                # Opponent-adjusted win rates
                opp_elo_pre = b_elo_pre if corner == "r" else r_elo_pre
                opp_slpm    = _safe(row.get(f"{opp_prefix}_pre_SLpM", 0))
                opp_td_avg  = _safe(row.get(f"{opp_prefix}_pre_td_avg", 0))
                if opp_elo_pre > 1600:
                    fs["vs_elite_fights"] += 1
                    if won:
                        fs["vs_elite_wins"] += 1
                if opp_slpm > 4.5:
                    fs["vs_striker_fights"] += 1
                    if won:
                        fs["vs_striker_wins"] += 1
                if opp_td_avg > 2.5:
                    fs["vs_grappler_fights"] += 1
                    if won:
                        fs["vs_grappler_wins"] += 1

        # Add ELO columns if missing
        if "r_elo_pre_fight" not in self.df.columns:
            self.df["r_elo_pre_fight"] = 1500.0
        if "b_elo_pre_fight" not in self.df.columns:
            self.df["b_elo_pre_fight"] = 1500.0

        # Recompute ELO cleanly with full K-factor modifiers
        self.feature_engineer.elo_ratings = {}
        fight_count_r = defaultdict(int)
        fight_count_b = defaultdict(int)
        win_streak_r = defaultdict(int)
        win_streak_b = defaultdict(int)
        for idx, row in self.df.iterrows():
            r = str(row.get("r_fighter", "")).strip()
            b = str(row.get("b_fighter", "")).strip()
            winner = str(row.get("winner", "")).strip()
            method = str(row.get("method", "")).strip()
            is_title = bool(row.get("is_title_bout", False))
            finish_round = int(float(row.get("finish_round", 0) or 0))
            r_elo_cur = self.feature_engineer.elo_get(r)
            b_elo_cur = self.feature_engineer.elo_get(b)
            pre_rA, pre_rB = self.feature_engineer.elo_update(
                r, b, winner, is_title, method,
                fight_count_r[r], fight_count_b[b],
                finish_round=finish_round,
                winner_streak=win_streak_r[r] if winner == "Red" else win_streak_b[b],
                opponent_elo=b_elo_cur if winner == "Red" else r_elo_cur,
            )
            self.df.at[idx, "r_elo_pre_fight"] = pre_rA
            self.df.at[idx, "b_elo_pre_fight"] = pre_rB
            fight_count_r[r] += 1
            fight_count_b[b] += 1
            if winner == "Red":
                win_streak_r[r] += 1
                win_streak_b[b] = 0
            elif winner == "Blue":
                win_streak_b[b] += 1
                win_streak_r[r] = 0
            else:
                win_streak_r[r] = 0
                win_streak_b[b] = 0

        self._log(f"Data leakage fix complete. DataFrame shape: {self.df.shape}")

    # ── BUILD FEATURES ───────────────────────────────────────────────────────
    def build_all_features(self):
        print_section("BUILDING FEATURES (13+ TIERS)")
        df = self.df

        # ── DROP LEAKY COLUMNS ────────────────────────────────────────────
        # These columns are either in-fight statistics (only known after the
        # fight ends) or post-fight career totals.  fix_data_leakage() has
        # already consumed them to build the safe r_pre_* / b_pre_* columns.
        # Leaving them in causes ~92% lab accuracy that collapses to ~54% on
        # real upcoming-fight cards.
        _LEAKY_COLS = [
            # Per-fight striking / grappling stats (outcome of the fight itself)
            "r_sig_str", "b_sig_str", "sig_str_diff",
            "r_sig_str_att", "b_sig_str_att", "sig_str_att_diff",
            "r_sig_str_acc", "b_sig_str_acc", "sig_str_acc_diff",
            "r_str", "b_str", "str_diff",
            "r_str_att", "b_str_att", "str_att_diff",
            "r_str_acc", "b_str_acc", "str_acc_diff",
            "r_kd", "b_kd", "kd_diff",
            "r_head", "b_head", "head_diff",
            "r_body", "b_body", "body_diff",
            "r_leg", "b_leg", "leg_diff",
            "r_distance", "b_distance", "distance_diff",
            "r_clinch", "b_clinch", "clinch_diff",
            "r_ground", "b_ground", "ground_diff",
            "r_td", "b_td", "td_diff",
            "r_td_att", "b_td_att", "td_att_diff",
            "r_td_acc", "b_td_acc", "td_acc_diff",
            "r_sub_att", "b_sub_att", "sub_att_diff",
            "r_rev", "b_rev", "rev_diff",
            "r_ctrl_sec", "b_ctrl_sec", "ctrl_sec_diff",
            "total_fight_time_sec",
            # Post-fight win/loss totals (pre-fight versions are r_pre_wins etc.)
            "r_wins", "b_wins", "wins_diff",
            "r_losses", "b_losses", "losses_diff",
            "r_draws", "b_draws", "draws_diff",
            "r_win_loss_ratio", "b_win_loss_ratio", "win_loss_ratio_diff",
            # UFC published career averages (may include current fight;
            # pre-fight versions are r_pre_SLpM, r_pre_sig_str_acc, etc.)
            "r_pro_SLpM", "b_pro_SLpM", "pro_SLpM_diff",
            "r_pro_sig_str_acc", "b_pro_sig_str_acc", "pro_sig_str_acc_diff",
            "r_pro_SApM", "b_pro_SApM", "pro_SApM_diff",
            "r_pro_str_def", "b_pro_str_def", "pro_str_def_diff",
            "r_pro_td_avg", "b_pro_td_avg", "pro_td_avg_diff",
            "r_pro_td_acc", "b_pro_td_acc", "pro_td_acc_diff",
            "r_pro_td_def", "b_pro_td_def", "pro_td_def_diff",
            "r_pro_sub_avg", "b_pro_sub_avg", "pro_sub_avg_diff",
        ]
        _leaky_present = [c for c in _LEAKY_COLS if c in df.columns]
        self.df = df = df.drop(columns=_leaky_present)
        print_step(f"Dropped {len(_leaky_present)} leaky in-fight columns.")

        # ── TIER 0: Raw column diffs (safe physical attributes only) ──────
        self._log("Tier 0: Raw column differences...")
        raw_pairs = [
            # Physical / biographical — available before the fight
            ("r_height", "b_height"), ("r_reach", "b_reach"),
            ("r_weight", "b_weight"),
            ("r_age_at_event", "b_age_at_event"),
            ("r_ape_index", "b_ape_index"),
        ]
        for rc, bc in raw_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = pd.to_numeric(df[rc], errors="coerce").fillna(0) - \
                                        pd.to_numeric(df[bc], errors="coerce").fillna(0)

        # ── TIER 1: Pre-fight stat diffs ──────────────────────────────────
        self._log("Tier 1: Pre-fight stat diffs...")
        pre_pairs = [
            ("r_pre_wins", "b_pre_wins"), ("r_pre_losses", "b_pre_losses"),
            ("r_pre_ko_wins", "b_pre_ko_wins"), ("r_pre_sub_wins", "b_pre_sub_wins"),
            ("r_pre_dec_wins", "b_pre_dec_wins"),
            ("r_pre_total_fights", "b_pre_total_fights"),
            ("r_pre_finish_rate", "b_pre_finish_rate"),
            ("r_pre_win_streak", "b_pre_win_streak"),
            ("r_pre_loss_streak", "b_pre_loss_streak"),
            ("r_pre_title_fights", "b_pre_title_fights"),
            ("r_pre_title_wins", "b_pre_title_wins"),
            ("r_pre_avg_fight_time", "b_pre_avg_fight_time"),
            ("r_pre_sig_str_acc", "b_pre_sig_str_acc"),
            ("r_pre_td_acc", "b_pre_td_acc"),
            ("r_pre_sub_att_rate", "b_pre_sub_att_rate"),
            ("r_pre_kd_rate", "b_pre_kd_rate"),
            ("r_pre_ctrl_avg", "b_pre_ctrl_avg"),
            ("r_pre_SLpM", "b_pre_SLpM"),
            ("r_pre_SApM", "b_pre_SApM"),
            ("r_pre_td_avg", "b_pre_td_avg"),
        ]
        for rc, bc in pre_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[6:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        # ── TIER 2: Rolling window diffs ──────────────────────────────────
        self._log("Tier 2: Rolling window diffs...")
        rolling_pairs = [
            ("r_rolling3_wins", "b_rolling3_wins"),
            ("r_rolling3_sig_str", "b_rolling3_sig_str"),
            ("r_rolling3_td", "b_rolling3_td"),
            ("r_rolling3_kd", "b_rolling3_kd"),
            ("r_rolling3_sub_att", "b_rolling3_sub_att"),
            ("r_rolling5_wins", "b_rolling5_wins"),
            ("r_rolling5_sig_str", "b_rolling5_sig_str"),
            ("r_rolling5_td", "b_rolling5_td"),
            ("r_rolling5_kd", "b_rolling5_kd"),
        ]
        for rc, bc in rolling_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        # ── TIER 3: ELO features ──────────────────────────────────────────
        self._log("Tier 3: ELO features...")
        if "r_elo_pre_fight" in df.columns and "b_elo_pre_fight" in df.columns:
            df["elo_diff"] = df["r_elo_pre_fight"] - df["b_elo_pre_fight"]
            df["elo_r"] = df["r_elo_pre_fight"]
            df["elo_b"] = df["b_elo_pre_fight"]
            df["elo_ratio"] = df["r_elo_pre_fight"] / (df["b_elo_pre_fight"] + 1e-6)
        else:
            df["elo_diff"] = 0.0
            df["elo_r"] = 1500.0
            df["elo_b"] = 1500.0
            df["elo_ratio"] = 1.0

        # ── TIER 4: Glicko-2 features ─────────────────────────────────────
        # Use pre-fight snapshots stored in fix_data_leakage() — no data leakage
        self._log("Tier 4: Glicko-2 features (pre-fight snapshots)...")
        if "r_glicko_pre_r" in df.columns and "b_glicko_pre_r" in df.columns:
            df["r_glicko_r"]   = df["r_glicko_pre_r"]
            df["r_glicko_rd"]  = df["r_glicko_pre_rd"]
            df["r_glicko_vol"] = df["r_glicko_pre_vol"]
            df["b_glicko_r"]   = df["b_glicko_pre_r"]
            df["b_glicko_rd"]  = df["b_glicko_pre_rd"]
            df["b_glicko_vol"] = df["b_glicko_pre_vol"]
        else:
            # Fallback: read final career values (slight leakage but graceful)
            r_glicko_r, r_glicko_rd, r_glicko_vol = [], [], []
            b_glicko_r, b_glicko_rd, b_glicko_vol = [], [], []
            for _, row in df.iterrows():
                rg = self.feature_engineer.glicko2_get(str(row.get("r_fighter", "")))
                bg = self.feature_engineer.glicko2_get(str(row.get("b_fighter", "")))
                r_glicko_r.append(rg[0])
                r_glicko_rd.append(rg[1])
                r_glicko_vol.append(rg[2])
                b_glicko_r.append(bg[0])
                b_glicko_rd.append(bg[1])
                b_glicko_vol.append(bg[2])
            df["r_glicko_r"]   = r_glicko_r
            df["r_glicko_rd"]  = r_glicko_rd
            df["r_glicko_vol"] = r_glicko_vol
            df["b_glicko_r"]   = b_glicko_r
            df["b_glicko_rd"]  = b_glicko_rd
            df["b_glicko_vol"] = b_glicko_vol
        df["glicko_diff"]    = df["r_glicko_r"] - df["b_glicko_r"]
        df["glicko_rd_diff"] = df["r_glicko_rd"] - df["b_glicko_rd"]

        # ── TIER 5: Weight-class Z-scores ─────────────────────────────────
        self._log("Tier 5: Weight-class Z-scores...")
        fe = self.feature_engineer
        z_feats = ["r_pre_SLpM", "r_pre_SApM", "r_pre_sig_str_acc",
                   "r_pre_td_avg", "r_pre_sub_att_rate", "r_pre_kd_rate",
                   "b_pre_SLpM", "b_pre_SApM", "b_pre_sig_str_acc",
                   "b_pre_td_avg", "b_pre_sub_att_rate", "b_pre_kd_rate"]

        # First pass: accumulate stats for Z-score computation
        for _, row in df.iterrows():
            wc = str(row.get("weight_class", ""))
            yr = row["event_date"].year if pd.notna(row.get("event_date")) else 2000
            stats_dict = {}
            for feat in z_feats:
                if feat in df.columns:
                    v = row.get(feat, 0)
                    try:
                        v = float(v)
                        if not math.isnan(v):
                            stats_dict[feat] = v
                    except (TypeError, ValueError):
                        pass
            fe.update_weight_class_stats(wc, yr, stats_dict)

        # Second pass: compute Z-scores
        for col in z_feats:
            df[f"z_{col}"] = 0.0
        for idx, row in df.iterrows():
            wc = str(row.get("weight_class", ""))
            yr = row["event_date"].year if pd.notna(row.get("event_date")) else 2000
            for feat in z_feats:
                if feat in df.columns:
                    v = row.get(feat, 0)
                    try:
                        v = float(v)
                        if not math.isnan(v):
                            df.at[idx, f"z_{feat}"] = fe.get_z_score(wc, yr, feat, v)
                    except (TypeError, ValueError):
                        pass

        # ── TIER 6: Common opponent features ──────────────────────────────
        self._log("Tier 6: Common opponent features...")
        n_common, r_wins_c, b_wins_c, co_edge = [], [], [], []
        for _, row in df.iterrows():
            r = str(row.get("r_fighter", ""))
            b = str(row.get("b_fighter", ""))
            feat = fe.get_common_opponent_features(r, b)
            n_common.append(feat["n_common_opponents"])
            r_wins_c.append(feat["r_wins_vs_common"])
            b_wins_c.append(feat["b_wins_vs_common"])
            co_edge.append(feat["common_opp_edge"])
        df["n_common_opponents"] = n_common
        df["r_wins_vs_common"] = r_wins_c
        df["b_wins_vs_common"] = b_wins_c
        df["common_opp_edge"] = co_edge

        # ── TIER 7: Style cluster features ────────────────────────────────
        self._log("Tier 7: Style cluster features...")
        # Restrict KMeans cluster fitting to the training portion of the data
        # (chronological first 90%) so test-set fighters don't influence cluster
        # centroids. Test-set fighters who never appeared in training will still
        # receive a cluster assignment via the fitted KMeans model; they just
        # don't shift the centroids during fitting.
        _n_train_proxy = int(len(df) * 0.9)
        _train_idx_set = set(
            df.sort_values("event_date", na_position="first").index[:_n_train_proxy]
        )
        fighter_style = {}
        for f in self.all_fighters:
            r_rows = df[(df["r_fighter"] == f) & (df.index.isin(_train_idx_set))]
            b_rows = df[(df["b_fighter"] == f) & (df.index.isin(_train_idx_set))]
            slpm = 0.0; sapm = 0.0; td = 0.0; sub = 0.0; finish = 0.0
            cnt = 0
            for _, row in r_rows.iterrows():
                slpm += float(row.get("r_pre_SLpM", 0) or 0)
                sapm += float(row.get("r_pre_SApM", 0) or 0)
                td += float(row.get("r_pre_td_avg", 0) or 0)
                sub += float(row.get("r_pre_sub_att_rate", 0) or 0)
                finish += float(row.get("r_pre_finish_rate", 0) or 0)
                cnt += 1
            for _, row in b_rows.iterrows():
                slpm += float(row.get("b_pre_SLpM", 0) or 0)
                sapm += float(row.get("b_pre_SApM", 0) or 0)
                td += float(row.get("b_pre_td_avg", 0) or 0)
                sub += float(row.get("b_pre_sub_att_rate", 0) or 0)
                finish += float(row.get("b_pre_finish_rate", 0) or 0)
                cnt += 1
            if cnt > 0:
                fighter_style[f] = {"SLpM": slpm/cnt, "SApM": sapm/cnt,
                                     "TD": td/cnt, "Sub": sub/cnt, "Finish": finish/cnt}
        fe.fit_clusters(fighter_style)

        # Populate style-cluster features using pre-fight chronological snapshots.
        # Single sorted pass: snapshot win-rates BEFORE updating with the outcome
        # so each row only sees data from fights that already happened.
        # The previous two-pass approach (replay all → assign features using final
        # cumulative rates) leaked future outcomes into every row's features.
        _style_snap = {}  # df index → (rc, bc, r_winrate, b_winrate, edge)
        for idx, row in df.sort_values("event_date", na_position="first").iterrows():
            r = str(row.get("r_fighter", ""))
            b = str(row.get("b_fighter", ""))
            winner = str(row.get("winner", ""))
            rc = fe.get_fighter_cluster(r)
            bc = fe.get_fighter_cluster(b)
            if rc >= 0 and bc >= 0:
                # Snapshot PRE-fight win rates before this outcome is recorded
                mf = fe.get_style_matchup_features(rc, bc)
                _style_snap[idx] = (
                    rc, bc,
                    mf["r_style_win_vs_opp_cluster"],
                    mf["b_style_win_vs_opp_cluster"],
                    mf["style_matchup_edge"],
                )
                # NOW update with this fight's outcome
                fe.update_style_performance(rc, bc, winner == "Red")
                fe.update_style_performance(bc, rc, winner == "Blue")
            else:
                _style_snap[idx] = (rc, bc, 0.5, 0.5, 0.0)

        r_cluster, b_cluster, style_edge = [], [], []
        r_style_win, b_style_win = [], []
        for idx, row in df.iterrows():
            snap = _style_snap.get(idx, (-1, -1, 0.5, 0.5, 0.0))
            rc, bc, rw, bw, edge = snap
            r_cluster.append(rc)
            b_cluster.append(bc)
            r_style_win.append(rw)
            b_style_win.append(bw)
            style_edge.append(edge)
        df["r_style_cluster"] = r_cluster
        df["b_style_cluster"] = b_cluster
        df["style_matchup_edge"] = style_edge
        df["r_style_win_vs_cluster"] = r_style_win
        df["b_style_win_vs_cluster"] = b_style_win

        # ── TIER 8: Stance encoding ───────────────────────────────────────
        self._log("Tier 8: Stance encoding...")
        stance_map = {"Orthodox": 0, "Southpaw": 1, "Switch": 2, "Open Stance": 3}
        for col in ["r_stance", "b_stance"]:
            if col in df.columns:
                df[f"{col}_enc"] = df[col].map(stance_map).fillna(-1)
        if "r_stance_enc" in df.columns and "b_stance_enc" in df.columns:
            df["stance_matchup"] = df["r_stance_enc"].astype(str) + "_" + df["b_stance_enc"].astype(str)
            df["same_stance"] = (df["r_stance_enc"] == df["b_stance_enc"]).astype(int)

        # ── TIER 9: Interaction features ─────────────────────────────────
        self._log("Tier 9: Interaction features...")
        if "elo_diff" in df.columns and "diff_finish_rate" in df.columns:
            df["elo_x_finish_rate"] = df["elo_diff"] * df["diff_finish_rate"]
        if "diff_pre_win_streak" in df.columns and "diff_pre_finish_rate" in df.columns:
            df["streak_x_finish"] = df["diff_pre_win_streak"] * df["diff_pre_finish_rate"]
        if "diff_pre_SLpM" in df.columns and "diff_pre_SApM" in df.columns:
            df["striking_exchange"] = df["diff_pre_SLpM"] - df["diff_pre_SApM"]
        if "diff_pre_td_avg" in df.columns and "diff_pre_td_acc" in df.columns:
            df["td_efficiency"] = df["diff_pre_td_avg"] * df["diff_pre_td_acc"]
        if "diff_pre_sig_str_acc" in df.columns and "diff_pre_ctrl_avg" in df.columns:
            df["control_accuracy"] = df["diff_pre_sig_str_acc"] * df["diff_pre_ctrl_avg"]

        # ── TIER 10: Polynomial features ─────────────────────────────────
        self._log("Tier 10: Polynomial features...")
        poly_cols = ["elo_diff", "glicko_diff", "diff_win_loss_ratio"]
        for col in poly_cols:
            if col in df.columns:
                df[f"{col}_sq"] = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()

        # ── TIER 11: Momentum indicators ─────────────────────────────────
        self._log("Tier 11: Momentum indicators...")
        if "r_rolling3_wins" in df.columns and "b_rolling3_wins" in df.columns:
            df["momentum_diff_3"] = df["r_rolling3_wins"] - df["b_rolling3_wins"]
        if "r_rolling5_wins" in df.columns and "b_rolling5_wins" in df.columns:
            df["momentum_diff_5"] = df["r_rolling5_wins"] - df["b_rolling5_wins"]
        if "r_pre_win_streak" in df.columns and "b_pre_win_streak" in df.columns:
            df["streak_differential"] = df["r_pre_win_streak"] - df["b_pre_win_streak"]

        # ── TIER 12: Method-specific features ────────────────────────────
        self._log("Tier 12: Method-specific features...")
        if "r_pre_ko_wins" in df.columns and "b_pre_ko_wins" in df.columns:
            df["ko_threat_diff"] = df["r_pre_ko_wins"] - df["b_pre_ko_wins"]
        if "r_pre_sub_wins" in df.columns and "b_pre_sub_wins" in df.columns:
            df["sub_threat_diff"] = df["r_pre_sub_wins"] - df["b_pre_sub_wins"]
        if "r_pre_dec_wins" in df.columns and "b_pre_dec_wins" in df.columns:
            df["dec_tendency_diff"] = df["r_pre_dec_wins"] - df["b_pre_dec_wins"]
        if "r_pre_finish_rate" in df.columns and "b_pre_finish_rate" in df.columns:
            df["r_finishing_tendency"] = df["r_pre_finish_rate"]
            df["b_finishing_tendency"] = df["b_pre_finish_rate"]
            df["finishing_matchup"] = df["r_pre_finish_rate"] * df["b_pre_finish_rate"]

        # ── CAREER PATTERNS ──────────────────────────────────────────────────
        print_step("Career pattern features...")
        for prefix in ['r', 'b']:
            # Decision-specific win rates
            dec_col = f'{prefix}_pre_dec_wins'
            ko_col = f'{prefix}_pre_ko_wins'
            sub_col = f'{prefix}_pre_sub_wins'
            total_col = f'{prefix}_pre_total_fights'

            if all(c in df.columns for c in [dec_col, ko_col, sub_col, total_col]):
                denom = df[total_col].clip(lower=1)
                df[f'{prefix}_decision_win_rate'] = df[dec_col] / denom
                df[f'{prefix}_ko_win_rate'] = df[ko_col] / denom
                df[f'{prefix}_sub_win_rate'] = df[sub_col] / denom
                df[f'{prefix}_finish_rate'] = (df[ko_col] + df[sub_col]) / denom

            # Title fight experience
            if f'{prefix}_pre_title_fights' in df.columns:
                df[f'{prefix}_title_fight_exp'] = df[f'{prefix}_pre_title_fights']

            # Main event experience (5-round fights)
            if f'{prefix}_pre_five_round_fights' in df.columns:
                df[f'{prefix}_main_event_exp'] = df[f'{prefix}_pre_five_round_fights']

        # Diffs
        for feat in ['decision_win_rate', 'ko_win_rate', 'sub_win_rate', 'finish_rate', 'title_fight_exp', 'main_event_exp']:
            r_col = f'r_{feat}'
            b_col = f'b_{feat}'
            if r_col in df.columns and b_col in df.columns:
                df[f'{feat}_diff'] = df[r_col] - df[b_col]

        # ── TIER 12b: Advanced combat metrics ────────────────────────────
        self._log("Tier 12b: Advanced combat metrics...")

        # ── RING RUST FACTOR ─────────────────────────────────────────────────
        for prefix in ['r', 'b']:
            col = f'{prefix}_days_since_last'
            if col in df.columns:
                days = df[col].fillna(365)
                df[f'{prefix}_ring_rust'] = np.where(days > 365, -0.15,
                                            np.where(days > 180, -0.08,
                                            np.where(days > 90, -0.03, 0.0)))
            else:
                df[f'{prefix}_ring_rust'] = 0.0
        if 'r_ring_rust' in df.columns and 'b_ring_rust' in df.columns:
            df['ring_rust_diff'] = df['r_ring_rust'] - df['b_ring_rust']

        # ── WEIGHT CLASS ADAPTATION FACTOR ──────────────────────────────────
        wc_map = {
            'Heavyweight': 1.4, 'Light Heavyweight': 1.2, 'Middleweight': 1.0,
            'Welterweight': 0.95, 'Lightweight': 0.9, 'Featherweight': 0.85,
            'Bantamweight': 0.8, 'Flyweight': 0.75,
            "Women's Bantamweight": 0.8, "Women's Flyweight": 0.75,
            "Women's Strawweight": 0.7, "Women's Featherweight": 0.72,
        }
        df['weight_class_ko_factor'] = df['weight_class'].map(wc_map).fillna(1.0)

        # ── STYLE CLASH SEVERITY ─────────────────────────────────────────────
        r_slpm = df.get('r_pre_SLpM', df.get('r_pro_SLpM', pd.Series(0, index=df.index))).fillna(0)
        b_slpm = df.get('b_pre_SLpM', df.get('b_pro_SLpM', pd.Series(0, index=df.index))).fillna(0)
        r_td   = df.get('r_pre_td_avg', df.get('r_pro_td_avg', pd.Series(0, index=df.index))).fillna(0)
        b_td   = df.get('b_pre_td_avg', df.get('b_pro_td_avg', pd.Series(0, index=df.index))).fillna(0)
        df['style_clash_severity'] = np.abs((r_slpm - b_slpm) * 0.5 + (r_td - b_td) * 0.3)

        # ── UPSET POTENTIAL ──────────────────────────────────────────────────
        r_form   = df.get('r_recent_form_3',   pd.Series(0.5, index=df.index)).fillna(0.5)
        b_form   = df.get('b_recent_form_3',   pd.Series(0.5, index=df.index)).fillna(0.5)
        r_streak = df.get('r_win_streak',       pd.Series(0,   index=df.index)).fillna(0)
        b_streak = df.get('b_win_streak',       pd.Series(0,   index=df.index)).fillna(0)
        r_age    = df.get('r_age_at_event',     pd.Series(28,  index=df.index)).fillna(28)
        b_age    = df.get('b_age_at_event',     pd.Series(28,  index=df.index)).fillna(28)
        r_fights = df.get('r_pre_total_fights', pd.Series(10,  index=df.index)).fillna(10)
        b_fights = df.get('b_pre_total_fights', pd.Series(10,  index=df.index)).fillna(10)
        df['upset_potential'] = ((b_form - r_form) * 0.4 +
                                  (b_streak - r_streak) * 0.1 +
                                  (b_age - r_age) * 0.01 +
                                  (b_fights - r_fights) * 0.005)

        # ── POWER VS TECHNIQUE ────────────────────────────────────────────────
        r_acc = df.get('r_pre_sig_str_acc', pd.Series(0.45, index=df.index)).fillna(0.45)
        b_acc = df.get('b_pre_sig_str_acc', pd.Series(0.45, index=df.index)).fillna(0.45)
        df['power_vs_technique'] = (r_slpm - b_slpm) * 0.6 + (r_acc - b_acc) * 0.4

        # ── CHAMPIONSHIP PRESSURE IMPACT ─────────────────────────────────────
        is_title       = df.get('is_title_bout',        pd.Series(0, index=df.index)).fillna(0)
        title_exp_diff = df.get('title_fight_exp_diff', pd.Series(0, index=df.index)).fillna(0)
        df['championship_pressure'] = title_exp_diff * is_title * (r_form - b_form)

        # ── CLINCH EFFECTIVENESS ─────────────────────────────────────────────
        r_clinch_pct = df.get('r_pre_clinch_pct', pd.Series(0, index=df.index)).fillna(0)
        b_clinch_pct = df.get('b_pre_clinch_pct', pd.Series(0, index=df.index)).fillna(0)
        df['r_clinch_effectiveness'] = r_clinch_pct * r_slpm * r_acc
        df['b_clinch_effectiveness'] = b_clinch_pct * b_slpm * b_acc
        df['clinch_effectiveness_diff'] = df['r_clinch_effectiveness'] - df['b_clinch_effectiveness']

        # ── FIVE-ROUND CARDIO ADVANTAGE ──────────────────────────────────────
        r_dec            = df.get('r_decision_win_rate', pd.Series(0, index=df.index)).fillna(0)
        b_dec            = df.get('b_decision_win_rate', pd.Series(0, index=df.index)).fillna(0)
        total_rounds_col = df.get('total_rounds',        pd.Series(3, index=df.index)).fillna(3)
        df['five_round_cardio_advantage'] = (r_dec - b_dec) * (total_rounds_col / 3)

        # ── CHIN DETERIORATION INDEX ─────────────────────────────────────────
        r_kd_absorbed  = df.get('r_pre_kd_absorbed', pd.Series(0, index=df.index)).fillna(0)
        b_kd_absorbed  = df.get('b_pre_kd_absorbed', pd.Series(0, index=df.index)).fillna(0)
        r_fights_safe  = r_fights.clip(lower=1)
        b_fights_safe  = b_fights.clip(lower=1)
        df['r_chin_deterioration'] = r_kd_absorbed / r_fights_safe
        df['b_chin_deterioration'] = b_kd_absorbed / b_fights_safe
        df['chin_deterioration_diff'] = df['r_chin_deterioration'] - df['b_chin_deterioration']

        # ── FINISHING PRESSURE (WEIGHTED COMBO) ──────────────────────────────
        r_ko_r  = df.get('r_ko_win_rate',             pd.Series(0, index=df.index)).fillna(0)
        b_ko_r  = df.get('b_ko_win_rate',             pd.Series(0, index=df.index)).fillna(0)
        r_sub_r = df.get('r_sub_win_rate',            pd.Series(0, index=df.index)).fillna(0)
        b_sub_r = df.get('b_sub_win_rate',            pd.Series(0, index=df.index)).fillna(0)
        r_fin_r = df.get('r_recent_finish_rate_3',    pd.Series(0, index=df.index)).fillna(0)
        b_fin_r = df.get('b_recent_finish_rate_3',    pd.Series(0, index=df.index)).fillna(0)
        df['finishing_pressure_diff'] = ((r_ko_r - b_ko_r) * 0.5 +
                                          (r_sub_r - b_sub_r) * 0.3 +
                                          (r_fin_r - b_fin_r) * 0.2)

        # ── OVERACTIVITY FLAG (< 60 DAYS BETWEEN FIGHTS) ─────────────────────
        r_days = df.get('r_days_since_last', pd.Series(180, index=df.index)).fillna(180)
        b_days = df.get('b_days_since_last', pd.Series(180, index=df.index)).fillna(180)
        df['r_overactive'] = (r_days < 60).astype(float)
        df['b_overactive'] = (b_days < 60).astype(float)
        df['overactivity_diff'] = df['r_overactive'] - df['b_overactive']

        # ── TIER 12c: Opponent quality & trajectory ───────────────────────
        self._log("Tier 12c: Opponent quality & trajectory slopes...")
        if "r_avg_opp_elo_L5" in df.columns and "b_avg_opp_elo_L5" in df.columns:
            df["opp_quality_diff"] = df["r_avg_opp_elo_L5"] - df["b_avg_opp_elo_L5"]
        if "r_trajectory_3" in df.columns and "b_trajectory_3" in df.columns:
            df["trajectory_diff"] = df["r_trajectory_3"] - df["b_trajectory_3"]

        # ── TIER 12d: Fighter-at-peak score ──────────────────────────────
        self._log("Tier 12d: Fighter-at-peak scoring...")
        for prefix in ["r", "b"]:
            age_col   = f"{prefix}_age_at_event"
            fights_col = f"{prefix}_pre_total_fights"
            form3_col  = f"{prefix}_rolling3_wins"
            age_s   = df.get(age_col,   pd.Series(28, index=df.index)).fillna(28)
            fights_s = df.get(fights_col, pd.Series(10, index=df.index)).fillna(10)
            form3_s  = df.get(form3_col,  pd.Series(1.5, index=df.index)).fillna(1.5)
            # Peak age: 27-32 scores highest, taper outside
            age_peak = 1.0 - np.abs(age_s - 29.5) / 10.0
            age_peak = age_peak.clip(0, 1)
            # Experience: sweet spot 10-25 fights
            exp_score = (fights_s.clip(1, 25) / 25.0)
            # Form: fraction of last 3 wins
            form_score = (form3_s / 3.0).clip(0, 1)
            df[f"{prefix}_peak_score"] = (age_peak * 0.4 + exp_score * 0.3 + form_score * 0.3)
        if "r_peak_score" in df.columns and "b_peak_score" in df.columns:
            df["peak_score_diff"] = df["r_peak_score"] - df["b_peak_score"]

        # ── TIER 12e: Uncertainty score ───────────────────────────────────
        self._log("Tier 12e: Uncertainty scoring...")
        # Multi-factor 0-1 uncertainty estimate for each fight
        r_slpm_u  = df.get("r_pre_SLpM",  pd.Series(3.0, index=df.index)).fillna(3.0)
        b_slpm_u  = df.get("b_pre_SLpM",  pd.Series(3.0, index=df.index)).fillna(3.0)
        r_td_u    = df.get("r_pre_td_avg", pd.Series(1.5, index=df.index)).fillna(1.5)
        b_td_u    = df.get("b_pre_td_avg", pd.Series(1.5, index=df.index)).fillna(1.5)
        r_nf_u   = df.get("r_pre_total_fights", pd.Series(5, index=df.index)).fillna(5)
        b_nf_u   = df.get("b_pre_total_fights", pd.Series(5, index=df.index)).fillna(5)
        elo_d_u  = df.get("elo_diff", pd.Series(0, index=df.index)).fillna(0).abs()
        # Striking similarity: 1 = very similar (high uncertainty), 0 = very different
        str_sim  = 1.0 - (np.abs(r_slpm_u - b_slpm_u) / (r_slpm_u + b_slpm_u + 1e-6)).clip(0, 1)
        # Grappling similarity
        gr_sim   = 1.0 - (np.abs(r_td_u - b_td_u) / (r_td_u + b_td_u + 1e-6)).clip(0, 1)
        # Experience factor: low experience -> high uncertainty
        exp_unc  = 1.0 - ((r_nf_u + b_nf_u) / 60.0).clip(0, 1)
        # ELO closeness: similar ELO -> high uncertainty
        elo_unc  = 1.0 - (elo_d_u / 400.0).clip(0, 1)
        df["uncertainty_score"] = (str_sim * 0.25 + gr_sim * 0.25 + exp_unc * 0.25 + elo_unc * 0.25)

        # ── TIER 13: SVD decomposition ────────────────────────────────────
        self._log("Tier 13: SVD decomposition per feature bucket...")
        striking_cols = [c for c in df.columns if any(x in c for x in
                         ["SLpM", "SApM", "sig_str", "str_def", "kd", "head", "body", "leg"])
                         and c in df.columns]
        grappling_cols = [c for c in df.columns if any(x in c for x in
                          ["td", "sub", "ctrl", "rev", "ground"])
                          and c in df.columns]
        physical_cols = [c for c in df.columns if any(x in c for x in
                         ["height", "reach", "weight", "age", "ape_index"])
                         and c in df.columns]
        form_cols = [c for c in df.columns if any(x in c for x in
                     ["rolling", "streak", "momentum", "finish_rate", "elo", "glicko"])
                     and c in df.columns]

        def apply_svd(cols, svd_obj, prefix, df, col_store_attr):
            if len(cols) < 2:
                return
            X = df[cols].fillna(0).values
            n_comp = min(svd_obj.n_components, X.shape[1] - 1, X.shape[0] - 1)
            if n_comp < 1:
                return
            svd_obj.n_components = n_comp
            try:
                X_svd = svd_obj.fit_transform(X)
                setattr(self, col_store_attr, cols)  # save exact fitted columns
                for i in range(n_comp):
                    df[f"{prefix}_svd_{i}"] = X_svd[:, i]
            except Exception:
                pass

        apply_svd(striking_cols,  self.svd_striking,  "striking",  df, "svd_striking_cols")
        apply_svd(grappling_cols, self.svd_grappling, "grappling", df, "svd_grappling_cols")
        apply_svd(physical_cols,  self.svd_physical,  "physical",  df, "svd_physical_cols")
        apply_svd(form_cols,      self.svd_form,      "form",      df, "svd_form_cols")
        self.svd_fitted = True

        # ── Encode title bout and is_title_bout ───────────────────────────
        if "is_title_bout" in df.columns:
            df["is_title_enc"] = df["is_title_bout"].astype(int)
        if "total_rounds" in df.columns:
            df["total_rounds_num"] = pd.to_numeric(df["total_rounds"], errors="coerce").fillna(3)

        # ── Gender encoding ───────────────────────────────────────────────
        if "gender" in df.columns:
            df["gender_enc"] = (df["gender"].fillna("").str.lower() == "women").astype(int)

        # ── TIER 14: Positional & Target Differentials ────────────────────
        self._log("Tier 14: Positional & target differentials...")
        for feat in ["distance_pct", "clinch_pct", "ground_pct", "head_pct", "body_pct", "leg_pct"]:
            r_col, b_col = f"r_pre_{feat}", f"b_pre_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"diff_{feat}"] = df[r_col].fillna(0) - df[b_col].fillna(0)
        df["positional_striking_advantage"] = (df.get("diff_distance_pct", pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
                                                df.get("diff_clinch_pct",  pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
                                                df.get("diff_ground_pct",  pd.Series(0.0, index=df.index)).fillna(0.0).abs())
        df["target_distribution_advantage"] = (df.get("diff_head_pct", pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
                                                df.get("diff_body_pct", pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
                                                df.get("diff_leg_pct",  pd.Series(0.0, index=df.index)).fillna(0.0).abs())

        # ── TIER 15: Defense Differentials ───────────────────────────────
        self._log("Tier 15: Defense differentials...")
        df["diff_str_def"] = df.get("r_pre_str_def", pd.Series(0, index=df.index)).fillna(0) - \
                             df.get("b_pre_str_def", pd.Series(0, index=df.index)).fillna(0)
        df["diff_td_def"]  = df.get("r_pre_td_def",  pd.Series(0, index=df.index)).fillna(0) - \
                             df.get("b_pre_td_def",  pd.Series(0, index=df.index)).fillna(0)
        df["defensive_composite"] = df["diff_str_def"].fillna(0) + df["diff_td_def"].fillna(0)

        # ── TIER 16: Deep Interaction Features ───────────────────────────
        self._log("Tier 16: Deep interaction features...")
        elo_d    = df.get("elo_diff",           pd.Series(0, index=df.index)).fillna(0)
        form3    = df.get("diff_rolling3_wins", pd.Series(0, index=df.index)).fillna(0)
        wlr      = df.get("diff_win_loss_ratio",pd.Series(0, index=df.index)).fillna(0)
        fin_r    = df.get("diff_finish_rate",   pd.Series(0, index=df.index)).fillna(0)
        kd_abs   = df.get("diff_kd_absorbed",   pd.Series(0, index=df.index)).fillna(0)

        df["elo_x_form"]       = elo_d * form3
        df["elo_x_win_ratio"]  = elo_d * wlr
        df["elo_x_finish"]     = elo_d * fin_r
        df["elo_x_durability"] = elo_d * kd_abs.abs()

        reach_d  = df.get("diff_reach",         pd.Series(0, index=df.index)).fillna(0)
        height_d = df.get("diff_height",        pd.Series(0, index=df.index)).fillna(0)
        slpm_d   = df.get("diff_pre_SLpM",      pd.Series(0, index=df.index)).fillna(0)
        td_d     = df.get("diff_pre_td_avg",    pd.Series(0, index=df.index)).fillna(0)
        acc_d    = df.get("diff_pre_sig_str_acc",pd.Series(0, index=df.index)).fillna(0)
        age_d    = df.get("diff_age_at_event",  pd.Series(0, index=df.index)).fillna(0)
        streak_d = df.get("diff_pre_win_streak",pd.Series(0, index=df.index)).fillna(0)
        exp_gap  = df.get("diff_pre_total_fights",pd.Series(0, index=df.index)).fillna(0)

        df["reach_x_striking"]    = reach_d * slpm_d
        df["height_x_reach"]      = height_d * reach_d
        df["physical_x_striking"] = (height_d + reach_d) * slpm_d

        df["age_x_striking"]   = age_d * slpm_d
        df["age_x_grappling"]  = age_d * td_d
        df["age_x_durability"] = age_d * kd_abs.abs()
        df["age_x_win_streak"] = age_d * streak_d
        df["experience_x_age"] = exp_gap * age_d

        str_def_d  = df.get("diff_str_def",      pd.Series(0, index=df.index)).fillna(0)
        td_def_d   = df.get("diff_td_def",        pd.Series(0, index=df.index)).fillna(0)
        sub_rate_d = df.get("sub_threat_diff",    pd.Series(0, index=df.index)).fillna(0)
        df["td_x_defense"]           = td_d * td_def_d
        df["submission_x_grappling"] = sub_rate_d * td_d

        df["striking_x_accuracy"] = slpm_d * acc_d
        df["striking_x_defense"]  = slpm_d * str_def_d
        df["ko_power_x_striking"] = df.get("ko_threat_diff", pd.Series(0, index=df.index)).fillna(0) * slpm_d

        momentum = df.get("momentum_diff_3", pd.Series(0, index=df.index)).fillna(0)
        df["momentum_x_win_streak"] = momentum * streak_d
        df["form_x_experience"]     = form3 * exp_gap
        df["finish_x_momentum"]     = fin_r * momentum
        df["form_x_durability"]     = form3 * kd_abs.abs()

        df["elite_finisher"]     = elo_d * fin_r * form3
        df["unstoppable_streak"] = streak_d * momentum * form3
        df["veteran_advantage"]  = wlr * exp_gap * (-age_d)

        # ── TIER 17: Extended Polynomial Features ─────────────────────────
        self._log("Tier 17: Extended polynomial features...")
        poly_extended = [
            "elo_diff", "glicko_diff", "diff_win_loss_ratio", "diff_age_at_event",
            "diff_reach", "diff_height", "diff_pre_SLpM", "diff_pre_sig_str_acc",
            "diff_pre_td_avg", "diff_pre_win_streak", "diff_finish_rate",
            "diff_pre_loss_streak", "diff_str_def", "diff_td_def",
            "diff_pre_kd_rate", "diff_pre_ctrl_avg",
            "elo_x_form", "streak_x_finish", "striking_exchange",
            "diff_distance_pct", "diff_clinch_pct", "diff_ground_pct",
        ]
        for col in poly_extended:
            if col in df.columns:
                df[f"{col}_sq"]  = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()
        if "diff_age_at_event" in df.columns:
            df["diff_age_cubed"] = df["diff_age_at_event"] ** 3

        # ── TIER 18: Opponent-Adjusted Performance ────────────────────────
        self._log("Tier 18: Opponent-adjusted performance...")
        if "r_vs_elite_win_rate" in df.columns and "b_vs_elite_win_rate" in df.columns:
            df["diff_win_rate_vs_elite"]     = df["r_vs_elite_win_rate"]    - df["b_vs_elite_win_rate"]
            df["diff_win_rate_vs_strikers"]  = df.get("r_vs_striker_win_rate",  pd.Series(0, index=df.index)).fillna(0) - \
                                               df.get("b_vs_striker_win_rate",  pd.Series(0, index=df.index)).fillna(0)
            df["diff_win_rate_vs_grapplers"] = df.get("r_vs_grappler_win_rate", pd.Series(0, index=df.index)).fillna(0) - \
                                               df.get("b_vs_grappler_win_rate", pd.Series(0, index=df.index)).fillna(0)
            df["championship_readiness"]     = df["diff_win_rate_vs_elite"] * df.get("elo_diff", pd.Series(0.0, index=df.index)).fillna(0.0)

        # ── TIER 19: Career Pattern Features ─────────────────────────────
        self._log("Tier 19: Career pattern features...")
        if "r_pre_early_finish_rate" in df.columns:
            df["diff_early_finish_rate"]    = df.get("r_pre_early_finish_rate", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_early_finish_rate", pd.Series(0.0, index=df.index)).fillna(0)
            df["diff_late_finish_rate"]     = df.get("r_pre_late_finish_rate",    pd.Series(0, index=df.index)).fillna(0) - \
                                              df.get("b_pre_late_finish_rate",    pd.Series(0, index=df.index)).fillna(0)
            df["diff_first_round_ko_rate"]  = df.get("r_pre_first_round_ko_rate", pd.Series(0, index=df.index)).fillna(0) - \
                                              df.get("b_pre_first_round_ko_rate", pd.Series(0, index=df.index)).fillna(0)
        if "r_pre_five_round_fights" in df.columns:
            df["diff_five_round_fights"] = df.get("r_pre_five_round_fights", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_five_round_fights", pd.Series(0.0, index=df.index)).fillna(0)

        age_r_t19 = df.get("r_age_at_event", pd.Series(28, index=df.index)).fillna(28)
        age_b_t19 = df.get("b_age_at_event", pd.Series(28, index=df.index)).fillna(28)
        df["r_prime_score"] = (1.0 - np.abs(age_r_t19 - 29.5) / 10.0).clip(0, 1)
        df["b_prime_score"] = (1.0 - np.abs(age_b_t19 - 29.5) / 10.0).clip(0, 1)
        df["prime_years_advantage"] = df["r_prime_score"] - df["b_prime_score"]

        if "r_fights_since_peak" in df.columns:
            df["diff_fights_since_peak"] = df.get("r_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0)
            df["declining_phase_diff"]   = df["diff_fights_since_peak"]

        if "r_last_fight_was_win" in df.columns:
            df["r_last_fight_momentum"] = df.get("r_last_fight_was_win", pd.Series(0.0, index=df.index)).fillna(0).astype(float) + df.get("r_last_fight_was_finish", pd.Series(0.0, index=df.index)).fillna(0).astype(float)
            df["b_last_fight_momentum"] = df.get("b_last_fight_was_win", pd.Series(0.0, index=df.index)).fillna(0).astype(float) + df.get("b_last_fight_was_finish", pd.Series(0.0, index=df.index)).fillna(0).astype(float)
            df["last_fight_momentum_diff"] = df["r_last_fight_momentum"] - df["b_last_fight_momentum"]

        for feat in ["wins", "sig_str", "td", "kd", "finishes"]:
            r_col10 = f"r_rolling10_{feat}"
            b_col10 = f"b_rolling10_{feat}"
            if r_col10 in df.columns and b_col10 in df.columns:
                df[f"diff_rolling10_{feat}"] = df[r_col10].fillna(0) - df[b_col10].fillna(0)

        # ── TIER 20: Rounds-Based Strategy ───────────────────────────────
        self._log("Tier 20: Rounds-based strategy features...")
        total_rds_t20 = df.get("total_rounds_num", pd.Series(3, index=df.index)).fillna(3)
        dec_rate_d_t20 = df.get("dec_tendency_diff", pd.Series(0, index=df.index)).fillna(0)
        if "diff_finish_rate" in df.columns:
            df["rounds_x_cardio"]      = total_rds_t20 * dec_rate_d_t20
            df["rounds_x_finish_rate"] = (5 - total_rds_t20) * df["diff_finish_rate"].fillna(0)
        kd_abs2_t20 = df.get("chin_deterioration_diff", pd.Series(0, index=df.index)).fillna(0)
        df["rounds_x_durability"] = total_rds_t20 * kd_abs2_t20

        # ── TIER 21: Matchup-Specific Features ───────────────────────────
        self._log("Tier 21: Matchup-specific features...")
        r_slpm_t21  = df.get("r_pre_SLpM",         pd.Series(3.0,  index=df.index)).fillna(3.0)
        b_slpm_t21  = df.get("b_pre_SLpM",         pd.Series(3.0,  index=df.index)).fillna(3.0)
        r_acc_t21   = df.get("r_pre_sig_str_acc",   pd.Series(0.45, index=df.index)).fillna(0.45)
        b_acc_t21   = df.get("b_pre_sig_str_acc",   pd.Series(0.45, index=df.index)).fillna(0.45)
        r_str_def_t21 = df.get("r_pre_str_def",     pd.Series(0.55, index=df.index)).fillna(0.55)
        b_str_def_t21 = df.get("b_pre_str_def",     pd.Series(0.55, index=df.index)).fillna(0.55)
        r_td_t21    = df.get("r_pre_td_avg",        pd.Series(1.5,  index=df.index)).fillna(1.5)
        b_td_t21    = df.get("b_pre_td_avg",        pd.Series(1.5,  index=df.index)).fillna(1.5)
        r_td_def_t21= df.get("r_pre_td_def",        pd.Series(0.65, index=df.index)).fillna(0.65)
        b_td_def_t21= df.get("b_pre_td_def",        pd.Series(0.65, index=df.index)).fillna(0.65)
        r_sub_avg_t21 = df.get("r_pre_sub_att_rate", pd.Series(0.3,  index=df.index)).fillna(0.3)
        b_sub_avg_t21 = df.get("b_pre_sub_att_rate", pd.Series(0.3,  index=df.index)).fillna(0.3)
        r_td_acc_t21= df.get("r_pre_td_acc",        pd.Series(0.4,  index=df.index)).fillna(0.4)
        b_td_acc_t21= df.get("b_pre_td_acc",        pd.Series(0.4,  index=df.index)).fillna(0.4)
        r_ctrl_t21  = df.get("r_pre_ctrl_avg",      pd.Series(60,   index=df.index)).fillna(60)
        b_ctrl_t21  = df.get("b_pre_ctrl_avg",      pd.Series(60,   index=df.index)).fillna(60)
        r_sub_rate_t21 = df.get("r_pre_sub_att_rate",pd.Series(0.3, index=df.index)).fillna(0.3)
        b_sub_rate_t21 = df.get("b_pre_sub_att_rate",pd.Series(0.3, index=df.index)).fillna(0.3)

        df["r_striking_vs_b_defense"]  = r_slpm_t21 * (1.0 - b_str_def_t21)
        df["b_striking_vs_r_defense"]  = b_slpm_t21 * (1.0 - r_str_def_t21)
        df["striking_exploitation_diff"] = df["r_striking_vs_b_defense"] - df["b_striking_vs_r_defense"]

        df["r_td_vs_b_td_defense"]   = r_td_t21 * (1.0 - b_td_def_t21)
        df["b_td_vs_r_td_defense"]   = b_td_t21 * (1.0 - r_td_def_t21)
        df["td_exploitation_diff"]    = df["r_td_vs_b_td_defense"] - df["b_td_vs_r_td_defense"]

        df["r_sub_setup_efficiency"]     = r_sub_rate_t21 * r_td_acc_t21
        df["b_sub_setup_efficiency"]     = b_sub_rate_t21 * b_td_acc_t21
        df["sub_setup_diff"]             = df["r_sub_setup_efficiency"] - df["b_sub_setup_efficiency"]
        df["r_sub_threat_vs_td_defense"] = r_sub_avg_t21 * (1.0 - b_td_def_t21)
        df["b_sub_threat_vs_td_defense"] = b_sub_avg_t21 * (1.0 - r_td_def_t21)
        df["sub_threat_vs_defense_diff"] = df["r_sub_threat_vs_td_defense"] - df["b_sub_threat_vs_td_defense"]

        df["r_striking_quality"]    = r_slpm_t21 * r_acc_t21
        df["b_striking_quality"]    = b_slpm_t21 * b_acc_t21
        df["striking_quality_diff"] = df["r_striking_quality"] - df["b_striking_quality"]
        df["r_accuracy_under_fire"] = r_acc_t21 / (b_slpm_t21 + 0.1)
        df["b_accuracy_under_fire"] = b_acc_t21 / (r_slpm_t21 + 0.1)
        df["accuracy_under_fire_diff"] = df["r_accuracy_under_fire"] - df["b_accuracy_under_fire"]

        # ── TIER 22: Statistical Ratio Features ──────────────────────────
        self._log("Tier 22: Statistical ratio features...")
        r_sapm_t22 = df.get("r_pre_SApM", pd.Series(3.0, index=df.index)).fillna(3.0)
        b_sapm_t22 = df.get("b_pre_SApM", pd.Series(3.0, index=df.index)).fillna(3.0)

        df["r_damage_ratio"]    = r_slpm_t21 / (r_sapm_t22 + 0.1)
        df["b_damage_ratio"]    = b_slpm_t21 / (b_sapm_t22 + 0.1)
        df["damage_ratio_diff"] = df["r_damage_ratio"] - df["b_damage_ratio"]

        df["r_striking_output_quality"] = r_slpm_t21 * r_acc_t21 / (r_sapm_t22 + 0.1)
        df["b_striking_output_quality"] = b_slpm_t21 * b_acc_t21 / (b_sapm_t22 + 0.1)
        df["striking_output_quality_diff"] = df["r_striking_output_quality"] - df["b_striking_output_quality"]

        df["r_grappling_quality"] = r_td_t21 * r_td_acc_t21 * (r_ctrl_t21 / 60.0)
        df["b_grappling_quality"] = b_td_t21 * b_td_acc_t21 * (b_ctrl_t21 / 60.0)
        df["grappling_quality_diff"] = df["r_grappling_quality"] - df["b_grappling_quality"]

        df["r_total_defense_index"] = r_str_def_t21 * r_td_def_t21
        df["b_total_defense_index"] = b_str_def_t21 * b_td_def_t21
        df["total_defense_diff"]    = df["r_total_defense_index"] - df["b_total_defense_index"]

        df["r_complete_fighter_index"] = (r_slpm_t21 + r_td_t21 + r_sub_avg_t21) * r_str_def_t21 * r_td_def_t21
        df["b_complete_fighter_index"] = (b_slpm_t21 + b_td_t21 + b_sub_avg_t21) * b_str_def_t21 * b_td_def_t21
        df["complete_fighter_diff"]    = df["r_complete_fighter_index"] - df["b_complete_fighter_index"]

        df["r_pressure_index"] = r_slpm_t21 * r_td_t21 * (r_ctrl_t21 / 60.0)
        df["b_pressure_index"] = b_slpm_t21 * b_td_t21 * (b_ctrl_t21 / 60.0)
        df["pressure_index_diff"] = df["r_pressure_index"] - df["b_pressure_index"]

        # ── TIER 23: Extended Statistical Ratio Features ──────────────────
        self._log("Tier 23: Extended statistical ratio features...")
        r_wins_t23   = df.get("r_pre_wins",        pd.Series(5.0,   index=df.index)).fillna(5.0)
        b_wins_t23   = df.get("b_pre_wins",        pd.Series(5.0,   index=df.index)).fillna(5.0)
        r_losses_t23 = df.get("r_pre_losses",      pd.Series(2.0,   index=df.index)).fillna(2.0)
        b_losses_t23 = df.get("b_pre_losses",      pd.Series(2.0,   index=df.index)).fillna(2.0)
        r_streak_t23 = df.get("r_pre_win_streak",  pd.Series(0.0,   index=df.index)).fillna(0.0)
        b_streak_t23 = df.get("b_pre_win_streak",  pd.Series(0.0,   index=df.index)).fillna(0.0)
        r_fr_t23     = df.get("r_pre_finish_rate", pd.Series(0.4,   index=df.index)).fillna(0.4)
        b_fr_t23     = df.get("b_pre_finish_rate", pd.Series(0.4,   index=df.index)).fillna(0.4)
        r_age_t23    = df.get("r_age_at_event",    pd.Series(28.0,  index=df.index)).fillna(28.0)
        b_age_t23    = df.get("b_age_at_event",    pd.Series(28.0,  index=df.index)).fillna(28.0)
        r_reach_t23  = df.get("r_reach",           pd.Series(71.0,  index=df.index)).fillna(71.0)
        b_reach_t23  = df.get("b_reach",           pd.Series(71.0,  index=df.index)).fillna(71.0)
        r_weight_t23 = df.get("r_weight",          pd.Series(155.0, index=df.index)).fillna(155.0)
        b_weight_t23 = df.get("b_weight",          pd.Series(155.0, index=df.index)).fillna(155.0)
        r_tf_t23     = df.get("r_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(10.0)
        b_tf_t23     = df.get("b_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(10.0)
        # Efficiency ratios
        df["r_defense_offense_balance"]    = (r_str_def_t21 + 0.01) / (r_acc_t21 + 0.01)
        df["b_defense_offense_balance"]    = (b_str_def_t21 + 0.01) / (b_acc_t21 + 0.01)
        df["defense_offense_balance_diff"] = df["r_defense_offense_balance"] - df["b_defense_offense_balance"]
        df["r_td_defense_offense_balance"]    = (r_td_def_t21 + 0.01) / (r_td_acc_t21 + 0.01)
        df["b_td_defense_offense_balance"]    = (b_td_def_t21 + 0.01) / (b_td_acc_t21 + 0.01)
        df["td_defense_offense_balance_diff"] = df["r_td_defense_offense_balance"] - df["b_td_defense_offense_balance"]
        df["finish_efficiency_diff"] = r_fr_t23 - b_fr_t23
        # Quality-over-quantity
        df["r_precision_striking"] = r_acc_t21 / (r_slpm_t21 + 0.1)
        df["b_precision_striking"] = b_acc_t21 / (b_slpm_t21 + 0.1)
        df["precision_striking_diff"] = df["r_precision_striking"] - df["b_precision_striking"]
        df["r_quality_grappling_23"] = r_td_acc_t21 * (r_td_t21 ** 0.5)
        df["b_quality_grappling_23"] = b_td_acc_t21 * (b_td_t21 ** 0.5)
        df["quality_grappling_diff"] = df["r_quality_grappling_23"] - df["b_quality_grappling_23"]
        df["r_submission_threat_ratio"] = (r_sub_avg_t21 + 0.01) / (r_td_t21 + 0.01)
        df["b_submission_threat_ratio"] = (b_sub_avg_t21 + 0.01) / (b_td_t21 + 0.01)
        df["submission_threat_ratio_diff"] = df["r_submission_threat_ratio"] - df["b_submission_threat_ratio"]
        # Defensive efficiency
        df["r_damage_absorption_efficiency"] = r_sapm_t22 / (r_str_def_t21 + 0.01)
        df["b_damage_absorption_efficiency"] = b_sapm_t22 / (b_str_def_t21 + 0.01)
        df["damage_absorption_efficiency_diff"] = df["r_damage_absorption_efficiency"] - df["b_damage_absorption_efficiency"]
        df["r_defense_versatility"] = (r_str_def_t21 * r_td_def_t21) ** 0.5
        df["b_defense_versatility"] = (b_str_def_t21 * b_td_def_t21) ** 0.5
        df["defense_versatility_diff"] = df["r_defense_versatility"] - df["b_defense_versatility"]
        # Offensive versatility
        df["r_total_offense_index"] = r_slpm_t21 + (r_td_t21 * 1.5)
        df["b_total_offense_index"] = b_slpm_t21 + (b_td_t21 * 1.5)
        df["total_offense_index_diff"] = df["r_total_offense_index"] - df["b_total_offense_index"]
        df["r_offensive_versatility"] = (r_slpm_t21 * r_td_t21) ** 0.5
        df["b_offensive_versatility"] = (b_slpm_t21 * b_td_t21) ** 0.5
        df["offensive_versatility_diff"] = df["r_offensive_versatility"] - df["b_offensive_versatility"]
        df["r_striker_index"] = (r_slpm_t21 + 0.1) / (r_td_t21 + 0.1)
        df["b_striker_index"] = (b_slpm_t21 + 0.1) / (b_td_t21 + 0.1)
        df["striker_index_diff"] = df["r_striker_index"] - df["b_striker_index"]
        # Win quality
        r_wlr_t23 = r_wins_t23 / (r_losses_t23 + 1.0)
        b_wlr_t23 = b_wins_t23 / (b_losses_t23 + 1.0)
        df["win_loss_ratio_squared_diff"] = (r_wlr_t23 ** 2) - (b_wlr_t23 ** 2)
        df["r_experience_quality"] = r_wins_t23 / (r_wins_t23 + r_losses_t23 + 1.0)
        df["b_experience_quality"] = b_wins_t23 / (b_wins_t23 + b_losses_t23 + 1.0)
        df["experience_quality_diff"] = df["r_experience_quality"] - df["b_experience_quality"]
        df["r_win_efficiency"] = r_wins_t23 / (r_age_t23 - 18.0 + 1.0)
        df["b_win_efficiency"] = b_wins_t23 / (b_age_t23 - 18.0 + 1.0)
        df["win_efficiency_diff"] = df["r_win_efficiency"] - df["b_win_efficiency"]
        df["r_momentum_quality"] = (r_streak_t23 + 1.0) / (r_wins_t23 + 1.0)
        df["b_momentum_quality"] = (b_streak_t23 + 1.0) / (b_wins_t23 + 1.0)
        df["momentum_quality_diff"] = df["r_momentum_quality"] - df["b_momentum_quality"]
        # Physical efficiency
        df["r_reach_efficiency"] = r_slpm_t21 / (r_reach_t23 + 1.0)
        df["b_reach_efficiency"] = b_slpm_t21 / (b_reach_t23 + 1.0)
        df["reach_efficiency_diff"] = df["r_reach_efficiency"] - df["b_reach_efficiency"]
        df["r_size_adjusted_striking"] = r_slpm_t21 / ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_striking"] = b_slpm_t21 / ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_striking_diff"] = df["r_size_adjusted_striking"] - df["b_size_adjusted_striking"]
        df["r_size_adjusted_grappling"] = r_td_t21 / ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_grappling"] = b_td_t21 / ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_grappling_diff"] = df["r_size_adjusted_grappling"] - df["b_size_adjusted_grappling"]
        # Advanced composites
        df["r_counter_fighter_index"] = (r_str_def_t21 + 0.1) / (r_slpm_t21 + 1.0)
        df["b_counter_fighter_index"] = (b_str_def_t21 + 0.1) / (b_slpm_t21 + 1.0)
        df["counter_fighter_index_diff"] = df["r_counter_fighter_index"] - df["b_counter_fighter_index"]
        df["r_finishing_threat_composite"] = (r_fr_t23 + 0.1) * (r_sub_avg_t21 + 0.1)
        df["b_finishing_threat_composite"] = (b_fr_t23 + 0.1) * (b_sub_avg_t21 + 0.1)
        df["finishing_threat_composite_diff"] = df["r_finishing_threat_composite"] - df["b_finishing_threat_composite"]
        df["r_complete_geo"] = ((r_slpm_t21 + 1.0) * (r_str_def_t21 + 0.1) * (r_fr_t23 + 0.1)) ** (1.0 / 3.0)
        df["b_complete_geo"] = ((b_slpm_t21 + 1.0) * (b_str_def_t21 + 0.1) * (b_fr_t23 + 0.1)) ** (1.0 / 3.0)
        df["complete_geo_diff"] = df["r_complete_geo"] - df["b_complete_geo"]
        df["r_pressure_fighter_index"] = (r_slpm_t21 + r_td_t21) / (r_str_def_t21 + 0.3)
        df["b_pressure_fighter_index"] = (b_slpm_t21 + b_td_t21) / (b_str_def_t21 + 0.3)
        df["pressure_fighter_index_diff"] = df["r_pressure_fighter_index"] - df["b_pressure_fighter_index"]
        # recent_form_ratio: recent win rate vs career win rate
        _r_roll3 = df.get("r_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(1.5)
        _b_roll3 = df.get("b_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(1.5)
        _r_cwr = r_wins_t23 / (r_tf_t23 + 1.0)
        _b_cwr = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_recent_form_ratio"] = (_r_roll3 / 3.0 + 0.01) / (_r_cwr + 0.01)
        df["b_recent_form_ratio"] = (_b_roll3 / 3.0 + 0.01) / (_b_cwr + 0.01)
        df["recent_form_ratio_diff"] = df["r_recent_form_ratio"] - df["b_recent_form_ratio"]
        # finish_method_diversity: how many distinct finish methods used
        _r_ko_d  = df.get("r_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_ko_d  = df.get("b_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        _r_sub_d = df.get("r_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_sub_d = df.get("b_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        _r_dec_d = df.get("r_pre_dec_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_dec_d = df.get("b_pre_dec_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["r_finish_method_diversity"] = (_r_ko_d > 0).astype(float) + (_r_sub_d > 0).astype(float) + (_r_dec_d > 0).astype(float)
        df["b_finish_method_diversity"] = (_b_ko_d > 0).astype(float) + (_b_sub_d > 0).astype(float) + (_b_dec_d > 0).astype(float)
        df["finish_method_diversity_diff"] = df["r_finish_method_diversity"] - df["b_finish_method_diversity"]
        # cross_domain_compensation: grappling compensates striking gap
        df["r_cross_domain_compensation"] = np.maximum(0.0, r_td_t21 - 1.5) - np.maximum(0.0, 4.0 - r_slpm_t21)
        df["b_cross_domain_compensation"] = np.maximum(0.0, b_td_t21 - 1.5) - np.maximum(0.0, 4.0 - b_slpm_t21)
        df["cross_domain_compensation_index_diff"] = df["r_cross_domain_compensation"] - df["b_cross_domain_compensation"]

        # ── TIER 24: Additional Matchup-Specific Features ─────────────────
        self._log("Tier 24: Additional matchup-specific features...")
        df["r_absorption_vuln"]               = r_sapm_t22 / (b_slpm_t21 + 0.1)
        df["b_absorption_vuln"]               = b_sapm_t22 / (r_slpm_t21 + 0.1)
        df["absorption_vulnerability_index_diff"] = df["r_absorption_vuln"] - df["b_absorption_vuln"]
        df["r_combined_def_hole"] = (1.0 - r_str_def_t21) * (1.0 - r_td_def_t21)
        df["b_combined_def_hole"] = (1.0 - b_str_def_t21) * (1.0 - b_td_def_t21)
        df["combined_defensive_hole_diff"] = df["r_combined_def_hole"] - df["b_combined_def_hole"]
        df["r_td_pressure_t24"] = (1.0 - r_td_def_t21) * b_td_t21
        df["b_td_pressure_t24"] = (1.0 - b_td_def_t21) * r_td_t21
        df["td_vulnerability_under_pressure_diff"] = df["r_td_pressure_t24"] - df["b_td_pressure_t24"]
        df["r_strike_pressure_t24"] = (1.0 - r_str_def_t21) * b_slpm_t21
        df["b_strike_pressure_t24"] = (1.0 - b_str_def_t21) * r_slpm_t21
        df["strike_defense_under_volume_diff"] = df["r_strike_pressure_t24"] - df["b_strike_pressure_t24"]
        df["r_ctrl_sub_ratio"] = (r_ctrl_t21 / 60.0) / (r_sub_avg_t21 + 0.1)
        df["b_ctrl_sub_ratio"] = (b_ctrl_t21 / 60.0) / (b_sub_avg_t21 + 0.1)
        df["grappling_control_vs_submission_ratio_diff"] = df["r_ctrl_sub_ratio"] - df["b_ctrl_sub_ratio"]
        df["r_sub_def_necessity"] = b_sub_avg_t21 / (r_td_def_t21 + 0.1)
        df["b_sub_def_necessity"] = r_sub_avg_t21 / (b_td_def_t21 + 0.1)
        df["submission_defense_necessity_diff"] = df["r_sub_def_necessity"] - df["b_sub_def_necessity"]
        df["r_strike_synergy"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5
        df["b_strike_synergy"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5
        df["striking_volume_accuracy_synergy_diff"] = df["r_strike_synergy"] - df["b_strike_synergy"]
        df["r_td_paradox"] = (r_td_acc_t21 + 0.01) / (r_td_t21 + 0.5)
        df["b_td_paradox"] = (b_td_acc_t21 + 0.01) / (b_td_t21 + 0.5)
        df["takedown_efficiency_paradox_diff"] = df["r_td_paradox"] - df["b_td_paradox"]
        df["r_total_off_eff"] = ((r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5
                                 + (r_td_t21 * (r_td_acc_t21 + 0.01)) ** 0.5)
        df["b_total_off_eff"] = ((b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5
                                 + (b_td_t21 * (b_td_acc_t21 + 0.01)) ** 0.5)
        df["total_offensive_efficiency_index_diff"] = df["r_total_off_eff"] - df["b_total_off_eff"]
        df["r_sg_corr"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) / (r_td_t21 * (r_td_acc_t21 + 0.01) + 0.1)
        df["b_sg_corr"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) / (b_td_t21 * (b_td_acc_t21 + 0.01) + 0.1)
        df["striking_grappling_efficiency_correlation_diff"] = df["r_sg_corr"] - df["b_sg_corr"]
        df["r_def_allocation_balance"] = (r_str_def_t21 - r_td_def_t21).abs()
        df["b_def_allocation_balance"] = (b_str_def_t21 - b_td_def_t21).abs()
        df["defense_allocation_balance_diff"] = df["r_def_allocation_balance"] - df["b_def_allocation_balance"]
        _r_cbt = ((r_slpm_t21 / 10.0 + 0.01) * (r_acc_t21 + 0.01) * (10.0 / (r_sapm_t22 + 0.01))
                  * (r_str_def_t21 + 0.01) * (r_td_t21 / 5.0 + 0.01) * (r_td_acc_t21 + 0.01)
                  * (r_td_def_t21 + 0.01) * (r_sub_avg_t21 / 2.0 + 0.01)) ** (1.0 / 8.0)
        _b_cbt = ((b_slpm_t21 / 10.0 + 0.01) * (b_acc_t21 + 0.01) * (10.0 / (b_sapm_t22 + 0.01))
                  * (b_str_def_t21 + 0.01) * (b_td_t21 / 5.0 + 0.01) * (b_td_acc_t21 + 0.01)
                  * (b_td_def_t21 + 0.01) * (b_sub_avg_t21 / 2.0 + 0.01)) ** (1.0 / 8.0)
        df["r_combat_eff"] = _r_cbt
        df["b_combat_eff"] = _b_cbt
        df["total_combat_efficiency_index_diff"] = _r_cbt - _b_cbt

        # ── TIER 25: Named Composite Features ─────────────────────────────
        self._log("Tier 25: Named composite features...")
        df["net_striking_advantage"] = (r_slpm_t21 - b_slpm_t21) - (r_sapm_t22 - b_sapm_t22)
        df["striker_advantage"]  = (r_slpm_t21 * r_acc_t21)  - (b_slpm_t21 * b_acc_t21)
        df["grappler_advantage"] = (r_td_t21 * r_td_acc_t21) - (b_td_t21 * b_td_acc_t21)
        df["experience_gap"] = r_tf_t23 - b_tf_t23
        r_ko_wins_t25  = df.get("r_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        b_ko_wins_t25  = df.get("b_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        r_sub_wins_t25 = df.get("r_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_sub_wins_t25 = df.get("b_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        r_ko_rate_t25  = r_ko_wins_t25  / (r_tf_t23 + 1.0)
        b_ko_rate_t25  = b_ko_wins_t25  / (b_tf_t23 + 1.0)
        r_sub_rate_t25 = r_sub_wins_t25 / (r_tf_t23 + 1.0)
        b_sub_rate_t25 = b_sub_wins_t25 / (b_tf_t23 + 1.0)
        df["ko_specialist_gap"]         = r_ko_rate_t25  - b_ko_rate_t25
        df["submission_specialist_gap"] = r_sub_rate_t25 - b_sub_rate_t25
        r_elo_t25  = df.get("elo_r", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_elo_t25  = df.get("elo_b", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        r_traj_t25 = df.get("r_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_traj_t25 = df.get("b_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["skill_momentum"] = (r_elo_t25 - b_elo_t25) * (r_traj_t25 - b_traj_t25)
        r_loss_streak_t25 = df.get("r_pre_loss_streak", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_loss_streak_t25 = df.get("b_pre_loss_streak", pd.Series(0.0, index=df.index)).fillna(0.0)
        r_win_rate_t25 = r_wins_t23 / (r_tf_t23 + 1.0)
        b_win_rate_t25 = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_desperation"] = r_loss_streak_t25 * (1.0 / (r_win_rate_t25 + 0.1))
        df["b_desperation"] = b_loss_streak_t25 * (1.0 / (b_win_rate_t25 + 0.1))
        df["desperation_diff"] = df["r_desperation"] - df["b_desperation"]
        r_days_t25 = df.get("r_days_since_last", pd.Series(180.0, index=df.index)).fillna(180.0)
        b_days_t25 = df.get("b_days_since_last", pd.Series(180.0, index=df.index)).fillna(180.0)
        df["r_freshness"] = np.exp(-((r_days_t25 - 135.0) ** 2) / (2.0 * 90.0 ** 2))
        df["b_freshness"] = np.exp(-((b_days_t25 - 135.0) ** 2) / (2.0 * 90.0 ** 2))
        df["freshness_advantage"] = df["r_freshness"] - df["b_freshness"]

        # ── TIER 26: Stance Directional Features ──────────────────────────
        self._log("Tier 26: Stance directional features...")
        _r_st26 = df.get("r_stance", pd.Series("", index=df.index)).fillna("").astype(str).str.strip().str.lower()
        _b_st26 = df.get("b_stance", pd.Series("", index=df.index)).fillna("").astype(str).str.strip().str.lower()
        df["orthodox_vs_southpaw_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "southpaw"), 1.0,
            np.where((_r_st26 == "southpaw") & (_b_st26 == "orthodox"), -1.0, 0.0)
        ).astype(float)
        df["orthodox_vs_switch_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "switch"), 1.0,
            np.where((_r_st26 == "switch") & (_b_st26 == "orthodox"), -1.0, 0.0)
        ).astype(float)
        df["southpaw_vs_switch_advantage"] = np.where(
            (_r_st26 == "southpaw") & (_b_st26 == "switch"), 1.0,
            np.where((_r_st26 == "switch") & (_b_st26 == "southpaw"), -1.0, 0.0)
        ).astype(float)
        df["mirror_matchup"] = (_r_st26 == _b_st26).astype(float)

        # ── TIER 27: Extended Polynomial Squared Terms ────────────────────
        self._log("Tier 27: Extended polynomial squared terms...")
        def _signed_sq_t27(s):
            return np.sign(s) * (s ** 2)
        for _feat_sq in [
            "net_striking_advantage", "striker_advantage", "grappler_advantage",
            "experience_gap", "ko_specialist_gap", "submission_specialist_gap",
            "skill_momentum", "desperation_diff", "freshness_advantage",
            "combined_defensive_hole_diff", "striking_volume_accuracy_synergy_diff",
            "total_offensive_efficiency_index_diff", "finish_efficiency_diff",
            "defense_versatility_diff", "offensive_versatility_diff",
        ]:
            if _feat_sq in df.columns:
                df[f"{_feat_sq}_sq"] = _signed_sq_t27(df[_feat_sq])

        # ── TIER 28: Volatility & Career Arc Features ──────────────────────
        self._log("Tier 28: Volatility and career arc features...")
        r_fr_l5_t28  = df.get("r_pre_finish_rate_l5",  pd.Series(0.4, index=df.index)).fillna(0.4)
        b_fr_l5_t28  = df.get("b_pre_finish_rate_l5",  pd.Series(0.4, index=df.index)).fillna(0.4)
        r_fr_l10_t28 = df.get("r_pre_finish_rate_l10", pd.Series(0.4, index=df.index)).fillna(0.4)
        b_fr_l10_t28 = df.get("b_pre_finish_rate_l10", pd.Series(0.4, index=df.index)).fillna(0.4)
        df["r_finish_rate_accel"]       = r_fr_l5_t28 - r_fr_l10_t28
        df["b_finish_rate_accel"]       = b_fr_l5_t28 - b_fr_l10_t28
        df["finish_rate_acceleration_diff"] = df["r_finish_rate_accel"] - df["b_finish_rate_accel"]
        r_slpm_cv_t28 = df.get("r_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(0.3)
        b_slpm_cv_t28 = df.get("b_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(0.3)
        df["slpm_coefficient_of_variation_diff"] = r_slpm_cv_t28 - b_slpm_cv_t28
        r_mil_t28 = df.get("r_pre_mileage_adj_age", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_mil_t28 = df.get("b_pre_mileage_adj_age", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["mileage_adjusted_age_diff"] = r_mil_t28 - b_mil_t28
        df["performance_decline_velocity_diff"] = (
            df.get("r_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0) -
            df.get("b_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0)
        ) * (-1.0)
        r_cur_elo_t28 = df.get("elo_r", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_cur_elo_t28 = df.get("elo_b", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        r_peak_t28    = df.get("r_career_elo_peak", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_peak_t28    = df.get("b_career_elo_peak", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        df["r_distance_from_peak"] = r_peak_t28 - r_cur_elo_t28
        df["b_distance_from_peak"] = b_peak_t28 - b_cur_elo_t28
        df["distance_from_career_peak_diff"] = df["r_distance_from_peak"] - df["b_distance_from_peak"]
        r_fsp_t28 = df.get("r_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_fsp_t28 = df.get("b_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["r_career_inflection"] = r_fsp_t28 / (r_tf_t23 + 1.0)
        df["b_career_inflection"] = b_fsp_t28 / (b_tf_t23 + 1.0)
        df["career_inflection_point_diff"] = df["r_career_inflection"] - df["b_career_inflection"]
        df["r_prime_exit_risk"] = (r_age_t23 > 33).astype(float) * np.clip(-r_traj_t25, 0.0, 1.0)
        df["b_prime_exit_risk"] = (b_age_t23 > 33).astype(float) * np.clip(-b_traj_t25, 0.0, 1.0)
        df["prime_exit_risk_diff"] = df["r_prime_exit_risk"] - df["b_prime_exit_risk"]
        df["r_aging_power_penalty"] = r_ko_rate_t25 * r_age_t23 * (r_age_t23 > 35).astype(float)
        df["b_aging_power_penalty"] = b_ko_rate_t25 * b_age_t23 * (b_age_t23 > 35).astype(float)
        df["aging_power_striker_penalty_diff"] = df["r_aging_power_penalty"] - df["b_aging_power_penalty"]
        df["r_bayesian_finish"] = (r_ko_wins_t25 + r_sub_wins_t25 + 2.0) / (r_tf_t23 + 4.0)
        df["b_bayesian_finish"] = (b_ko_wins_t25 + b_sub_wins_t25 + 2.0) / (b_tf_t23 + 4.0)
        df["bayesian_finish_rate_diff"] = df["r_bayesian_finish"] - df["b_bayesian_finish"]
        df["r_layoff_veteran"] = r_days_t25 * r_tf_t23
        df["b_layoff_veteran"] = b_days_t25 * b_tf_t23
        df["layoff_veteran_interaction_diff"] = df["r_layoff_veteran"] - df["b_layoff_veteran"]
        df["r_elo_momentum"] = r_cur_elo_t28 * r_traj_t25
        df["b_elo_momentum"] = b_cur_elo_t28 * b_traj_t25
        df["elo_momentum_vs_competition_diff"] = df["r_elo_momentum"] - df["b_elo_momentum"]
        r_avg_opp_elo_t28 = df.get("r_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_avg_opp_elo_t28 = df.get("b_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        df["r_title_proximity"] = r_streak_t23 * r_avg_opp_elo_t28 * r_cur_elo_t28 / 1.0e6
        df["b_title_proximity"] = b_streak_t23 * b_avg_opp_elo_t28 * b_cur_elo_t28 / 1.0e6
        df["title_shot_proximity_score_diff"] = df["r_title_proximity"] - df["b_title_proximity"]
        df["r_elo_volatility"] = r_cur_elo_t28 * r_slpm_cv_t28
        df["b_elo_volatility"] = b_cur_elo_t28 * b_slpm_cv_t28
        df["elo_volatility_interaction_diff"] = df["r_elo_volatility"] - df["b_elo_volatility"]
        r_fin_l10_t28 = df.get("r_rolling10_finishes", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_fin_l10_t28 = df.get("b_rolling10_finishes", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["elite_performance_frequency_l10_diff"] = (r_fin_l10_t28 / 10.0) - (b_fin_l10_t28 / 10.0)
        _r_dr = df.get("r_damage_ratio", pd.Series(1.0, index=df.index)).fillna(1.0)
        _b_dr = df.get("b_damage_ratio", pd.Series(1.0, index=df.index)).fillna(1.0)
        df["r_conf_damage_ratio"] = _r_dr * (1.0 - 1.0 / (r_tf_t23 ** 0.5 + 1.0))
        df["b_conf_damage_ratio"] = _b_dr * (1.0 - 1.0 / (b_tf_t23 ** 0.5 + 1.0))
        df["confidence_weighted_damage_ratio_diff"] = df["r_conf_damage_ratio"] - df["b_conf_damage_ratio"]
        # recent_vs_career_striking: rolling SLpM relative to career SLpM
        _r_r5slpm = df.get("r_pre_rolling5_slpm", pd.Series(3.0, index=df.index)).fillna(3.0)
        _b_r5slpm = df.get("b_pre_rolling5_slpm", pd.Series(3.0, index=df.index)).fillna(3.0)
        df["r_recent_vs_career_striking"] = _r_r5slpm / (r_slpm_t21 + 0.1)
        df["b_recent_vs_career_striking"] = _b_r5slpm / (b_slpm_t21 + 0.1)
        df["recent_vs_career_striking_diff"] = df["r_recent_vs_career_striking"] - df["b_recent_vs_career_striking"]
        # striking_consistency_ratio: inverse of SLpM std (lower variance = more consistent)
        _r_slpmstd = df.get("r_pre_slpm_std_l10", pd.Series(1.0, index=df.index)).fillna(1.0)
        _b_slpmstd = df.get("b_pre_slpm_std_l10", pd.Series(1.0, index=df.index)).fillna(1.0)
        df["r_striking_consistency_ratio"] = 1.0 / (_r_slpmstd + 0.1)
        df["b_striking_consistency_ratio"] = 1.0 / (_b_slpmstd + 0.1)
        df["striking_consistency_ratio_diff"] = df["r_striking_consistency_ratio"] - df["b_striking_consistency_ratio"]
        # performance_volatility_l10: std of damage ratio over last 10 fights
        _r_drstd = df.get("r_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)).fillna(0.3)
        _b_drstd = df.get("b_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)).fillna(0.3)
        df["performance_volatility_l10_diff"] = _r_drstd - _b_drstd
        # tactical_evolution_score: change in distance-based fighting style (recent vs career)
        _r_tact = df.get("r_pre_tactical_evolution", pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_tact = df.get("b_pre_tactical_evolution", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["tactical_evolution_score_diff"] = _r_tact - _b_tact

        # ── TIER 29: New signal features ─────────────────────────────────────
        self._log("Tier 29: New signal features...")

        # -- Item 5: Loss method breakdown (chin & submission vulnerability) ---
        for _px in ("r", "b"):
            _ko_l  = df.get(f"{_px}_pre_ko_losses",  pd.Series(0.0, index=df.index)).fillna(0.0)
            _sub_l = df.get(f"{_px}_pre_sub_losses", pd.Series(0.0, index=df.index)).fillna(0.0)
            _tot_l = df.get(f"{_px}_pre_losses",     pd.Series(1.0, index=df.index)).fillna(1.0).clip(lower=1)
            df[f"{_px}_ko_loss_rate"]  = _ko_l  / _tot_l
            df[f"{_px}_sub_loss_rate"] = _sub_l / _tot_l
        df["chin_vulnerability_diff"] = df["r_ko_loss_rate"]  - df["b_ko_loss_rate"]
        df["sub_vulnerability_diff"]  = df["r_sub_loss_rate"] - df["b_sub_loss_rate"]

        # -- Item 6: Strike accuracy differential --------------------------------
        _r_sa = df.get("r_pre_sig_str_acc", pd.Series(0.43, index=df.index)).fillna(0.43)
        _b_sa = df.get("b_pre_sig_str_acc", pd.Series(0.43, index=df.index)).fillna(0.43)
        df["diff_sig_str_acc"] = _r_sa - _b_sa
        # Volume × accuracy: distinguish high-volume/low-acc from low-volume/high-acc
        _r_slpm_t29 = df.get("r_pre_SLpM", pd.Series(3.0, index=df.index)).fillna(3.0)
        _b_slpm_t29 = df.get("b_pre_SLpM", pd.Series(3.0, index=df.index)).fillna(3.0)
        df["r_effective_striking"] = _r_slpm_t29 * _r_sa
        df["b_effective_striking"] = _b_slpm_t29 * _b_sa
        df["effective_striking_diff"] = df["r_effective_striking"] - df["b_effective_striking"]

        # -- Item 7: Physical matchup interaction features -----------------------
        _reach_diff_t29  = df.get("diff_reach",     pd.Series(0.0, index=df.index)).fillna(0.0)
        _height_diff_t29 = df.get("diff_height",    pd.Series(0.0, index=df.index)).fillna(0.0)
        _style_edge_t29  = df.get("style_matchup_edge", pd.Series(0.0, index=df.index)).fillna(0.0)
        _r_fin_t29       = df.get("r_pre_finish_rate", pd.Series(0.4, index=df.index)).fillna(0.4)
        _b_fin_t29       = df.get("b_pre_finish_rate", pd.Series(0.4, index=df.index)).fillna(0.4)
        # Reach advantage amplified by strike accuracy edge
        df["reach_x_str_acc_diff"]  = _reach_diff_t29 * df["diff_sig_str_acc"]
        # Reach advantage interacting with style matchup edge
        df["reach_x_style_edge"]    = _reach_diff_t29 * _style_edge_t29
        # Height advantage interacting with finishing tendency
        df["height_x_finish_diff"]  = _height_diff_t29 * (_r_fin_t29 - _b_fin_t29)

        # -- Item 3: Layoff × age interaction ------------------------------------
        _r_days_t29 = df.get("r_days_since_last", pd.Series(180.0, index=df.index)).fillna(180.0)
        _b_days_t29 = df.get("b_days_since_last", pd.Series(180.0, index=df.index)).fillna(180.0)
        _r_age_t29  = df.get("r_current_age",     pd.Series(28.0,  index=df.index)).fillna(28.0)
        _b_age_t29  = df.get("b_current_age",     pd.Series(28.0,  index=df.index)).fillna(28.0)
        # Older fighters coming back from longer layoffs suffer more ring rust
        df["r_layoff_age_penalty"] = (_r_days_t29 / 365.0) * np.clip(_r_age_t29 - 28.0, 0.0, None)
        df["b_layoff_age_penalty"] = (_b_days_t29 / 365.0) * np.clip(_b_age_t29 - 28.0, 0.0, None)
        df["layoff_age_penalty_diff"] = df["r_layoff_age_penalty"] - df["b_layoff_age_penalty"]

        # -- Item 4: Weight-class-specific age curve modeling -------------------
        # Approximate peak competitive age by weight class (lighter = peaks earlier)
        _WC_PEAK_AGE = {
            "Heavyweight": 31, "Light Heavyweight": 30,
            "Middleweight": 30, "Welterweight": 29,
            "Lightweight": 28, "Featherweight": 27,
            "Bantamweight": 27, "Flyweight": 26,
            "Women's Featherweight": 28, "Women's Bantamweight": 27,
            "Women's Flyweight": 26, "Women's Strawweight": 25,
        }
        _wc_col = df.get("weight_class", pd.Series("Lightweight", index=df.index)).fillna("Lightweight")
        _wc_peak = _wc_col.map(_WC_PEAK_AGE).fillna(29.0)
        df["r_age_vs_wc_peak"] = _r_age_t29 - _wc_peak
        df["b_age_vs_wc_peak"] = _b_age_t29 - _wc_peak
        df["age_vs_peak_diff"] = df["r_age_vs_wc_peak"] - df["b_age_vs_wc_peak"]
        # Past-prime penalty: only penalise fighters past their weight-class peak
        df["r_past_prime"] = np.clip(df["r_age_vs_wc_peak"], 0.0, None)
        df["b_past_prime"] = np.clip(df["b_age_vs_wc_peak"], 0.0, None)
        df["past_prime_diff"] = df["r_past_prime"] - df["b_past_prime"]

        # r_ewm_* / b_ewm_* pairs are handled by the D+I antisymmetric
        # decomposition automatically — no manual diff columns needed here.

        self.df = df
        print_metric("Feature columns added:", len(df.columns))

    # ── ASSEMBLE FEATURE MATRIX ───────────────────────────────────────────────
    def _get_feature_cols(self):
        exclude = {
            "event_date", "event_name", "event_location",
            "r_fighter", "b_fighter", "weight_class", "gender",
            "winner", "method", "referee", "r_stance", "b_stance",
            "r_date_of_birth", "b_date_of_birth",
            "finish_round", "time_sec", "r_name", "b_name",
            "stance_matchup", "_history",
        }
        cols = []
        for col in self.df.columns:
            if col in exclude:
                continue
            if self.df[col].dtype in [object, "object"]:
                continue
            cols.append(col)
        return cols

    def _build_X_y(self, df=None):
        if df is None:
            df = self.df
        feat_cols = self._get_feature_cols()
        # Only keep columns that exist in df
        feat_cols = [c for c in feat_cols if c in df.columns]
        X = df[feat_cols].fillna(0).replace([np.inf, -np.inf], 0).values
        y_raw = df["winner"].values
        y = np.array([1 if w == "Red" else 0 for w in y_raw])
        return X, y, feat_cols

    # ── ANTISYMMETRIC DECOMPOSITION ───────────────────────────────────────────
    def _decompose_features(self, X_orig, X_swap):
        """
        Antisymmetric decomposition:
        D = 0.5 * (X_orig - X_swap)  # Directional: flips sign on corner swap
        I = 0.5 * (X_orig + X_swap)  # Invariant: unchanged on corner swap
        Returns concatenation [D | I] with _inv suffix on invariant columns.
        """
        if isinstance(X_orig, pd.DataFrame):
            cols = X_orig.columns.tolist()
            D = 0.5 * (X_orig.values - X_swap.values)
            Inv = 0.5 * (X_orig.values + X_swap.values)
            D_df = pd.DataFrame(D, columns=cols, index=X_orig.index)
            I_df = pd.DataFrame(Inv, columns=[c + '_inv' for c in cols], index=X_orig.index)
            return pd.concat([D_df, I_df], axis=1)
        else:
            D = 0.5 * (X_orig - X_swap)
            Inv = 0.5 * (X_orig + X_swap)
            return np.concatenate([D, Inv], axis=1)

    # ── TRAIN ─────────────────────────────────────────────────────────────────
    def train(self):
        print_section("TRAINING ENSEMBLE MODEL")
        self._log("Detecting GPU...")
        self.gpu_info = detect_gpu()
        print_metric("XGB GPU:", self.gpu_info["xgb"])
        print_metric("LGB GPU:", self.gpu_info["lgb"])
        print_metric("CAT GPU:", self.gpu_info["cat"])

        # Build X, y
        df_train = self.df.copy()
        # Temporal 2-way split: 90% train / 10% holdout test.
        # The former 10% validation window is now included in training — those
        # are the most recent fights before the test period, so their outcomes
        # update ELO, win streaks, and style clusters that the test-set fighters
        # rely on. Holding them out starved the model of the most relevant signal.
        # Optuna and the StackingClassifier meta-learner both use TimeSeriesSplit
        # internally on the 90%, preserving temporal order for CV.
        # A diagnostic window (last 10% of training) is kept as df_val purely
        # for in-training accuracy display; it is NOT used for any optimisation.
        n = len(df_train)
        train_end = int(n * 0.90)
        df_tr   = df_train.iloc[:train_end]
        df_test = df_train.iloc[train_end:]
        # Diagnostic window — last ~10% of training data (overlaps with training)
        diag_start = int(n * 0.80)
        df_val  = df_train.iloc[diag_start:train_end]

        # Corner-swap augmentation on train set
        n_orig = len(df_tr)   # number of original (non-augmented) training fights
        df_aug = self._corner_swap(df_tr)
        df_tr_aug = pd.concat([df_tr, df_aug], ignore_index=True)

        X_tr, y_tr, feat_cols = self._build_X_y(df_tr_aug)
        X_val, y_val, _ = self._build_X_y(df_val)
        self.feature_cols = feat_cols

        # Sample weighting removed: any mismatch between OOF fold training
        # (unweighted) and final base model fits (weighted) miscalibrates the
        # meta-learner probabilities and raises test log-loss.  The recency
        # signal is already captured through the ewm rolling features in Tier 29.
        sw_tr = None

        print_metric("Train samples (augmented):", len(X_tr))
        print_metric("Diag window samples (in-training):", len(X_val))
        print_metric("Test samples (holdout):", len(df_test))
        print_metric("Features (pre-decomposition):", len(feat_cols))

        # ── Antisymmetric feature decomposition (D + I) ───────────────────
        print_step("Applying antisymmetric feature decomposition (D + I)...")
        df_train_feat = df_tr_aug.copy()
        df_train_swap = self._corner_swap(df_train_feat)
        feat_col_list = self._get_feature_cols()
        feat_col_list = [c for c in feat_col_list if c in df_train_feat.columns]
        X_orig_df = df_train_feat[feat_col_list].fillna(0)
        X_swap_df = df_train_swap[feat_col_list].fillna(0)
        X_decomposed = self._decompose_features(X_orig_df, X_swap_df)
        self._decomposed_cols = X_decomposed.columns.tolist()
        print_metric("  Features after decomposition:", X_decomposed.shape[1])

        # Build val decomposed features
        df_val_feat = df_val.copy()
        df_val_swap = self._corner_swap(df_val_feat)
        X_val_orig_df = df_val_feat[feat_col_list].fillna(0)
        X_val_swap_df = df_val_swap[feat_col_list].fillna(0)
        X_val_decomposed = self._decompose_features(X_val_orig_df, X_val_swap_df)

        # Build test decomposed features
        df_test_feat = df_test.copy()
        df_test_swap = self._corner_swap(df_test_feat)
        X_test_orig_df = df_test_feat[feat_col_list].fillna(0)
        X_test_swap_df = df_test_swap[feat_col_list].fillna(0)
        X_test_decomposed = self._decompose_features(X_test_orig_df, X_test_swap_df)

        # Use decomposed arrays going forward
        X_tr_raw   = X_decomposed.values
        X_val_raw  = X_val_decomposed.values
        X_test_raw = X_test_decomposed.values
        y_tr   = np.array([1 if w == "Red" else 0 for w in df_train_feat["winner"].values])
        y_val  = np.array([1 if w == "Red" else 0 for w in df_val_feat["winner"].values])
        y_test = np.array([1 if w == "Red" else 0 for w in df_test_feat["winner"].values])

        # Scale
        X_tr_s   = self.scaler.fit_transform(X_tr_raw)
        X_val_s  = self.scaler.transform(X_val_raw)
        X_test_s = self.scaler.transform(X_test_raw)

        # Track D-feature indices (no _inv suffix) for corner bias diagnostic
        self._d_indices = [i for i, c in enumerate(self._decomposed_cols) if not c.endswith('_inv')]

        # Global model-based feature selection — keep top 50% of the decomposed
        # features by LightGBM split-count importance.
        # ANOVA F-score (the univariate alternative) only detects linear
        # correlations; LGB importance captures feature interactions and
        # non-linear signal (e.g. age has a non-linear relationship with
        # performance).  A lightweight LGB (100 trees) is trained purely to rank
        # features; it is completely separate from the Optuna-tuned LGB in the
        # ensemble.
        print_step("Running model-based feature selection (LGB importance, top 50%)...")
        from sklearn.feature_selection import SelectFromModel as _SFM
        if HAS_LGB:
            _sel_lgb = lgb.LGBMClassifier(
                n_estimators=100, num_leaves=31, max_depth=5,
                learning_rate=0.1, subsample=0.8, colsample_bytree=0.8,
                random_state=RANDOM_SEED, verbose=-1, n_jobs=SAFE_N_JOBS,
            )
            _sel_lgb.fit(X_tr_s, y_tr)
            # Use max_features to guarantee exactly top 1/3 by count (~400).
            # threshold="median" fails when median importance=0 (>50% zero-importance
            # features), causing SelectFromModel to keep all 1200 features.
            _n_keep = X_tr_s.shape[1] // 2   # keep top 50% by LGB importance
            self._global_selector = _SFM(
                _sel_lgb, prefit=True, max_features=_n_keep, threshold=-np.inf
            )
        else:
            # Fallback: ANOVA F-score when LGB unavailable
            from sklearn.feature_selection import SelectPercentile as _SP, f_classif as _fc
            _sp = _SP(_fc, percentile=50)
            _sp.fit(X_tr_s, y_tr)
            self._global_selector = _sp
        X_tr_sel   = self._global_selector.transform(X_tr_s)
        X_val_sel  = self._global_selector.transform(X_val_s)
        X_test_sel = self._global_selector.transform(X_test_s)
        _gsel_idx  = self._global_selector.get_support(indices=True)
        # Subset of decomposed column names that survived selection
        self._selected_decomposed_cols = [self._decomposed_cols[i] for i in _gsel_idx]
        # Recompute D-feature indices within the reduced feature space
        self._d_indices = [
            new_i for new_i, orig_i in enumerate(_gsel_idx)
            if not self._decomposed_cols[orig_i].endswith('_inv')
        ]
        print_metric("Features after model-based selection:", X_tr_sel.shape[1])

        # ── TimeSeriesSplit cross-validation ──────────────────────────────
        t0_cv = time.time()
        print_step("Running TimeSeriesSplit cross-validation (5 folds)...")
        tscv = TimeSeriesSplit(n_splits=5)
        fold_scores = []
        for fold, (tr_idx, val_idx) in enumerate(tscv.split(X_tr_sel), 1):
            # use RandomForest as quick CV estimator
            fold_rf = RandomForestClassifier(n_estimators=100, n_jobs=SAFE_N_JOBS, random_state=42)
            fold_rf.fit(X_tr_sel[tr_idx], y_tr[tr_idx])
            fold_pred = fold_rf.predict_proba(X_tr_sel[val_idx])[:, 1]
            fold_acc = (fold_rf.predict(X_tr_sel[val_idx]) == y_tr[val_idx]).mean()
            fold_ll = log_loss(y_tr[val_idx], fold_pred)
            fold_scores.append(fold_acc)
            print_metric(f"  Fold {fold} Accuracy:", f"{fold_acc:.4f}  |  Log-Loss: {fold_ll:.4f}")
        print_metric("  Mean CV Accuracy:", f"{np.mean(fold_scores):.4f} \u00b1 {np.std(fold_scores):.4f}")
        print_metric("  CV time:", f"{time.time()-t0_cv:.1f}s")

        # ── Optuna hyperparameter tuning for XGBoost ──────────────────────
        if HAS_OPTUNA and HAS_XGB:
            print_step("Running Optuna hyperparameter search for XGBoost (25 trials)...")
            optuna_start = time.time()

            def optuna_objective(trial):
                params = {
                    # Cap n_estimators at 1000: with low learning rates (0.005),
                    # 2000 trees can memorize a small UFC dataset. 1000 is sufficient.
                    'n_estimators': trial.suggest_int('n_estimators', 200, 1000),
                    'max_depth': trial.suggest_int('max_depth', 3, 8),
                    'learning_rate': trial.suggest_float('learning_rate', 0.01, 0.15, log=True),
                    'subsample': trial.suggest_float('subsample', 0.6, 1.0),
                    'colsample_bytree': trial.suggest_float('colsample_bytree', 0.5, 1.0),
                    'reg_alpha': trial.suggest_float('reg_alpha', 0.0, 5.0),
                    'reg_lambda': trial.suggest_float('reg_lambda', 1.0, 8.0),
                    # min_child_weight floor at 5 — on ~600 training fights this
                    # prevents leaves from containing only 1-4 samples.
                    'min_child_weight': trial.suggest_int('min_child_weight', 5, 30),
                    'gamma': trial.suggest_float('gamma', 0.0, 5.0),
                }
                xgb_trial = xgb.XGBClassifier(
                    **params,
                    device='cuda' if self.gpu_info.get('xgb') else 'cpu',
                    tree_method='hist',
                    eval_metric='logloss',
                    random_state=42,
                    n_jobs=1,
                    verbosity=0,
                )
                # quick 3-fold CV on training data
                tscv3 = TimeSeriesSplit(n_splits=3)
                scores = []
                for tr_i, val_i in tscv3.split(X_tr_sel):
                    xgb_trial.fit(X_tr_sel[tr_i], y_tr[tr_i],
                                  eval_set=[(X_tr_sel[val_i], y_tr[val_i])],
                                  verbose=False)
                    p = xgb_trial.predict_proba(X_tr_sel[val_i])[:, 1]
                    scores.append(-log_loss(y_tr[val_i], p))
                return np.mean(scores)

            optuna.logging.set_verbosity(optuna.logging.WARNING)
            study = optuna.create_study(
                direction='maximize',
                sampler=optuna.samplers.TPESampler(seed=RANDOM_SEED),
            )

            N_TRIALS = 25
            _optuna_best_so_far = [None]

            def _optuna_callback(study, trial):
                completed = len(study.trials)
                best_val  = -study.best_value if study.best_value is not None else float('nan')
                bar_fill  = int(completed / N_TRIALS * 30)
                bar       = "█" * bar_fill + "░" * (30 - bar_fill)
                pct       = int(completed / N_TRIALS * 100)
                print(f"\r  [{bar}] {pct:3d}%  trial {completed:2d}/{N_TRIALS}  best log-loss: {best_val:.4f}",
                      end="", flush=True)

            study.optimize(
                optuna_objective,
                n_trials=N_TRIALS,
                timeout=None,
                show_progress_bar=False,
                callbacks=[_optuna_callback],
            )
            print()  # newline after bar completes
            best_params = study.best_params
            elapsed = time.time() - optuna_start
            print_metric("Optuna best log-loss:", f"{-study.best_value:.4f}")
            print_metric("Time elapsed:", f"{elapsed:.1f}s")
            print("  Best params:")
            for k, v in best_params.items():
                print(f"    {k}: {v}")
            # Store for use in winner model
            self._optuna_best_xgb_params = best_params
        else:
            self._optuna_best_xgb_params = {}
            if not HAS_OPTUNA:
                print_step("Optuna not available — using default XGBoost params")

        # ── Optuna hyperparameter tuning for LightGBM ─────────────────────
        if HAS_OPTUNA and HAS_LGB:
            print_step("Running Optuna hyperparameter search for LightGBM (15 trials)...")
            lgb_optuna_start = time.time()
            LGB_TRIALS = 15

            def lgb_optuna_objective(trial):
                params = {
                    "n_estimators": trial.suggest_int("n_estimators", 200, 1000),
                    "num_leaves": trial.suggest_int("num_leaves", 15, 63),
                    "max_depth": trial.suggest_int("max_depth", 3, 8),
                    "learning_rate": trial.suggest_float("learning_rate", 0.01, 0.1, log=True),
                    "min_child_samples": trial.suggest_int("min_child_samples", 15, 50),
                    "reg_alpha": trial.suggest_float("reg_alpha", 0.0, 5.0),
                    "reg_lambda": trial.suggest_float("reg_lambda", 0.0, 5.0),
                    "colsample_bytree": trial.suggest_float("colsample_bytree", 0.5, 1.0),
                    "subsample": trial.suggest_float("subsample", 0.6, 1.0),
                }
                lgb_trial = lgb.LGBMClassifier(
                    **params,
                    random_state=42, verbose=-1,
                    n_jobs=SAFE_N_JOBS, class_weight="balanced",
                )
                tscv3 = TimeSeriesSplit(n_splits=3)
                scores = []
                for tr_i, val_i in tscv3.split(X_tr_sel):
                    lgb_trial.fit(X_tr_sel[tr_i], y_tr[tr_i])
                    p = lgb_trial.predict_proba(X_tr_sel[val_i])[:, 1]
                    scores.append(-log_loss(y_tr[val_i], p))
                return np.mean(scores)

            lgb_study = optuna.create_study(
                direction="maximize",
                sampler=optuna.samplers.TPESampler(seed=RANDOM_SEED),
            )
            _lgb_trials_done = [0]

            def _lgb_callback(study, trial):
                _lgb_trials_done[0] += 1
                best_val = -study.best_value if study.best_value is not None else float("nan")
                bar_fill = int(_lgb_trials_done[0] / LGB_TRIALS * 30)
                bar = "█" * bar_fill + "░" * (30 - bar_fill)
                pct = int(_lgb_trials_done[0] / LGB_TRIALS * 100)
                print(f"\r  [{bar}] {pct:3d}%  trial {_lgb_trials_done[0]:2d}/{LGB_TRIALS}"
                      f"  best log-loss: {best_val:.4f}", end="", flush=True)

            lgb_study.optimize(
                lgb_optuna_objective, n_trials=LGB_TRIALS,
                show_progress_bar=False, callbacks=[_lgb_callback],
            )
            print()
            lgb_best_params = lgb_study.best_params
            lgb_elapsed = time.time() - lgb_optuna_start
            print_metric("LGB Optuna best log-loss:", f"{-lgb_study.best_value:.4f}")
            print_metric("Time elapsed:", f"{lgb_elapsed:.1f}s")
            self._optuna_best_lgb_params = lgb_best_params
        else:
            self._optuna_best_lgb_params = {}

        # Build and fit base estimators — these fitted copies are used for the
        # per-model val accuracy display and cloned into the stacking ensemble.
        t0_estimators = time.time()
        estimators = self._build_estimators(X_tr_sel, y_tr, sample_weight=sw_tr)
        print_metric("Base estimators time:", f"{time.time()-t0_estimators:.1f}s")

        # ── Manual OOF stacking with KFold-3 ───────────────────────────────
        # Replaces StackingClassifier to fix the NaN stacking bug: sklearn's
        # StackingClassifier uses cross_val_predict internally, which produces
        # NaN meta-features for rf/mlp/cat in our setup, leaving those models
        # with NaN LR coefficients and making them useless.  _ManualStackingEnsemble
        # does the same KFold-3 OOF stacking explicitly, catching NaN/exceptions
        # per model per fold so all 6 base models contribute valid meta-features.
        print_step("Building stacking ensemble (manual KFold-3 OOF meta-learning)...")
        t0_stack = time.time()
        _stk = _ManualStackingEnsemble(
            estimators=estimators,   # already fitted by _build_estimators
            meta_C=0.05,
            n_splits=3,
            random_state=RANDOM_SEED,
        )
        # OOF meta-features must use only original (non-augmented) rows.
        # The augmented dataset has corner-swapped duplicates at rows n_orig..2*n_orig.
        # With KFold-3 on the full 2×n_orig set, the mirror of each validation fight
        # lands in the training fold — pure data leakage that inflates OOF accuracy
        # by ~4% relative to true holdout.  Base models are still fitted on the full
        # augmented set (via _build_estimators above); only the OOF phase is restricted.
        _stk.fit(X_tr_sel[:n_orig], y_tr[:n_orig], sample_weight=sw_tr)
        # Store for downstream use
        self._base_ensemble = _stk
        self.stacking_clf   = _stk

        # Show meta-learner coefficients (model contributions)
        try:
            _meta = _stk.final_estimator_
            if hasattr(_meta, "coef_"):
                _coefs     = _meta.coef_[0]
                _est_names = [n for n, _ in estimators]
                nc         = len(_stk.classes_)
                print_step("Meta-learner model contributions "
                           "(LR coef on class-1 probability column):")
                for _ei, _ename in enumerate(_est_names):
                    _ci = _ei * nc + 1   # class-1 column
                    _c  = float(_coefs[_ci]) if _ci < len(_coefs) else float("nan")
                    print(f"    {_ename:<6s}  coef={_c:+.4f}")
        except Exception:
            pass
        print_metric("Stacking ensemble time:", f"{time.time()-t0_stack:.1f}s")

        # ── Winner model evaluation on diagnostic window (in-training) ────
        # This window overlaps with training data — accuracy is optimistic.
        # Use HOLDOUT TEST SET METRICS below for true out-of-sample performance.
        print_section("WINNER MODEL — DIAGNOSTIC WINDOW (in-training, last 10% of train)")
        try:
            val_pred  = self.stacking_clf.predict(X_val_sel)
            val_proba = self.stacking_clf.predict_proba(X_val_sel)
            # predict returns 0/1; class order matches stacking_clf.classes_
            classes = list(self.stacking_clf.classes_)
            r_idx = classes.index(1) if 1 in classes else 1
            val_proba_pos = val_proba[:, r_idx]

            val_acc = accuracy_score(y_val, val_pred)
            val_ll  = log_loss(y_val, np.column_stack([1.0 - val_proba_pos, val_proba_pos]))

            # Baseline: always predict majority class
            majority_acc = max(np.mean(y_val == 1), np.mean(y_val == 0))

            print_metric("Val Accuracy:",        f"{val_acc:.4f}")
            print_metric("Val Log-Loss:",         f"{val_ll:.4f}")
            print_metric("Majority-class baseline:", f"{majority_acc:.4f}")
            print_metric("Lift over baseline:",   f"{val_acc - majority_acc:+.4f}")

            # Per-model breakdown
            print_step("Per-estimator val accuracy:")
            _base_ests = None
            if isinstance(self.stacking_clf, _ManualStackingEnsemble):
                _base_ests = self.stacking_clf.estimators  # list of (name, est)
            elif hasattr(self.stacking_clf, "named_estimators_"):
                _base_ests = list(self.stacking_clf.named_estimators_.items())
            if _base_ests is not None:
                for est_name, est in _base_ests:
                    try:
                        est_pred = est.predict(X_val_sel)
                        est_acc  = accuracy_score(y_val, est_pred)
                        print_metric(f"  {est_name}:", f"{est_acc:.4f}")
                    except Exception:
                        pass
        except Exception as _e:
            print(f"  (metrics unavailable: {_e})")

        # Feature importance — all non-zero features, per model.
        # Uses _selected_decomposed_cols (the 600 names that survived global selection).
        try:
            _feat_names = (getattr(self, '_selected_decomposed_cols', None) or
                           getattr(self, '_decomposed_cols', None) or
                           self.feature_cols)
            # Resolve base estimator list — handles _ManualStackingEnsemble and
            # StackingClassifier (and calibration wrappers around either).
            _fi_ests = None
            winner_inner = getattr(self, '_base_ensemble', None)
            if isinstance(winner_inner, _ManualStackingEnsemble):
                _fi_ests = winner_inner.estimators   # list of (name, est)
            elif winner_inner is not None and hasattr(winner_inner, "named_estimators_"):
                _fi_ests = list(winner_inner.named_estimators_.items())
            elif hasattr(self.stacking_clf, "named_estimators_"):
                _fi_ests = list(self.stacking_clf.named_estimators_.items())
            if _fi_ests is not None:
                for _model_name, _est in _fi_ests:
                    # Unwrap Pipeline to get the underlying estimator's importances
                    _base = _est
                    if hasattr(_est, 'named_steps'):
                        for _step in reversed(list(_est.named_steps.values())):
                            if hasattr(_step, 'feature_importances_'):
                                _base = _step
                                break
                    if not hasattr(_base, 'feature_importances_'):
                        continue
                    _imp = _base.feature_importances_
                    # Pair each feature with its importance; filter zero-importance
                    _pairs = [(i, _feat_names[i], _imp[i])
                              for i in range(len(_imp))
                              if i < len(_feat_names) and _imp[i] > 1e-6]
                    _pairs.sort(key=lambda x: x[2], reverse=True)
                    print_section(f"FEATURE IMPORTANCES — {_model_name.upper()} "
                                  f"(top 20 of {len(_pairs)} non-zero / {len(_imp)} total)")
                    for _rank, (_i, _fname, _fval) in enumerate(_pairs[:20], 1):
                        print(f"  {_rank:4d}. {_fname:55s}: {_fval:.6f}")
        except Exception as _fie:
            print(f"  (feature importance unavailable: {_fie})")

        # ── Corner bias diagnostic ────────────────────────────────────────
        print_step("Running corner bias diagnostic...")
        self._corner_bias_diagnostic(X_val_sel, y_val)

        # ── Holdout test set evaluation (never used in any optimisation step) ──
        print_section("WINNER MODEL — HOLDOUT TEST SET METRICS")
        try:
            test_pred  = self.stacking_clf.predict(X_test_sel)
            test_proba = self.stacking_clf.predict_proba(X_test_sel)
            classes_t  = list(self.stacking_clf.classes_)
            r_idx_t    = classes_t.index(1) if 1 in classes_t else 1
            test_proba_pos = test_proba[:, r_idx_t]

            test_acc = accuracy_score(y_test, test_pred)
            test_ll  = log_loss(y_test, np.column_stack([1.0 - test_proba_pos, test_proba_pos]))
            majority_acc_t = max(np.mean(y_test == 1), np.mean(y_test == 0))

            print_metric("Test Accuracy:",           f"{test_acc:.4f}")
            print_metric("Test Log-Loss:",           f"{test_ll:.4f}")
            print_metric("Test samples:",            len(y_test))
            print_metric("Majority-class baseline:", f"{majority_acc_t:.4f}")
            print_metric("Lift over baseline:",      f"{test_acc - majority_acc_t:+.4f}")
        except Exception as _e:
            print(f"  (test metrics unavailable: {_e})")

        # ── Method classifier ─────────────────────────────────────────────
        self._log("Training method classifier (Decision/KO/Submission)...")
        self._train_method_clf(df_tr_aug, df_val)

        self.is_trained = True
        self._log("Training complete!")

    # ── CORNER BIAS DIAGNOSTIC ────────────────────────────────────────────────
    def _corner_bias_diagnostic(self, X_val, _y_val):
        """
        Checks if model has learned corner position bias.
        Flips directional features (negate first half = D features) and checks parity.
        P(red wins | orig) + P(red wins | flipped) should ≈ 1.0
        """
        print_section("CORNER BIAS DIAGNOSTIC")

        # Flip only the D features (directional, no _inv suffix) in the decomposed space
        d_indices = getattr(self, '_d_indices', None)
        if d_indices is None or len(d_indices) == 0:
            print("  Skipping: D-feature indices not available.")
            return 0.0, 0.0, 0.0

        X_flipped = X_val.copy()
        X_flipped[:, d_indices] = -X_val[:, d_indices]

        # Predict on original and flipped
        p_orig = self.stacking_clf.predict_proba(X_val)[:, 1]  # P(red wins | orig)
        p_flip = self.stacking_clf.predict_proba(X_flipped)[:, 1]  # P(red wins | flipped)

        parity = p_orig + p_flip  # Should be 1.0 for perfectly unbiased model
        parity_error = np.abs(parity - 1.0)

        mean_error = parity_error.mean()
        max_error = parity_error.max()

        # Inherent red bias: average P(red) across all original predictions
        red_bias = p_orig.mean() - 0.5

        print_metric("  Mean parity error:", f"{mean_error:.4f}")
        print_metric("  Max parity error:", f"{max_error:.4f}")
        print_metric("  Inherent red bias:", f"{red_bias:+.4f} ({red_bias*100:+.2f}%)")

        # UFC books better-ranked fighter red, so ~56-58% red wins is real.
        # A model bias of ~3-6% is therefore expected legitimate signal.
        if abs(red_bias) < 0.03:
            print("  GOOD: Bias < 3% -- Consistent with fighter skill signal")
        elif abs(red_bias) < 0.07:
            print("  NOTE: Bias 3-7% -- Consistent with real UFC red-corner advantage (~56-58% red wins)")
        else:
            print("  WARNING: Bias > 7% -- Model may have over-learned corner position!")

        # Store for informational tracking (not used to adjust predictions)
        self._red_corner_bias = red_bias

        return mean_error, max_error, red_bias

    def _corner_swap(self, df):
        """
        Create corner-swapped augmentation. Swaps r_/b_ base columns and recomputes
        ALL derived features from scratch to ensure correctness for every tier.
        """
        df_s = df.copy()

        # 1. Swap all r_*/b_* column pairs
        r_cols = [c for c in df_s.columns if c.startswith("r_")]
        b_cols = [c for c in df_s.columns if c.startswith("b_")]

        rename_map = {}
        for c in r_cols:
            rename_map[c] = "b_" + c[2:]
        for c in b_cols:
            rename_map[c] = "r_" + c[2:]
        df_s = df_s.rename(columns=rename_map)

        # 2. Flip winner label
        def flip_winner(w):
            if w == "Red":
                return "Blue"
            if w == "Blue":
                return "Red"
            return w
        df_s["winner"] = df_s["winner"].apply(flip_winner)

        # 3. Drop all derived/computed columns — keep only base r_/b_ stats,
        #    metadata, and pre-fight snapshot columns that are truly per-fighter
        derived_patterns = (
            "diff_", "elo_diff", "elo_r", "elo_b", "elo_ratio",
            "glicko_diff", "glicko_rd_diff", "r_glicko_r", "b_glicko_r",
            "r_glicko_rd", "b_glicko_rd", "r_glicko_vol", "b_glicko_vol",
            "n_common_", "common_opp_", "r_wins_vs_", "b_wins_vs_",
            "r_style_cluster", "b_style_cluster", "style_matchup_edge",
            "r_style_win_vs_cluster", "b_style_win_vs_cluster",
            "same_stance", "stance_matchup",
            "momentum_diff_", "streak_differential",
            "ko_threat_diff", "sub_threat_diff", "dec_tendency_diff",
            "r_finishing_tendency", "b_finishing_tendency", "finishing_matchup",
            "r_ring_rust", "b_ring_rust", "ring_rust_diff",
            "weight_class_ko_factor", "style_clash_severity",
            "upset_potential", "power_vs_technique", "championship_pressure",
            "r_clinch_effectiveness", "b_clinch_effectiveness", "clinch_effectiveness_diff",
            "five_round_cardio_advantage",
            "r_chin_deterioration", "b_chin_deterioration", "chin_deterioration_diff",
            "finishing_pressure_diff",
            "r_overactive", "b_overactive", "overactivity_diff",
            "opp_quality_diff", "trajectory_diff",
            "r_peak_score", "b_peak_score", "peak_score_diff",
            "uncertainty_score",
            "striking_svd_", "grappling_svd_", "physical_svd_", "form_svd_",
            "z_r_", "z_b_",
            "is_title_enc", "total_rounds_num", "gender_enc",
            "positional_striking_advantage", "target_distribution_advantage",
            "defensive_composite",
            "elo_x_form", "elo_x_win_ratio", "elo_x_finish", "elo_x_durability",
            "reach_x_striking", "height_x_reach", "physical_x_striking",
            "age_x_striking", "age_x_grappling", "age_x_durability",
            "age_x_win_streak", "experience_x_age",
            "td_x_defense", "submission_x_grappling",
            "striking_x_accuracy", "striking_x_defense", "ko_power_x_striking",
            "momentum_x_win_streak", "form_x_experience", "finish_x_momentum",
            "form_x_durability", "elite_finisher", "unstoppable_streak",
            "veteran_advantage",
            "diff_age_cubed",
            "diff_win_rate_vs_elite", "diff_win_rate_vs_strikers",
            "diff_win_rate_vs_grapplers", "championship_readiness",
            "r_prime_score", "b_prime_score", "prime_years_advantage",
            "declining_phase_diff",
            "r_last_fight_momentum", "b_last_fight_momentum", "last_fight_momentum_diff",
            "rounds_x_cardio", "rounds_x_finish_rate", "rounds_x_durability",
            "r_striking_vs_b_defense", "b_striking_vs_r_defense", "striking_exploitation_diff",
            "r_td_vs_b_td_defense", "b_td_vs_r_td_defense", "td_exploitation_diff",
            "r_sub_setup_efficiency", "b_sub_setup_efficiency", "sub_setup_diff",
            "r_sub_threat_vs_td_defense", "b_sub_threat_vs_td_defense", "sub_threat_vs_defense_diff",
            "r_striking_quality", "b_striking_quality", "striking_quality_diff",
            "r_accuracy_under_fire", "b_accuracy_under_fire", "accuracy_under_fire_diff",
            "r_damage_ratio", "b_damage_ratio", "damage_ratio_diff",
            "r_striking_output_quality", "b_striking_output_quality", "striking_output_quality_diff",
            "r_grappling_quality", "b_grappling_quality", "grappling_quality_diff",
            "r_total_defense_index", "b_total_defense_index", "total_defense_diff",
            "r_complete_fighter_index", "b_complete_fighter_index", "complete_fighter_diff",
            "r_pressure_index", "b_pressure_index", "pressure_index_diff",
            "elo_x_finish_rate", "streak_x_finish", "striking_exchange",
            "td_efficiency", "control_accuracy",
            "r_stance_enc", "b_stance_enc",
            "r_decision_win_rate", "b_decision_win_rate",
            "r_ko_win_rate", "b_ko_win_rate",
            "r_sub_win_rate", "b_sub_win_rate",
            "r_finish_rate", "b_finish_rate",
            "r_title_fight_exp", "b_title_fight_exp",
            "r_main_event_exp", "b_main_event_exp",
            "decision_win_rate_diff", "ko_win_rate_diff", "sub_win_rate_diff",
            "finish_rate_diff", "title_fight_exp_diff", "main_event_exp_diff",
            # Tier 23 — statistical ratio features
            "r_defense_offense_balance", "b_defense_offense_balance", "defense_offense_balance_diff",
            "r_td_defense_offense_balance", "b_td_defense_offense_balance", "td_defense_offense_balance_diff",
            "finish_efficiency_diff",
            "r_precision_striking", "b_precision_striking", "precision_striking_diff",
            "r_quality_grappling_23", "b_quality_grappling_23", "quality_grappling_diff",
            "r_submission_threat_ratio", "b_submission_threat_ratio", "submission_threat_ratio_diff",
            "r_damage_absorption_efficiency", "b_damage_absorption_efficiency", "damage_absorption_efficiency_diff",
            "r_defense_versatility", "b_defense_versatility", "defense_versatility_diff",
            "r_total_offense_index", "b_total_offense_index", "total_offense_index_diff",
            "r_offensive_versatility", "b_offensive_versatility", "offensive_versatility_diff",
            "r_striker_index", "b_striker_index", "striker_index_diff",
            "win_loss_ratio_squared_diff",
            "r_experience_quality", "b_experience_quality", "experience_quality_diff",
            "r_win_efficiency", "b_win_efficiency", "win_efficiency_diff",
            "r_momentum_quality", "b_momentum_quality", "momentum_quality_diff",
            "r_reach_efficiency", "b_reach_efficiency", "reach_efficiency_diff",
            "r_size_adjusted_striking", "b_size_adjusted_striking", "size_adjusted_striking_diff",
            "r_size_adjusted_grappling", "b_size_adjusted_grappling", "size_adjusted_grappling_diff",
            "r_counter_fighter_index", "b_counter_fighter_index", "counter_fighter_index_diff",
            "r_finishing_threat_composite", "b_finishing_threat_composite", "finishing_threat_composite_diff",
            "r_complete_geo", "b_complete_geo", "complete_geo_diff",
            "r_pressure_fighter_index", "b_pressure_fighter_index", "pressure_fighter_index_diff",
            # Tier 24 — additional matchup features
            "r_absorption_vuln", "b_absorption_vuln", "absorption_vulnerability_index_diff",
            "r_combined_def_hole", "b_combined_def_hole", "combined_defensive_hole_diff",
            "r_td_pressure_t24", "b_td_pressure_t24", "td_vulnerability_under_pressure_diff",
            "r_strike_pressure_t24", "b_strike_pressure_t24", "strike_defense_under_volume_diff",
            "r_ctrl_sub_ratio", "b_ctrl_sub_ratio", "grappling_control_vs_submission_ratio_diff",
            "r_sub_def_necessity", "b_sub_def_necessity", "submission_defense_necessity_diff",
            "r_strike_synergy", "b_strike_synergy", "striking_volume_accuracy_synergy_diff",
            "r_td_paradox", "b_td_paradox", "takedown_efficiency_paradox_diff",
            "r_total_off_eff", "b_total_off_eff", "total_offensive_efficiency_index_diff",
            "r_sg_corr", "b_sg_corr", "striking_grappling_efficiency_correlation_diff",
            "r_def_allocation_balance", "b_def_allocation_balance", "defense_allocation_balance_diff",
            "r_combat_eff", "b_combat_eff", "total_combat_efficiency_index_diff",
            # Tier 25 — named composites
            "net_striking_advantage", "striker_advantage", "grappler_advantage", "experience_gap",
            "ko_specialist_gap", "submission_specialist_gap", "skill_momentum",
            "r_desperation", "b_desperation", "desperation_diff",
            "r_freshness", "b_freshness", "freshness_advantage",
            # Tier 26 — stance directional
            "orthodox_vs_southpaw_advantage", "orthodox_vs_switch_advantage",
            "southpaw_vs_switch_advantage", "mirror_matchup",
            # Tier 27 — extended squared terms (matched by _sq suffix below)
            # Tier 28 — volatility & career arc
            "r_finish_rate_accel", "b_finish_rate_accel", "finish_rate_acceleration_diff",
            "slpm_coefficient_of_variation_diff", "mileage_adjusted_age_diff",
            "performance_decline_velocity_diff",
            "r_distance_from_peak", "b_distance_from_peak", "distance_from_career_peak_diff",
            "r_career_inflection", "b_career_inflection", "career_inflection_point_diff",
            "r_prime_exit_risk", "b_prime_exit_risk", "prime_exit_risk_diff",
            "r_aging_power_penalty", "b_aging_power_penalty", "aging_power_striker_penalty_diff",
            "r_bayesian_finish", "b_bayesian_finish", "bayesian_finish_rate_diff",
            "r_layoff_veteran", "b_layoff_veteran", "layoff_veteran_interaction_diff",
            "r_elo_momentum", "b_elo_momentum", "elo_momentum_vs_competition_diff",
            "r_title_proximity", "b_title_proximity", "title_shot_proximity_score_diff",
            "r_elo_volatility", "b_elo_volatility", "elo_volatility_interaction_diff",
            "elite_performance_frequency_l10_diff",
            "r_conf_damage_ratio", "b_conf_damage_ratio", "confidence_weighted_damage_ratio_diff",
            # 7 missing features (Tier 23 additions)
            "r_recent_form_ratio", "b_recent_form_ratio", "recent_form_ratio_diff",
            "r_finish_method_diversity", "b_finish_method_diversity", "finish_method_diversity_diff",
            "r_cross_domain_compensation", "b_cross_domain_compensation", "cross_domain_compensation_index_diff",
            # 7 missing features (Tier 28 additions)
            "r_recent_vs_career_striking", "b_recent_vs_career_striking", "recent_vs_career_striking_diff",
            "r_striking_consistency_ratio", "b_striking_consistency_ratio", "striking_consistency_ratio_diff",
            "performance_volatility_l10_diff",
            "tactical_evolution_score_diff",
        )

        # Drop cols that match derived patterns (prefix match or exact)
        cols_to_drop_set = set()
        for col in df_s.columns:
            for pat in derived_patterns:
                if pat.endswith("_"):
                    if col.startswith(pat):
                        cols_to_drop_set.add(col)
                        break
                else:
                    if col == pat or col.endswith("_sq") or col.endswith("_abs"):
                        cols_to_drop_set.add(col)
                        break
        # Also drop SVD output columns by prefix
        for col in df_s.columns:
            for svd_prefix in ("striking_svd_", "grappling_svd_", "physical_svd_", "form_svd_"):
                if col.startswith(svd_prefix):
                    cols_to_drop_set.add(col)
            for z_prefix in ("z_r_", "z_b_"):
                if col.startswith(z_prefix):
                    cols_to_drop_set.add(col)

        df_s = df_s.drop(columns=[c for c in cols_to_drop_set if c in df_s.columns], errors="ignore")

        # 4. Recompute all derived features via the shared vectorized function
        df_s = self._recompute_derived_features(df_s)

        return df_s

    def _recompute_derived_features(self, df):
        """
        Recompute ALL derived features for a DataFrame that already has r_/b_ base columns.
        Used by both _corner_swap() (training augmentation) and _compute_row_features()
        (prediction time). Does NOT re-fit SVD/clustering — uses already-fitted models.
        """

        # ── TIER 0: Raw column diffs ──────────────────────────────────────────
        raw_pairs = [
            # Physical / biographical — available before the fight
            ("r_height", "b_height"), ("r_reach", "b_reach"),
            ("r_weight", "b_weight"),
            ("r_age_at_event", "b_age_at_event"),
            ("r_ape_index", "b_ape_index"),
        ]
        for rc, bc in raw_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = pd.to_numeric(df[rc], errors="coerce").fillna(0) - \
                                        pd.to_numeric(df[bc], errors="coerce").fillna(0)

        # ── TIER 1: Pre-fight stat diffs ──────────────────────────────────────
        pre_pairs = [
            ("r_pre_wins", "b_pre_wins"), ("r_pre_losses", "b_pre_losses"),
            ("r_pre_ko_wins", "b_pre_ko_wins"), ("r_pre_sub_wins", "b_pre_sub_wins"),
            ("r_pre_dec_wins", "b_pre_dec_wins"),
            ("r_pre_total_fights", "b_pre_total_fights"),
            ("r_pre_finish_rate", "b_pre_finish_rate"),
            ("r_pre_win_streak", "b_pre_win_streak"),
            ("r_pre_loss_streak", "b_pre_loss_streak"),
            ("r_pre_title_fights", "b_pre_title_fights"),
            ("r_pre_title_wins", "b_pre_title_wins"),
            ("r_pre_avg_fight_time", "b_pre_avg_fight_time"),
            ("r_pre_sig_str_acc", "b_pre_sig_str_acc"),
            ("r_pre_td_acc", "b_pre_td_acc"),
            ("r_pre_sub_att_rate", "b_pre_sub_att_rate"),
            ("r_pre_kd_rate", "b_pre_kd_rate"),
            ("r_pre_ctrl_avg", "b_pre_ctrl_avg"),
            ("r_pre_SLpM", "b_pre_SLpM"),
            ("r_pre_SApM", "b_pre_SApM"),
            ("r_pre_td_avg", "b_pre_td_avg"),
        ]
        for rc, bc in pre_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[6:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        # ── TIER 2: Rolling window diffs ──────────────────────────────────────
        rolling_pairs = [
            ("r_rolling3_wins", "b_rolling3_wins"),
            ("r_rolling3_sig_str", "b_rolling3_sig_str"),
            ("r_rolling3_td", "b_rolling3_td"),
            ("r_rolling3_kd", "b_rolling3_kd"),
            ("r_rolling3_sub_att", "b_rolling3_sub_att"),
            ("r_rolling5_wins", "b_rolling5_wins"),
            ("r_rolling5_sig_str", "b_rolling5_sig_str"),
            ("r_rolling5_td", "b_rolling5_td"),
            ("r_rolling5_kd", "b_rolling5_kd"),
        ]
        for rc, bc in rolling_pairs:
            if rc in df.columns and bc in df.columns:
                df[f"diff_{rc[2:]}"] = df[rc].fillna(0) - df[bc].fillna(0)

        # ── TIER 3: ELO features ──────────────────────────────────────────────
        if "r_elo_pre_fight" in df.columns and "b_elo_pre_fight" in df.columns:
            df["elo_diff"]  = df["r_elo_pre_fight"] - df["b_elo_pre_fight"]
            df["elo_r"]     = df["r_elo_pre_fight"]
            df["elo_b"]     = df["b_elo_pre_fight"]
            df["elo_ratio"] = df["r_elo_pre_fight"] / (df["b_elo_pre_fight"] + 1e-6)
        else:
            df["elo_diff"]  = 0.0
            df["elo_r"]     = 1500.0
            df["elo_b"]     = 1500.0
            df["elo_ratio"] = 1.0

        # ── TIER 4: Glicko-2 features ─────────────────────────────────────────
        if "r_glicko_pre_r" in df.columns and "b_glicko_pre_r" in df.columns:
            df["r_glicko_r"]   = df["r_glicko_pre_r"]
            df["r_glicko_rd"]  = df["r_glicko_pre_rd"]
            df["r_glicko_vol"] = df["r_glicko_pre_vol"]
            df["b_glicko_r"]   = df["b_glicko_pre_r"]
            df["b_glicko_rd"]  = df["b_glicko_pre_rd"]
            df["b_glicko_vol"] = df["b_glicko_pre_vol"]
        df["glicko_diff"]    = df.get("r_glicko_r",  pd.Series(1500.0, index=df.index)).fillna(1500.0) - \
                               df.get("b_glicko_r",  pd.Series(1500.0, index=df.index)).fillna(1500.0)
        df["glicko_rd_diff"] = df.get("r_glicko_rd", pd.Series(200.0,  index=df.index)).fillna(200.0) - \
                               df.get("b_glicko_rd", pd.Series(200.0,  index=df.index)).fillna(200.0)

        # ── TIER 5: Weight-class Z-scores (apply existing distributions) ───────
        fe = self.feature_engineer
        z_feats = ["r_pre_SLpM", "r_pre_SApM", "r_pre_sig_str_acc",
                   "r_pre_td_avg", "r_pre_sub_att_rate", "r_pre_kd_rate",
                   "b_pre_SLpM", "b_pre_SApM", "b_pre_sig_str_acc",
                   "b_pre_td_avg", "b_pre_sub_att_rate", "b_pre_kd_rate"]
        for col in z_feats:
            df[f"z_{col}"] = 0.0
        if "weight_class" in df.columns and "event_date" in df.columns:
            for idx, row in df.iterrows():
                wc = str(row.get("weight_class", ""))
                yr_val = row.get("event_date", None)
                try:
                    yr = pd.Timestamp(yr_val).year if pd.notna(yr_val) else 2000
                except Exception:
                    yr = 2000
                for feat in z_feats:
                    if feat in df.columns:
                        v = row.get(feat, 0)
                        try:
                            v = float(v)
                            if not math.isnan(v):
                                df.at[idx, f"z_{feat}"] = fe.get_z_score(wc, yr, feat, v)
                        except (TypeError, ValueError):
                            pass

        # ── TIER 6: Common opponent features ──────────────────────────────────
        if "r_fighter" in df.columns and "b_fighter" in df.columns:
            n_common, r_wins_c, b_wins_c, co_edge = [], [], [], []
            for _, row in df.iterrows():
                r = str(row.get("r_fighter", ""))
                b = str(row.get("b_fighter", ""))
                feat = fe.get_common_opponent_features(r, b)
                n_common.append(feat["n_common_opponents"])
                r_wins_c.append(feat["r_wins_vs_common"])
                b_wins_c.append(feat["b_wins_vs_common"])
                co_edge.append(feat["common_opp_edge"])
            df["n_common_opponents"] = n_common
            df["r_wins_vs_common"]   = r_wins_c
            df["b_wins_vs_common"]   = b_wins_c
            df["common_opp_edge"]    = co_edge

        # ── TIER 7: Style cluster features ────────────────────────────────────
        if "r_fighter" in df.columns and "b_fighter" in df.columns:
            r_cluster, b_cluster, style_edge_list = [], [], []
            r_style_win, b_style_win = [], []
            for _, row in df.iterrows():
                r = str(row.get("r_fighter", ""))
                b = str(row.get("b_fighter", ""))
                rc = fe.get_fighter_cluster(r)
                bc = fe.get_fighter_cluster(b)
                mf = fe.get_style_matchup_features(rc, bc)
                r_cluster.append(rc)
                b_cluster.append(bc)
                style_edge_list.append(mf["style_matchup_edge"])
                r_style_win.append(mf["r_style_win_vs_opp_cluster"])
                b_style_win.append(mf["b_style_win_vs_opp_cluster"])
            df["r_style_cluster"]        = r_cluster
            df["b_style_cluster"]        = b_cluster
            df["style_matchup_edge"]     = style_edge_list
            df["r_style_win_vs_cluster"] = r_style_win
            df["b_style_win_vs_cluster"] = b_style_win

        # ── TIER 8: Stance encoding ────────────────────────────────────────────
        stance_map = {"Orthodox": 0, "Southpaw": 1, "Switch": 2, "Open Stance": 3}
        for col in ["r_stance", "b_stance"]:
            if col in df.columns:
                df[f"{col}_enc"] = df[col].map(stance_map).fillna(-1)
        if "r_stance_enc" in df.columns and "b_stance_enc" in df.columns:
            df["stance_matchup"] = df["r_stance_enc"].astype(str) + "_" + df["b_stance_enc"].astype(str)
            df["same_stance"]    = (df["r_stance_enc"] == df["b_stance_enc"]).astype(int)

        # ── TIER 9: Interaction features ───────────────────────────────────────
        if "elo_diff" in df.columns and "diff_finish_rate" in df.columns:
            df["elo_x_finish_rate"] = df["elo_diff"] * df["diff_finish_rate"]
        if "diff_pre_win_streak" in df.columns and "diff_pre_finish_rate" in df.columns:
            df["streak_x_finish"] = df["diff_pre_win_streak"] * df["diff_pre_finish_rate"]
        if "diff_pre_SLpM" in df.columns and "diff_pre_SApM" in df.columns:
            df["striking_exchange"] = df["diff_pre_SLpM"] - df["diff_pre_SApM"]
        if "diff_pre_td_avg" in df.columns and "diff_pre_td_acc" in df.columns:
            df["td_efficiency"] = df["diff_pre_td_avg"] * df["diff_pre_td_acc"]
        if "diff_pre_sig_str_acc" in df.columns and "diff_pre_ctrl_avg" in df.columns:
            df["control_accuracy"] = df["diff_pre_sig_str_acc"] * df["diff_pre_ctrl_avg"]

        # ── TIER 10: Polynomial features ───────────────────────────────────────
        poly_cols = ["elo_diff", "glicko_diff", "diff_win_loss_ratio"]
        for col in poly_cols:
            if col in df.columns:
                df[f"{col}_sq"]  = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()

        # ── TIER 11: Momentum indicators ──────────────────────────────────────
        if "r_rolling3_wins" in df.columns and "b_rolling3_wins" in df.columns:
            df["momentum_diff_3"] = df["r_rolling3_wins"] - df["b_rolling3_wins"]
        if "r_rolling5_wins" in df.columns and "b_rolling5_wins" in df.columns:
            df["momentum_diff_5"] = df["r_rolling5_wins"] - df["b_rolling5_wins"]
        if "r_pre_win_streak" in df.columns:
            df["streak_differential"] = df.get("r_pre_win_streak", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_win_streak", pd.Series(0.0, index=df.index)).fillna(0)

        # ── TIER 12: Method-specific features ──────────────────────────────────
        if "r_pre_ko_wins" in df.columns:
            df["ko_threat_diff"] = df.get("r_pre_ko_wins", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_ko_wins", pd.Series(0.0, index=df.index)).fillna(0)
        if "r_pre_sub_wins" in df.columns:
            df["sub_threat_diff"] = df.get("r_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0)
        if "r_pre_dec_wins" in df.columns:
            df["dec_tendency_diff"] = df.get("r_pre_dec_wins", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_dec_wins", pd.Series(0.0, index=df.index)).fillna(0)
        if "r_pre_finish_rate" in df.columns:
            df["r_finishing_tendency"] = df.get("r_pre_finish_rate", pd.Series(0.0, index=df.index)).fillna(0)
            df["b_finishing_tendency"] = df.get("b_pre_finish_rate", pd.Series(0.0, index=df.index)).fillna(0)
            df["finishing_matchup"]    = df["r_finishing_tendency"] * df["b_finishing_tendency"]

        # ── Career patterns ────────────────────────────────────────────────────
        for prefix in ["r", "b"]:
            dec_col   = f"{prefix}_pre_dec_wins"
            ko_col    = f"{prefix}_pre_ko_wins"
            sub_col   = f"{prefix}_pre_sub_wins"
            total_col = f"{prefix}_pre_total_fights"
            if all(c in df.columns for c in [dec_col, ko_col, sub_col, total_col]):
                denom = df[total_col].clip(lower=1)
                df[f"{prefix}_decision_win_rate"] = df[dec_col]  / denom
                df[f"{prefix}_ko_win_rate"]        = df[ko_col]   / denom
                df[f"{prefix}_sub_win_rate"]       = df[sub_col]  / denom
                df[f"{prefix}_finish_rate"]        = (df[ko_col] + df[sub_col]) / denom
            if f"{prefix}_pre_title_fights" in df.columns:
                df[f"{prefix}_title_fight_exp"] = df[f"{prefix}_pre_title_fights"]
            if f"{prefix}_pre_five_round_fights" in df.columns:
                df[f"{prefix}_main_event_exp"] = df[f"{prefix}_pre_five_round_fights"]
        for feat in ["decision_win_rate", "ko_win_rate", "sub_win_rate",
                     "finish_rate", "title_fight_exp", "main_event_exp"]:
            r_col = f"r_{feat}"
            b_col = f"b_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"{feat}_diff"] = df[r_col] - df[b_col]

        # ── TIER 12b: Advanced combat metrics ──────────────────────────────────
        for pfx in ["r", "b"]:
            col = f"{pfx}_days_since_last"
            if col in df.columns:
                days = df[col].fillna(365)
                df[f"{pfx}_ring_rust"] = np.where(days > 365, -0.15,
                                          np.where(days > 180, -0.08,
                                          np.where(days > 90, -0.03, 0.0)))
            else:
                df[f"{pfx}_ring_rust"] = 0.0
        if "r_ring_rust" in df.columns and "b_ring_rust" in df.columns:
            df["ring_rust_diff"] = df["r_ring_rust"] - df["b_ring_rust"]

        wc_map = {
            "Heavyweight": 1.4, "Light Heavyweight": 1.2, "Middleweight": 1.0,
            "Welterweight": 0.95, "Lightweight": 0.9, "Featherweight": 0.85,
            "Bantamweight": 0.8, "Flyweight": 0.75,
            "Women's Bantamweight": 0.8, "Women's Flyweight": 0.75,
            "Women's Strawweight": 0.7, "Women's Featherweight": 0.72,
        }
        if "weight_class" in df.columns:
            df["weight_class_ko_factor"] = df["weight_class"].map(wc_map).fillna(1.0)
        else:
            df["weight_class_ko_factor"] = 1.0

        r_slpm = df.get("r_pre_SLpM", df.get("r_pro_SLpM", pd.Series(0.0, index=df.index))).fillna(0)
        b_slpm = df.get("b_pre_SLpM", df.get("b_pro_SLpM", pd.Series(0.0, index=df.index))).fillna(0)
        r_td   = df.get("r_pre_td_avg", df.get("r_pro_td_avg", pd.Series(0.0, index=df.index))).fillna(0)
        b_td   = df.get("b_pre_td_avg", df.get("b_pro_td_avg", pd.Series(0.0, index=df.index))).fillna(0)
        df["style_clash_severity"] = np.abs((r_slpm - b_slpm) * 0.5 + (r_td - b_td) * 0.3)

        r_form   = df.get("r_recent_form_3",   pd.Series(0.5, index=df.index)).fillna(0.5)
        b_form   = df.get("b_recent_form_3",   pd.Series(0.5, index=df.index)).fillna(0.5)
        r_streak = df.get("r_win_streak",       pd.Series(0.0, index=df.index)).fillna(0)
        b_streak = df.get("b_win_streak",       pd.Series(0.0, index=df.index)).fillna(0)
        r_age    = df.get("r_age_at_event",     pd.Series(28.0, index=df.index)).fillna(28)
        b_age    = df.get("b_age_at_event",     pd.Series(28.0, index=df.index)).fillna(28)
        r_fights = df.get("r_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(10)
        b_fights = df.get("b_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(10)
        df["upset_potential"] = ((b_form - r_form) * 0.4 +
                                  (b_streak - r_streak) * 0.1 +
                                  (b_age - r_age) * 0.01 +
                                  (b_fights - r_fights) * 0.005)

        r_acc = df.get("r_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(0.45)
        b_acc = df.get("b_pre_sig_str_acc", pd.Series(0.45, index=df.index)).fillna(0.45)
        df["power_vs_technique"] = (r_slpm - b_slpm) * 0.6 + (r_acc - b_acc) * 0.4

        is_title_col   = df.get("is_title_bout",        pd.Series(0.0, index=df.index)).fillna(0)
        title_exp_diff = df.get("title_fight_exp_diff", pd.Series(0.0, index=df.index)).fillna(0)
        df["championship_pressure"] = title_exp_diff * is_title_col * (r_form - b_form)

        r_clinch_pct = df.get("r_pre_clinch_pct", pd.Series(0.0, index=df.index)).fillna(0)
        b_clinch_pct = df.get("b_pre_clinch_pct", pd.Series(0.0, index=df.index)).fillna(0)
        df["r_clinch_effectiveness"]    = r_clinch_pct * r_slpm * r_acc
        df["b_clinch_effectiveness"]    = b_clinch_pct * b_slpm * b_acc
        df["clinch_effectiveness_diff"] = df["r_clinch_effectiveness"] - df["b_clinch_effectiveness"]

        r_dec            = df.get("r_decision_win_rate", pd.Series(0.0, index=df.index)).fillna(0)
        b_dec            = df.get("b_decision_win_rate", pd.Series(0.0, index=df.index)).fillna(0)
        total_rounds_col = df.get("total_rounds",        pd.Series(3.0, index=df.index)).fillna(3)
        df["five_round_cardio_advantage"] = (r_dec - b_dec) * (total_rounds_col / 3)

        r_kd_absorbed = df.get("r_pre_kd_absorbed", pd.Series(0.0, index=df.index)).fillna(0)
        b_kd_absorbed = df.get("b_pre_kd_absorbed", pd.Series(0.0, index=df.index)).fillna(0)
        r_fights_safe  = r_fights.clip(lower=1)
        b_fights_safe  = b_fights.clip(lower=1)
        df["r_chin_deterioration"]    = r_kd_absorbed / r_fights_safe
        df["b_chin_deterioration"]    = b_kd_absorbed / b_fights_safe
        df["chin_deterioration_diff"] = df["r_chin_deterioration"] - df["b_chin_deterioration"]

        r_ko_r  = df.get("r_ko_win_rate",          pd.Series(0.0, index=df.index)).fillna(0)
        b_ko_r  = df.get("b_ko_win_rate",          pd.Series(0.0, index=df.index)).fillna(0)
        r_sub_r = df.get("r_sub_win_rate",         pd.Series(0.0, index=df.index)).fillna(0)
        b_sub_r = df.get("b_sub_win_rate",         pd.Series(0.0, index=df.index)).fillna(0)
        r_fin_r = df.get("r_recent_finish_rate_3", pd.Series(0.0, index=df.index)).fillna(0)
        b_fin_r = df.get("b_recent_finish_rate_3", pd.Series(0.0, index=df.index)).fillna(0)
        df["finishing_pressure_diff"] = ((r_ko_r - b_ko_r) * 0.5 +
                                          (r_sub_r - b_sub_r) * 0.3 +
                                          (r_fin_r - b_fin_r) * 0.2)

        r_days_col = df.get("r_days_since_last", pd.Series(180.0, index=df.index)).fillna(180)
        b_days_col = df.get("b_days_since_last", pd.Series(180.0, index=df.index)).fillna(180)
        df["r_overactive"]    = (r_days_col < 60).astype(float)
        df["b_overactive"]    = (b_days_col < 60).astype(float)
        df["overactivity_diff"] = df["r_overactive"] - df["b_overactive"]

        # ── TIER 12c: Opponent quality & trajectory ───────────────────────────
        if "r_avg_opp_elo_L5" in df.columns and "b_avg_opp_elo_L5" in df.columns:
            df["opp_quality_diff"] = df["r_avg_opp_elo_L5"] - df["b_avg_opp_elo_L5"]
        if "r_trajectory_3" in df.columns and "b_trajectory_3" in df.columns:
            df["trajectory_diff"] = df["r_trajectory_3"] - df["b_trajectory_3"]

        # ── TIER 12d: Fighter-at-peak score ──────────────────────────────────
        for prefix in ["r", "b"]:
            age_s    = df.get(f"{prefix}_age_at_event",   pd.Series(28.0, index=df.index)).fillna(28)
            fights_s = df.get(f"{prefix}_pre_total_fights",pd.Series(10.0, index=df.index)).fillna(10)
            form3_s  = df.get(f"{prefix}_rolling3_wins",  pd.Series(1.5, index=df.index)).fillna(1.5)
            age_peak  = (1.0 - np.abs(age_s - 29.5) / 10.0).clip(0, 1)
            exp_score = (fights_s.clip(1, 25) / 25.0)
            form_score = (form3_s / 3.0).clip(0, 1)
            df[f"{prefix}_peak_score"] = (age_peak * 0.4 + exp_score * 0.3 + form_score * 0.3)
        if "r_peak_score" in df.columns and "b_peak_score" in df.columns:
            df["peak_score_diff"] = df["r_peak_score"] - df["b_peak_score"]

        # ── TIER 12e: Uncertainty score ───────────────────────────────────────
        r_slpm_u = df.get("r_pre_SLpM",         pd.Series(3.0, index=df.index)).fillna(3.0)
        b_slpm_u = df.get("b_pre_SLpM",         pd.Series(3.0, index=df.index)).fillna(3.0)
        r_td_u   = df.get("r_pre_td_avg",        pd.Series(1.5, index=df.index)).fillna(1.5)
        b_td_u   = df.get("b_pre_td_avg",        pd.Series(1.5, index=df.index)).fillna(1.5)
        r_nf_u   = df.get("r_pre_total_fights",  pd.Series(5.0, index=df.index)).fillna(5)
        b_nf_u   = df.get("b_pre_total_fights",  pd.Series(5.0, index=df.index)).fillna(5)
        elo_d_u  = df.get("elo_diff",            pd.Series(0.0, index=df.index)).fillna(0).abs()
        str_sim  = 1.0 - (np.abs(r_slpm_u - b_slpm_u) / (r_slpm_u + b_slpm_u + 1e-6)).clip(0, 1)
        gr_sim   = 1.0 - (np.abs(r_td_u - b_td_u) / (r_td_u + b_td_u + 1e-6)).clip(0, 1)
        exp_unc  = 1.0 - ((r_nf_u + b_nf_u) / 60.0).clip(0, 1)
        elo_unc  = 1.0 - (elo_d_u / 400.0).clip(0, 1)
        df["uncertainty_score"] = (str_sim * 0.25 + gr_sim * 0.25 + exp_unc * 0.25 + elo_unc * 0.25)

        # ── TIER 13: SVD — apply transform only (do NOT re-fit) ───────────────
        if self.svd_fitted:
            def _apply_svd_transform(fitted_cols, svd_obj, prefix):
                if not fitted_cols:
                    return
                try:
                    # Build X using exactly the columns the SVD was fitted on;
                    # fill any column absent from df (e.g. at prediction time) with zeros.
                    X = np.column_stack([
                        df[c].fillna(0).values if c in df.columns else np.zeros(len(df))
                        for c in fitted_cols
                    ])
                    X_svd = svd_obj.transform(X)
                    for i in range(X_svd.shape[1]):
                        df[f"{prefix}_svd_{i}"] = X_svd[:, i]
                except Exception:
                    pass

            _apply_svd_transform(self.svd_striking_cols,  self.svd_striking,  "striking")
            _apply_svd_transform(self.svd_grappling_cols, self.svd_grappling, "grappling")
            _apply_svd_transform(self.svd_physical_cols,  self.svd_physical,  "physical")
            _apply_svd_transform(self.svd_form_cols,      self.svd_form,      "form")

        # ── Encode title bout, total_rounds, gender ───────────────────────────
        if "is_title_bout" in df.columns:
            df["is_title_enc"] = df["is_title_bout"].astype(int)
        if "total_rounds" in df.columns:
            df["total_rounds_num"] = pd.to_numeric(df["total_rounds"], errors="coerce").fillna(3)
        if "gender" in df.columns:
            df["gender_enc"] = (df["gender"].fillna("").str.lower() == "women").astype(int)

        # ── TIER 14: Positional & Target Differentials ────────────────────────
        for feat in ["distance_pct", "clinch_pct", "ground_pct", "head_pct", "body_pct", "leg_pct"]:
            r_col, b_col = f"r_pre_{feat}", f"b_pre_{feat}"
            if r_col in df.columns and b_col in df.columns:
                df[f"diff_{feat}"] = df[r_col].fillna(0) - df[b_col].fillna(0)
        df["positional_striking_advantage"] = (
            df.get("diff_distance_pct", pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
            df.get("diff_clinch_pct",   pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
            df.get("diff_ground_pct",   pd.Series(0.0, index=df.index)).fillna(0.0).abs()
        )
        df["target_distribution_advantage"] = (
            df.get("diff_head_pct", pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
            df.get("diff_body_pct", pd.Series(0.0, index=df.index)).fillna(0.0).abs() +
            df.get("diff_leg_pct",  pd.Series(0.0, index=df.index)).fillna(0.0).abs()
        )

        # ── TIER 15: Defense Differentials ────────────────────────────────────
        df["diff_str_def"] = df.get("r_pre_str_def", pd.Series(0.0, index=df.index)).fillna(0) - \
                             df.get("b_pre_str_def", pd.Series(0.0, index=df.index)).fillna(0)
        df["diff_td_def"]  = df.get("r_pre_td_def",  pd.Series(0.0, index=df.index)).fillna(0) - \
                             df.get("b_pre_td_def",  pd.Series(0.0, index=df.index)).fillna(0)
        df["defensive_composite"] = df["diff_str_def"].fillna(0) + df["diff_td_def"].fillna(0)

        # ── TIER 16: Deep Interaction Features ────────────────────────────────
        elo_d    = df.get("elo_diff",            pd.Series(0.0, index=df.index)).fillna(0)
        form3    = df.get("diff_rolling3_wins",  pd.Series(0.0, index=df.index)).fillna(0)
        wlr      = df.get("diff_win_loss_ratio", pd.Series(0.0, index=df.index)).fillna(0)
        fin_r    = df.get("diff_finish_rate",    pd.Series(0.0, index=df.index)).fillna(0)
        kd_abs   = df.get("diff_kd_absorbed",    pd.Series(0.0, index=df.index)).fillna(0)

        df["elo_x_form"]       = elo_d * form3
        df["elo_x_win_ratio"]  = elo_d * wlr
        df["elo_x_finish"]     = elo_d * fin_r
        df["elo_x_durability"] = elo_d * kd_abs.abs()

        reach_d  = df.get("diff_reach",            pd.Series(0.0, index=df.index)).fillna(0)
        height_d = df.get("diff_height",           pd.Series(0.0, index=df.index)).fillna(0)
        slpm_d   = df.get("diff_pre_SLpM",         pd.Series(0.0, index=df.index)).fillna(0)
        td_d     = df.get("diff_pre_td_avg",       pd.Series(0.0, index=df.index)).fillna(0)
        acc_d    = df.get("diff_pre_sig_str_acc",  pd.Series(0.0, index=df.index)).fillna(0)
        age_d    = df.get("diff_age_at_event",     pd.Series(0.0, index=df.index)).fillna(0)
        streak_d = df.get("diff_pre_win_streak",   pd.Series(0.0, index=df.index)).fillna(0)
        exp_gap  = df.get("diff_pre_total_fights", pd.Series(0.0, index=df.index)).fillna(0)

        df["reach_x_striking"]    = reach_d * slpm_d
        df["height_x_reach"]      = height_d * reach_d
        df["physical_x_striking"] = (height_d + reach_d) * slpm_d

        df["age_x_striking"]   = age_d * slpm_d
        df["age_x_grappling"]  = age_d * td_d
        df["age_x_durability"] = age_d * kd_abs.abs()
        df["age_x_win_streak"] = age_d * streak_d
        df["experience_x_age"] = exp_gap * age_d

        str_def_d  = df.get("diff_str_def",   pd.Series(0.0, index=df.index)).fillna(0)
        td_def_d   = df.get("diff_td_def",    pd.Series(0.0, index=df.index)).fillna(0)
        sub_rate_d = df.get("sub_threat_diff", pd.Series(0.0, index=df.index)).fillna(0)
        df["td_x_defense"]           = td_d * td_def_d
        df["submission_x_grappling"] = sub_rate_d * td_d

        df["striking_x_accuracy"] = slpm_d * acc_d
        df["striking_x_defense"]  = slpm_d * str_def_d
        df["ko_power_x_striking"] = df.get("ko_threat_diff", pd.Series(0.0, index=df.index)).fillna(0) * slpm_d

        momentum = df.get("momentum_diff_3", pd.Series(0.0, index=df.index)).fillna(0)
        df["momentum_x_win_streak"] = momentum * streak_d
        df["form_x_experience"]     = form3 * exp_gap
        df["finish_x_momentum"]     = fin_r * momentum
        df["form_x_durability"]     = form3 * kd_abs.abs()

        df["elite_finisher"]     = elo_d * fin_r * form3
        df["unstoppable_streak"] = streak_d * momentum * form3
        df["veteran_advantage"]  = wlr * exp_gap * (-age_d)

        # ── TIER 17: Extended Polynomial Features ─────────────────────────────
        poly_extended = [
            "elo_diff", "glicko_diff", "diff_win_loss_ratio", "diff_age_at_event",
            "diff_reach", "diff_height", "diff_pre_SLpM", "diff_pre_sig_str_acc",
            "diff_pre_td_avg", "diff_pre_win_streak", "diff_finish_rate",
            "diff_pre_loss_streak", "diff_str_def", "diff_td_def",
            "diff_pre_kd_rate", "diff_pre_ctrl_avg",
            "elo_x_form", "streak_x_finish", "striking_exchange",
            "diff_distance_pct", "diff_clinch_pct", "diff_ground_pct",
        ]
        for col in poly_extended:
            if col in df.columns:
                df[f"{col}_sq"]  = df[col] ** 2
                df[f"{col}_abs"] = df[col].abs()
        if "diff_age_at_event" in df.columns:
            df["diff_age_cubed"] = df["diff_age_at_event"] ** 3

        # ── TIER 18: Opponent-Adjusted Performance ────────────────────────────
        if "r_vs_elite_win_rate" in df.columns and "b_vs_elite_win_rate" in df.columns:
            df["diff_win_rate_vs_elite"]     = df["r_vs_elite_win_rate"]    - df["b_vs_elite_win_rate"]
            df["diff_win_rate_vs_strikers"]  = df.get("r_vs_striker_win_rate",  pd.Series(0.0, index=df.index)).fillna(0) - \
                                               df.get("b_vs_striker_win_rate",  pd.Series(0.0, index=df.index)).fillna(0)
            df["diff_win_rate_vs_grapplers"] = df.get("r_vs_grappler_win_rate", pd.Series(0.0, index=df.index)).fillna(0) - \
                                               df.get("b_vs_grappler_win_rate", pd.Series(0.0, index=df.index)).fillna(0)
            df["championship_readiness"]     = df["diff_win_rate_vs_elite"] * df.get("elo_diff", pd.Series(0.0, index=df.index)).fillna(0.0)

        # ── TIER 19: Career Pattern Features ──────────────────────────────────
        if "r_pre_early_finish_rate" in df.columns:
            df["diff_early_finish_rate"]   = df.get("r_pre_early_finish_rate", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_early_finish_rate", pd.Series(0.0, index=df.index)).fillna(0)
            df["diff_late_finish_rate"]    = df.get("r_pre_late_finish_rate",    pd.Series(0.0, index=df.index)).fillna(0) - \
                                             df.get("b_pre_late_finish_rate",    pd.Series(0.0, index=df.index)).fillna(0)
            df["diff_first_round_ko_rate"] = df.get("r_pre_first_round_ko_rate", pd.Series(0.0, index=df.index)).fillna(0) - \
                                             df.get("b_pre_first_round_ko_rate", pd.Series(0.0, index=df.index)).fillna(0)
        if "r_pre_five_round_fights" in df.columns:
            df["diff_five_round_fights"] = df.get("r_pre_five_round_fights", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_pre_five_round_fights", pd.Series(0.0, index=df.index)).fillna(0)

        age_r_t19 = df.get("r_age_at_event", pd.Series(28.0, index=df.index)).fillna(28)
        age_b_t19 = df.get("b_age_at_event", pd.Series(28.0, index=df.index)).fillna(28)
        df["r_prime_score"]       = (1.0 - np.abs(age_r_t19 - 29.5) / 10.0).clip(0, 1)
        df["b_prime_score"]       = (1.0 - np.abs(age_b_t19 - 29.5) / 10.0).clip(0, 1)
        df["prime_years_advantage"] = df["r_prime_score"] - df["b_prime_score"]

        if "r_fights_since_peak" in df.columns:
            df["diff_fights_since_peak"] = df.get("r_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0) - df.get("b_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0)
            df["declining_phase_diff"]   = df["diff_fights_since_peak"]

        if "r_last_fight_was_win" in df.columns:
            df["r_last_fight_momentum"] = df.get("r_last_fight_was_win", pd.Series(0.0, index=df.index)).fillna(0).astype(float) + df.get("r_last_fight_was_finish", pd.Series(0.0, index=df.index)).fillna(0).astype(float)
            df["b_last_fight_momentum"] = df.get("b_last_fight_was_win", pd.Series(0.0, index=df.index)).fillna(0).astype(float) + df.get("b_last_fight_was_finish", pd.Series(0.0, index=df.index)).fillna(0).astype(float)
            df["last_fight_momentum_diff"] = df["r_last_fight_momentum"] - df["b_last_fight_momentum"]

        for feat in ["wins", "sig_str", "td", "kd", "finishes"]:
            r_col10 = f"r_rolling10_{feat}"
            b_col10 = f"b_rolling10_{feat}"
            if r_col10 in df.columns and b_col10 in df.columns:
                df[f"diff_rolling10_{feat}"] = df[r_col10].fillna(0) - df[b_col10].fillna(0)

        # ── TIER 20: Rounds-Based Strategy ───────────────────────────────────
        total_rds_t20  = df.get("total_rounds_num", pd.Series(3.0, index=df.index)).fillna(3)
        dec_rate_d_t20 = df.get("dec_tendency_diff", pd.Series(0.0, index=df.index)).fillna(0)
        if "diff_finish_rate" in df.columns:
            df["rounds_x_cardio"]      = total_rds_t20 * dec_rate_d_t20
            df["rounds_x_finish_rate"] = (5 - total_rds_t20) * df["diff_finish_rate"].fillna(0)
        kd_abs2_t20 = df.get("chin_deterioration_diff", pd.Series(0.0, index=df.index)).fillna(0)
        df["rounds_x_durability"] = total_rds_t20 * kd_abs2_t20

        # ── TIER 21: Matchup-Specific Features ───────────────────────────────
        r_slpm_t21   = df.get("r_pre_SLpM",          pd.Series(3.0,  index=df.index)).fillna(3.0)
        b_slpm_t21   = df.get("b_pre_SLpM",          pd.Series(3.0,  index=df.index)).fillna(3.0)
        r_acc_t21    = df.get("r_pre_sig_str_acc",    pd.Series(0.45, index=df.index)).fillna(0.45)
        b_acc_t21    = df.get("b_pre_sig_str_acc",    pd.Series(0.45, index=df.index)).fillna(0.45)
        r_str_def_t21= df.get("r_pre_str_def",        pd.Series(0.55, index=df.index)).fillna(0.55)
        b_str_def_t21= df.get("b_pre_str_def",        pd.Series(0.55, index=df.index)).fillna(0.55)
        r_td_t21     = df.get("r_pre_td_avg",         pd.Series(1.5,  index=df.index)).fillna(1.5)
        b_td_t21     = df.get("b_pre_td_avg",         pd.Series(1.5,  index=df.index)).fillna(1.5)
        r_td_def_t21 = df.get("r_pre_td_def",         pd.Series(0.65, index=df.index)).fillna(0.65)
        b_td_def_t21 = df.get("b_pre_td_def",         pd.Series(0.65, index=df.index)).fillna(0.65)
        r_sub_avg_t21= df.get("r_pre_sub_att_rate",   pd.Series(0.3,  index=df.index)).fillna(0.3)
        b_sub_avg_t21= df.get("b_pre_sub_att_rate",   pd.Series(0.3,  index=df.index)).fillna(0.3)
        r_td_acc_t21 = df.get("r_pre_td_acc",         pd.Series(0.4,  index=df.index)).fillna(0.4)
        b_td_acc_t21 = df.get("b_pre_td_acc",         pd.Series(0.4,  index=df.index)).fillna(0.4)
        r_ctrl_t21   = df.get("r_pre_ctrl_avg",       pd.Series(60.0, index=df.index)).fillna(60)
        b_ctrl_t21   = df.get("b_pre_ctrl_avg",       pd.Series(60.0, index=df.index)).fillna(60)
        r_sub_rate_t21 = df.get("r_pre_sub_att_rate", pd.Series(0.3,  index=df.index)).fillna(0.3)
        b_sub_rate_t21 = df.get("b_pre_sub_att_rate", pd.Series(0.3,  index=df.index)).fillna(0.3)

        df["r_striking_vs_b_defense"]    = r_slpm_t21 * (1.0 - b_str_def_t21)
        df["b_striking_vs_r_defense"]    = b_slpm_t21 * (1.0 - r_str_def_t21)
        df["striking_exploitation_diff"] = df["r_striking_vs_b_defense"] - df["b_striking_vs_r_defense"]

        df["r_td_vs_b_td_defense"]  = r_td_t21 * (1.0 - b_td_def_t21)
        df["b_td_vs_r_td_defense"]  = b_td_t21 * (1.0 - r_td_def_t21)
        df["td_exploitation_diff"]  = df["r_td_vs_b_td_defense"] - df["b_td_vs_r_td_defense"]

        df["r_sub_setup_efficiency"]     = r_sub_rate_t21 * r_td_acc_t21
        df["b_sub_setup_efficiency"]     = b_sub_rate_t21 * b_td_acc_t21
        df["sub_setup_diff"]             = df["r_sub_setup_efficiency"] - df["b_sub_setup_efficiency"]
        df["r_sub_threat_vs_td_defense"] = r_sub_avg_t21 * (1.0 - b_td_def_t21)
        df["b_sub_threat_vs_td_defense"] = b_sub_avg_t21 * (1.0 - r_td_def_t21)
        df["sub_threat_vs_defense_diff"] = df["r_sub_threat_vs_td_defense"] - df["b_sub_threat_vs_td_defense"]

        df["r_striking_quality"]       = r_slpm_t21 * r_acc_t21
        df["b_striking_quality"]       = b_slpm_t21 * b_acc_t21
        df["striking_quality_diff"]    = df["r_striking_quality"] - df["b_striking_quality"]
        df["r_accuracy_under_fire"]    = r_acc_t21 / (b_slpm_t21 + 0.1)
        df["b_accuracy_under_fire"]    = b_acc_t21 / (r_slpm_t21 + 0.1)
        df["accuracy_under_fire_diff"] = df["r_accuracy_under_fire"] - df["b_accuracy_under_fire"]

        # ── TIER 22: Statistical Ratio Features ───────────────────────────────
        r_sapm_t22 = df.get("r_pre_SApM", pd.Series(3.0, index=df.index)).fillna(3.0)
        b_sapm_t22 = df.get("b_pre_SApM", pd.Series(3.0, index=df.index)).fillna(3.0)

        df["r_damage_ratio"]    = r_slpm_t21 / (r_sapm_t22 + 0.1)
        df["b_damage_ratio"]    = b_slpm_t21 / (b_sapm_t22 + 0.1)
        df["damage_ratio_diff"] = df["r_damage_ratio"] - df["b_damage_ratio"]

        df["r_striking_output_quality"]    = r_slpm_t21 * r_acc_t21 / (r_sapm_t22 + 0.1)
        df["b_striking_output_quality"]    = b_slpm_t21 * b_acc_t21 / (b_sapm_t22 + 0.1)
        df["striking_output_quality_diff"] = df["r_striking_output_quality"] - df["b_striking_output_quality"]

        df["r_grappling_quality"]  = r_td_t21 * r_td_acc_t21 * (r_ctrl_t21 / 60.0)
        df["b_grappling_quality"]  = b_td_t21 * b_td_acc_t21 * (b_ctrl_t21 / 60.0)
        df["grappling_quality_diff"] = df["r_grappling_quality"] - df["b_grappling_quality"]

        df["r_total_defense_index"] = r_str_def_t21 * r_td_def_t21
        df["b_total_defense_index"] = b_str_def_t21 * b_td_def_t21
        df["total_defense_diff"]    = df["r_total_defense_index"] - df["b_total_defense_index"]

        df["r_complete_fighter_index"] = (r_slpm_t21 + r_td_t21 + r_sub_avg_t21) * r_str_def_t21 * r_td_def_t21
        df["b_complete_fighter_index"] = (b_slpm_t21 + b_td_t21 + b_sub_avg_t21) * b_str_def_t21 * b_td_def_t21
        df["complete_fighter_diff"]    = df["r_complete_fighter_index"] - df["b_complete_fighter_index"]

        df["r_pressure_index"] = r_slpm_t21 * r_td_t21 * (r_ctrl_t21 / 60.0)
        df["b_pressure_index"] = b_slpm_t21 * b_td_t21 * (b_ctrl_t21 / 60.0)
        df["pressure_index_diff"] = df["r_pressure_index"] - df["b_pressure_index"]

        # ── TIER 23: Extended Statistical Ratio Features ──────────────────────
        r_wins_t23   = df.get("r_pre_wins",        pd.Series(5.0,   index=df.index)).fillna(5.0)
        b_wins_t23   = df.get("b_pre_wins",        pd.Series(5.0,   index=df.index)).fillna(5.0)
        r_losses_t23 = df.get("r_pre_losses",      pd.Series(2.0,   index=df.index)).fillna(2.0)
        b_losses_t23 = df.get("b_pre_losses",      pd.Series(2.0,   index=df.index)).fillna(2.0)
        r_streak_t23 = df.get("r_pre_win_streak",  pd.Series(0.0,   index=df.index)).fillna(0.0)
        b_streak_t23 = df.get("b_pre_win_streak",  pd.Series(0.0,   index=df.index)).fillna(0.0)
        r_fr_t23     = df.get("r_pre_finish_rate", pd.Series(0.4,   index=df.index)).fillna(0.4)
        b_fr_t23     = df.get("b_pre_finish_rate", pd.Series(0.4,   index=df.index)).fillna(0.4)
        r_age_t23    = df.get("r_age_at_event",    pd.Series(28.0,  index=df.index)).fillna(28.0)
        b_age_t23    = df.get("b_age_at_event",    pd.Series(28.0,  index=df.index)).fillna(28.0)
        r_reach_t23  = df.get("r_reach",           pd.Series(71.0,  index=df.index)).fillna(71.0)
        b_reach_t23  = df.get("b_reach",           pd.Series(71.0,  index=df.index)).fillna(71.0)
        r_weight_t23 = df.get("r_weight",          pd.Series(155.0, index=df.index)).fillna(155.0)
        b_weight_t23 = df.get("b_weight",          pd.Series(155.0, index=df.index)).fillna(155.0)
        r_tf_t23     = df.get("r_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(10.0)
        b_tf_t23     = df.get("b_pre_total_fights", pd.Series(10.0, index=df.index)).fillna(10.0)
        df["r_defense_offense_balance"]    = (r_str_def_t21 + 0.01) / (r_acc_t21 + 0.01)
        df["b_defense_offense_balance"]    = (b_str_def_t21 + 0.01) / (b_acc_t21 + 0.01)
        df["defense_offense_balance_diff"] = df["r_defense_offense_balance"] - df["b_defense_offense_balance"]
        df["r_td_defense_offense_balance"]    = (r_td_def_t21 + 0.01) / (r_td_acc_t21 + 0.01)
        df["b_td_defense_offense_balance"]    = (b_td_def_t21 + 0.01) / (b_td_acc_t21 + 0.01)
        df["td_defense_offense_balance_diff"] = df["r_td_defense_offense_balance"] - df["b_td_defense_offense_balance"]
        df["finish_efficiency_diff"] = r_fr_t23 - b_fr_t23
        df["r_precision_striking"] = r_acc_t21 / (r_slpm_t21 + 0.1)
        df["b_precision_striking"] = b_acc_t21 / (b_slpm_t21 + 0.1)
        df["precision_striking_diff"] = df["r_precision_striking"] - df["b_precision_striking"]
        df["r_quality_grappling_23"] = r_td_acc_t21 * (r_td_t21 ** 0.5)
        df["b_quality_grappling_23"] = b_td_acc_t21 * (b_td_t21 ** 0.5)
        df["quality_grappling_diff"] = df["r_quality_grappling_23"] - df["b_quality_grappling_23"]
        df["r_submission_threat_ratio"] = (r_sub_avg_t21 + 0.01) / (r_td_t21 + 0.01)
        df["b_submission_threat_ratio"] = (b_sub_avg_t21 + 0.01) / (b_td_t21 + 0.01)
        df["submission_threat_ratio_diff"] = df["r_submission_threat_ratio"] - df["b_submission_threat_ratio"]
        df["r_damage_absorption_efficiency"] = r_sapm_t22 / (r_str_def_t21 + 0.01)
        df["b_damage_absorption_efficiency"] = b_sapm_t22 / (b_str_def_t21 + 0.01)
        df["damage_absorption_efficiency_diff"] = df["r_damage_absorption_efficiency"] - df["b_damage_absorption_efficiency"]
        df["r_defense_versatility"] = (r_str_def_t21 * r_td_def_t21) ** 0.5
        df["b_defense_versatility"] = (b_str_def_t21 * b_td_def_t21) ** 0.5
        df["defense_versatility_diff"] = df["r_defense_versatility"] - df["b_defense_versatility"]
        df["r_total_offense_index"] = r_slpm_t21 + (r_td_t21 * 1.5)
        df["b_total_offense_index"] = b_slpm_t21 + (b_td_t21 * 1.5)
        df["total_offense_index_diff"] = df["r_total_offense_index"] - df["b_total_offense_index"]
        df["r_offensive_versatility"] = (r_slpm_t21 * r_td_t21) ** 0.5
        df["b_offensive_versatility"] = (b_slpm_t21 * b_td_t21) ** 0.5
        df["offensive_versatility_diff"] = df["r_offensive_versatility"] - df["b_offensive_versatility"]
        df["r_striker_index"] = (r_slpm_t21 + 0.1) / (r_td_t21 + 0.1)
        df["b_striker_index"] = (b_slpm_t21 + 0.1) / (b_td_t21 + 0.1)
        df["striker_index_diff"] = df["r_striker_index"] - df["b_striker_index"]
        r_wlr_t23 = r_wins_t23 / (r_losses_t23 + 1.0)
        b_wlr_t23 = b_wins_t23 / (b_losses_t23 + 1.0)
        df["win_loss_ratio_squared_diff"] = (r_wlr_t23 ** 2) - (b_wlr_t23 ** 2)
        df["r_experience_quality"] = r_wins_t23 / (r_wins_t23 + r_losses_t23 + 1.0)
        df["b_experience_quality"] = b_wins_t23 / (b_wins_t23 + b_losses_t23 + 1.0)
        df["experience_quality_diff"] = df["r_experience_quality"] - df["b_experience_quality"]
        df["r_win_efficiency"] = r_wins_t23 / (r_age_t23 - 18.0 + 1.0)
        df["b_win_efficiency"] = b_wins_t23 / (b_age_t23 - 18.0 + 1.0)
        df["win_efficiency_diff"] = df["r_win_efficiency"] - df["b_win_efficiency"]
        df["r_momentum_quality"] = (r_streak_t23 + 1.0) / (r_wins_t23 + 1.0)
        df["b_momentum_quality"] = (b_streak_t23 + 1.0) / (b_wins_t23 + 1.0)
        df["momentum_quality_diff"] = df["r_momentum_quality"] - df["b_momentum_quality"]
        df["r_reach_efficiency"] = r_slpm_t21 / (r_reach_t23 + 1.0)
        df["b_reach_efficiency"] = b_slpm_t21 / (b_reach_t23 + 1.0)
        df["reach_efficiency_diff"] = df["r_reach_efficiency"] - df["b_reach_efficiency"]
        df["r_size_adjusted_striking"] = r_slpm_t21 / ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_striking"] = b_slpm_t21 / ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_striking_diff"] = df["r_size_adjusted_striking"] - df["b_size_adjusted_striking"]
        df["r_size_adjusted_grappling"] = r_td_t21 / ((r_weight_t23 / 100.0) + 0.01)
        df["b_size_adjusted_grappling"] = b_td_t21 / ((b_weight_t23 / 100.0) + 0.01)
        df["size_adjusted_grappling_diff"] = df["r_size_adjusted_grappling"] - df["b_size_adjusted_grappling"]
        df["r_counter_fighter_index"] = (r_str_def_t21 + 0.1) / (r_slpm_t21 + 1.0)
        df["b_counter_fighter_index"] = (b_str_def_t21 + 0.1) / (b_slpm_t21 + 1.0)
        df["counter_fighter_index_diff"] = df["r_counter_fighter_index"] - df["b_counter_fighter_index"]
        df["r_finishing_threat_composite"] = (r_fr_t23 + 0.1) * (r_sub_avg_t21 + 0.1)
        df["b_finishing_threat_composite"] = (b_fr_t23 + 0.1) * (b_sub_avg_t21 + 0.1)
        df["finishing_threat_composite_diff"] = df["r_finishing_threat_composite"] - df["b_finishing_threat_composite"]
        df["r_complete_geo"] = ((r_slpm_t21 + 1.0) * (r_str_def_t21 + 0.1) * (r_fr_t23 + 0.1)) ** (1.0 / 3.0)
        df["b_complete_geo"] = ((b_slpm_t21 + 1.0) * (b_str_def_t21 + 0.1) * (b_fr_t23 + 0.1)) ** (1.0 / 3.0)
        df["complete_geo_diff"] = df["r_complete_geo"] - df["b_complete_geo"]
        df["r_pressure_fighter_index"] = (r_slpm_t21 + r_td_t21) / (r_str_def_t21 + 0.3)
        df["b_pressure_fighter_index"] = (b_slpm_t21 + b_td_t21) / (b_str_def_t21 + 0.3)
        df["pressure_fighter_index_diff"] = df["r_pressure_fighter_index"] - df["b_pressure_fighter_index"]
        # recent_form_ratio: recent win rate vs career win rate
        _r_roll3 = df.get("r_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(1.5)
        _b_roll3 = df.get("b_pre_rolling3_wins", pd.Series(1.5, index=df.index)).fillna(1.5)
        _r_cwr = r_wins_t23 / (r_tf_t23 + 1.0)
        _b_cwr = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_recent_form_ratio"] = (_r_roll3 / 3.0 + 0.01) / (_r_cwr + 0.01)
        df["b_recent_form_ratio"] = (_b_roll3 / 3.0 + 0.01) / (_b_cwr + 0.01)
        df["recent_form_ratio_diff"] = df["r_recent_form_ratio"] - df["b_recent_form_ratio"]
        # finish_method_diversity: how many distinct finish methods used
        _r_ko_d  = df.get("r_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_ko_d  = df.get("b_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        _r_sub_d = df.get("r_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_sub_d = df.get("b_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        _r_dec_d = df.get("r_pre_dec_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_dec_d = df.get("b_pre_dec_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["r_finish_method_diversity"] = (_r_ko_d > 0).astype(float) + (_r_sub_d > 0).astype(float) + (_r_dec_d > 0).astype(float)
        df["b_finish_method_diversity"] = (_b_ko_d > 0).astype(float) + (_b_sub_d > 0).astype(float) + (_b_dec_d > 0).astype(float)
        df["finish_method_diversity_diff"] = df["r_finish_method_diversity"] - df["b_finish_method_diversity"]
        # cross_domain_compensation: grappling compensates striking gap
        df["r_cross_domain_compensation"] = np.maximum(0.0, r_td_t21 - 1.5) - np.maximum(0.0, 4.0 - r_slpm_t21)
        df["b_cross_domain_compensation"] = np.maximum(0.0, b_td_t21 - 1.5) - np.maximum(0.0, 4.0 - b_slpm_t21)
        df["cross_domain_compensation_index_diff"] = df["r_cross_domain_compensation"] - df["b_cross_domain_compensation"]

        # ── TIER 24: Additional Matchup-Specific Features ─────────────────────
        df["r_absorption_vuln"]               = r_sapm_t22 / (b_slpm_t21 + 0.1)
        df["b_absorption_vuln"]               = b_sapm_t22 / (r_slpm_t21 + 0.1)
        df["absorption_vulnerability_index_diff"] = df["r_absorption_vuln"] - df["b_absorption_vuln"]
        df["r_combined_def_hole"] = (1.0 - r_str_def_t21) * (1.0 - r_td_def_t21)
        df["b_combined_def_hole"] = (1.0 - b_str_def_t21) * (1.0 - b_td_def_t21)
        df["combined_defensive_hole_diff"] = df["r_combined_def_hole"] - df["b_combined_def_hole"]
        df["r_td_pressure_t24"] = (1.0 - r_td_def_t21) * b_td_t21
        df["b_td_pressure_t24"] = (1.0 - b_td_def_t21) * r_td_t21
        df["td_vulnerability_under_pressure_diff"] = df["r_td_pressure_t24"] - df["b_td_pressure_t24"]
        df["r_strike_pressure_t24"] = (1.0 - r_str_def_t21) * b_slpm_t21
        df["b_strike_pressure_t24"] = (1.0 - b_str_def_t21) * r_slpm_t21
        df["strike_defense_under_volume_diff"] = df["r_strike_pressure_t24"] - df["b_strike_pressure_t24"]
        df["r_ctrl_sub_ratio"] = (r_ctrl_t21 / 60.0) / (r_sub_avg_t21 + 0.1)
        df["b_ctrl_sub_ratio"] = (b_ctrl_t21 / 60.0) / (b_sub_avg_t21 + 0.1)
        df["grappling_control_vs_submission_ratio_diff"] = df["r_ctrl_sub_ratio"] - df["b_ctrl_sub_ratio"]
        df["r_sub_def_necessity"] = b_sub_avg_t21 / (r_td_def_t21 + 0.1)
        df["b_sub_def_necessity"] = r_sub_avg_t21 / (b_td_def_t21 + 0.1)
        df["submission_defense_necessity_diff"] = df["r_sub_def_necessity"] - df["b_sub_def_necessity"]
        df["r_strike_synergy"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5
        df["b_strike_synergy"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5
        df["striking_volume_accuracy_synergy_diff"] = df["r_strike_synergy"] - df["b_strike_synergy"]
        df["r_td_paradox"] = (r_td_acc_t21 + 0.01) / (r_td_t21 + 0.5)
        df["b_td_paradox"] = (b_td_acc_t21 + 0.01) / (b_td_t21 + 0.5)
        df["takedown_efficiency_paradox_diff"] = df["r_td_paradox"] - df["b_td_paradox"]
        df["r_total_off_eff"] = ((r_slpm_t21 * (r_acc_t21 + 0.01)) ** 0.5
                                 + (r_td_t21 * (r_td_acc_t21 + 0.01)) ** 0.5)
        df["b_total_off_eff"] = ((b_slpm_t21 * (b_acc_t21 + 0.01)) ** 0.5
                                 + (b_td_t21 * (b_td_acc_t21 + 0.01)) ** 0.5)
        df["total_offensive_efficiency_index_diff"] = df["r_total_off_eff"] - df["b_total_off_eff"]
        df["r_sg_corr"] = (r_slpm_t21 * (r_acc_t21 + 0.01)) / (r_td_t21 * (r_td_acc_t21 + 0.01) + 0.1)
        df["b_sg_corr"] = (b_slpm_t21 * (b_acc_t21 + 0.01)) / (b_td_t21 * (b_td_acc_t21 + 0.01) + 0.1)
        df["striking_grappling_efficiency_correlation_diff"] = df["r_sg_corr"] - df["b_sg_corr"]
        df["r_def_allocation_balance"] = (r_str_def_t21 - r_td_def_t21).abs()
        df["b_def_allocation_balance"] = (b_str_def_t21 - b_td_def_t21).abs()
        df["defense_allocation_balance_diff"] = df["r_def_allocation_balance"] - df["b_def_allocation_balance"]
        _r_cbt = ((r_slpm_t21 / 10.0 + 0.01) * (r_acc_t21 + 0.01) * (10.0 / (r_sapm_t22 + 0.01))
                  * (r_str_def_t21 + 0.01) * (r_td_t21 / 5.0 + 0.01) * (r_td_acc_t21 + 0.01)
                  * (r_td_def_t21 + 0.01) * (r_sub_avg_t21 / 2.0 + 0.01)) ** (1.0 / 8.0)
        _b_cbt = ((b_slpm_t21 / 10.0 + 0.01) * (b_acc_t21 + 0.01) * (10.0 / (b_sapm_t22 + 0.01))
                  * (b_str_def_t21 + 0.01) * (b_td_t21 / 5.0 + 0.01) * (b_td_acc_t21 + 0.01)
                  * (b_td_def_t21 + 0.01) * (b_sub_avg_t21 / 2.0 + 0.01)) ** (1.0 / 8.0)
        df["r_combat_eff"] = _r_cbt
        df["b_combat_eff"] = _b_cbt
        df["total_combat_efficiency_index_diff"] = _r_cbt - _b_cbt

        # ── TIER 25: Named Composite Features ─────────────────────────────────
        df["net_striking_advantage"] = (r_slpm_t21 - b_slpm_t21) - (r_sapm_t22 - b_sapm_t22)
        df["striker_advantage"]  = (r_slpm_t21 * r_acc_t21)  - (b_slpm_t21 * b_acc_t21)
        df["grappler_advantage"] = (r_td_t21 * r_td_acc_t21) - (b_td_t21 * b_td_acc_t21)
        df["experience_gap"] = r_tf_t23 - b_tf_t23
        r_ko_wins_t25  = df.get("r_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        b_ko_wins_t25  = df.get("b_pre_ko_wins",  pd.Series(0.0, index=df.index)).fillna(0.0)
        r_sub_wins_t25 = df.get("r_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_sub_wins_t25 = df.get("b_pre_sub_wins", pd.Series(0.0, index=df.index)).fillna(0.0)
        r_ko_rate_t25  = r_ko_wins_t25  / (r_tf_t23 + 1.0)
        b_ko_rate_t25  = b_ko_wins_t25  / (b_tf_t23 + 1.0)
        r_sub_rate_t25 = r_sub_wins_t25 / (r_tf_t23 + 1.0)
        b_sub_rate_t25 = b_sub_wins_t25 / (b_tf_t23 + 1.0)
        df["ko_specialist_gap"]         = r_ko_rate_t25  - b_ko_rate_t25
        df["submission_specialist_gap"] = r_sub_rate_t25 - b_sub_rate_t25
        r_elo_t25  = df.get("elo_r", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_elo_t25  = df.get("elo_b", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        r_traj_t25 = df.get("r_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_traj_t25 = df.get("b_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["skill_momentum"] = (r_elo_t25 - b_elo_t25) * (r_traj_t25 - b_traj_t25)
        r_loss_streak_t25 = df.get("r_pre_loss_streak", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_loss_streak_t25 = df.get("b_pre_loss_streak", pd.Series(0.0, index=df.index)).fillna(0.0)
        r_win_rate_t25 = r_wins_t23 / (r_tf_t23 + 1.0)
        b_win_rate_t25 = b_wins_t23 / (b_tf_t23 + 1.0)
        df["r_desperation"] = r_loss_streak_t25 * (1.0 / (r_win_rate_t25 + 0.1))
        df["b_desperation"] = b_loss_streak_t25 * (1.0 / (b_win_rate_t25 + 0.1))
        df["desperation_diff"] = df["r_desperation"] - df["b_desperation"]
        r_days_t25 = df.get("r_days_since_last", pd.Series(180.0, index=df.index)).fillna(180.0)
        b_days_t25 = df.get("b_days_since_last", pd.Series(180.0, index=df.index)).fillna(180.0)
        df["r_freshness"] = np.exp(-((r_days_t25 - 135.0) ** 2) / (2.0 * 90.0 ** 2))
        df["b_freshness"] = np.exp(-((b_days_t25 - 135.0) ** 2) / (2.0 * 90.0 ** 2))
        df["freshness_advantage"] = df["r_freshness"] - df["b_freshness"]

        # ── TIER 26: Stance Directional Features ──────────────────────────────
        _r_st26 = df.get("r_stance", pd.Series("", index=df.index)).fillna("").astype(str).str.strip().str.lower()
        _b_st26 = df.get("b_stance", pd.Series("", index=df.index)).fillna("").astype(str).str.strip().str.lower()
        df["orthodox_vs_southpaw_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "southpaw"), 1.0,
            np.where((_r_st26 == "southpaw") & (_b_st26 == "orthodox"), -1.0, 0.0)
        ).astype(float)
        df["orthodox_vs_switch_advantage"] = np.where(
            (_r_st26 == "orthodox") & (_b_st26 == "switch"), 1.0,
            np.where((_r_st26 == "switch") & (_b_st26 == "orthodox"), -1.0, 0.0)
        ).astype(float)
        df["southpaw_vs_switch_advantage"] = np.where(
            (_r_st26 == "southpaw") & (_b_st26 == "switch"), 1.0,
            np.where((_r_st26 == "switch") & (_b_st26 == "southpaw"), -1.0, 0.0)
        ).astype(float)
        df["mirror_matchup"] = (_r_st26 == _b_st26).astype(float)

        # ── TIER 27: Extended Polynomial Squared Terms ────────────────────────
        def _signed_sq_t27(s):
            return np.sign(s) * (s ** 2)
        for _feat_sq in [
            "net_striking_advantage", "striker_advantage", "grappler_advantage",
            "experience_gap", "ko_specialist_gap", "submission_specialist_gap",
            "skill_momentum", "desperation_diff", "freshness_advantage",
            "combined_defensive_hole_diff", "striking_volume_accuracy_synergy_diff",
            "total_offensive_efficiency_index_diff", "finish_efficiency_diff",
            "defense_versatility_diff", "offensive_versatility_diff",
        ]:
            if _feat_sq in df.columns:
                df[f"{_feat_sq}_sq"] = _signed_sq_t27(df[_feat_sq])

        # ── TIER 28: Volatility & Career Arc Features ──────────────────────────
        r_fr_l5_t28  = df.get("r_pre_finish_rate_l5",  pd.Series(0.4, index=df.index)).fillna(0.4)
        b_fr_l5_t28  = df.get("b_pre_finish_rate_l5",  pd.Series(0.4, index=df.index)).fillna(0.4)
        r_fr_l10_t28 = df.get("r_pre_finish_rate_l10", pd.Series(0.4, index=df.index)).fillna(0.4)
        b_fr_l10_t28 = df.get("b_pre_finish_rate_l10", pd.Series(0.4, index=df.index)).fillna(0.4)
        df["r_finish_rate_accel"]       = r_fr_l5_t28 - r_fr_l10_t28
        df["b_finish_rate_accel"]       = b_fr_l5_t28 - b_fr_l10_t28
        df["finish_rate_acceleration_diff"] = df["r_finish_rate_accel"] - df["b_finish_rate_accel"]
        r_slpm_cv_t28 = df.get("r_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(0.3)
        b_slpm_cv_t28 = df.get("b_pre_slpm_cv", pd.Series(0.3, index=df.index)).fillna(0.3)
        df["slpm_coefficient_of_variation_diff"] = r_slpm_cv_t28 - b_slpm_cv_t28
        r_mil_t28 = df.get("r_pre_mileage_adj_age", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_mil_t28 = df.get("b_pre_mileage_adj_age", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["mileage_adjusted_age_diff"] = r_mil_t28 - b_mil_t28
        df["performance_decline_velocity_diff"] = (
            df.get("r_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0) -
            df.get("b_trajectory_3", pd.Series(0.0, index=df.index)).fillna(0.0)
        ) * (-1.0)
        r_cur_elo_t28 = df.get("elo_r", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_cur_elo_t28 = df.get("elo_b", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        r_peak_t28    = df.get("r_career_elo_peak", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_peak_t28    = df.get("b_career_elo_peak", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        df["r_distance_from_peak"] = r_peak_t28 - r_cur_elo_t28
        df["b_distance_from_peak"] = b_peak_t28 - b_cur_elo_t28
        df["distance_from_career_peak_diff"] = df["r_distance_from_peak"] - df["b_distance_from_peak"]
        r_fsp_t28 = df.get("r_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_fsp_t28 = df.get("b_fights_since_peak", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["r_career_inflection"] = r_fsp_t28 / (r_tf_t23 + 1.0)
        df["b_career_inflection"] = b_fsp_t28 / (b_tf_t23 + 1.0)
        df["career_inflection_point_diff"] = df["r_career_inflection"] - df["b_career_inflection"]
        df["r_prime_exit_risk"] = (r_age_t23 > 33).astype(float) * np.clip(-r_traj_t25, 0.0, 1.0)
        df["b_prime_exit_risk"] = (b_age_t23 > 33).astype(float) * np.clip(-b_traj_t25, 0.0, 1.0)
        df["prime_exit_risk_diff"] = df["r_prime_exit_risk"] - df["b_prime_exit_risk"]
        df["r_aging_power_penalty"] = r_ko_rate_t25 * r_age_t23 * (r_age_t23 > 35).astype(float)
        df["b_aging_power_penalty"] = b_ko_rate_t25 * b_age_t23 * (b_age_t23 > 35).astype(float)
        df["aging_power_striker_penalty_diff"] = df["r_aging_power_penalty"] - df["b_aging_power_penalty"]
        df["r_bayesian_finish"] = (r_ko_wins_t25 + r_sub_wins_t25 + 2.0) / (r_tf_t23 + 4.0)
        df["b_bayesian_finish"] = (b_ko_wins_t25 + b_sub_wins_t25 + 2.0) / (b_tf_t23 + 4.0)
        df["bayesian_finish_rate_diff"] = df["r_bayesian_finish"] - df["b_bayesian_finish"]
        df["r_layoff_veteran"] = r_days_t25 * r_tf_t23
        df["b_layoff_veteran"] = b_days_t25 * b_tf_t23
        df["layoff_veteran_interaction_diff"] = df["r_layoff_veteran"] - df["b_layoff_veteran"]
        df["r_elo_momentum"] = r_cur_elo_t28 * r_traj_t25
        df["b_elo_momentum"] = b_cur_elo_t28 * b_traj_t25
        df["elo_momentum_vs_competition_diff"] = df["r_elo_momentum"] - df["b_elo_momentum"]
        r_avg_opp_elo_t28 = df.get("r_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        b_avg_opp_elo_t28 = df.get("b_avg_opp_elo_L5", pd.Series(1500.0, index=df.index)).fillna(1500.0)
        df["r_title_proximity"] = r_streak_t23 * r_avg_opp_elo_t28 * r_cur_elo_t28 / 1.0e6
        df["b_title_proximity"] = b_streak_t23 * b_avg_opp_elo_t28 * b_cur_elo_t28 / 1.0e6
        df["title_shot_proximity_score_diff"] = df["r_title_proximity"] - df["b_title_proximity"]
        df["r_elo_volatility"] = r_cur_elo_t28 * r_slpm_cv_t28
        df["b_elo_volatility"] = b_cur_elo_t28 * b_slpm_cv_t28
        df["elo_volatility_interaction_diff"] = df["r_elo_volatility"] - df["b_elo_volatility"]
        r_fin_l10_t28 = df.get("r_rolling10_finishes", pd.Series(0.0, index=df.index)).fillna(0.0)
        b_fin_l10_t28 = df.get("b_rolling10_finishes", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["elite_performance_frequency_l10_diff"] = (r_fin_l10_t28 / 10.0) - (b_fin_l10_t28 / 10.0)
        _r_dr28 = df.get("r_damage_ratio", pd.Series(1.0, index=df.index)).fillna(1.0)
        _b_dr28 = df.get("b_damage_ratio", pd.Series(1.0, index=df.index)).fillna(1.0)
        df["r_conf_damage_ratio"] = _r_dr28 * (1.0 - 1.0 / (r_tf_t23 ** 0.5 + 1.0))
        df["b_conf_damage_ratio"] = _b_dr28 * (1.0 - 1.0 / (b_tf_t23 ** 0.5 + 1.0))
        df["confidence_weighted_damage_ratio_diff"] = df["r_conf_damage_ratio"] - df["b_conf_damage_ratio"]
        # recent_vs_career_striking: rolling SLpM relative to career SLpM
        _r_r5slpm = df.get("r_pre_rolling5_slpm", pd.Series(3.0, index=df.index)).fillna(3.0)
        _b_r5slpm = df.get("b_pre_rolling5_slpm", pd.Series(3.0, index=df.index)).fillna(3.0)
        df["r_recent_vs_career_striking"] = _r_r5slpm / (r_slpm_t21 + 0.1)
        df["b_recent_vs_career_striking"] = _b_r5slpm / (b_slpm_t21 + 0.1)
        df["recent_vs_career_striking_diff"] = df["r_recent_vs_career_striking"] - df["b_recent_vs_career_striking"]
        # striking_consistency_ratio: inverse of SLpM std (lower variance = more consistent)
        _r_slpmstd = df.get("r_pre_slpm_std_l10", pd.Series(1.0, index=df.index)).fillna(1.0)
        _b_slpmstd = df.get("b_pre_slpm_std_l10", pd.Series(1.0, index=df.index)).fillna(1.0)
        df["r_striking_consistency_ratio"] = 1.0 / (_r_slpmstd + 0.1)
        df["b_striking_consistency_ratio"] = 1.0 / (_b_slpmstd + 0.1)
        df["striking_consistency_ratio_diff"] = df["r_striking_consistency_ratio"] - df["b_striking_consistency_ratio"]
        # performance_volatility_l10: std of damage ratio over last 10 fights
        _r_drstd = df.get("r_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)).fillna(0.3)
        _b_drstd = df.get("b_pre_damage_ratio_std_l10", pd.Series(0.3, index=df.index)).fillna(0.3)
        df["performance_volatility_l10_diff"] = _r_drstd - _b_drstd
        # tactical_evolution_score: change in distance-based fighting style (recent vs career)
        _r_tact = df.get("r_pre_tactical_evolution", pd.Series(0.0, index=df.index)).fillna(0.0)
        _b_tact = df.get("b_pre_tactical_evolution", pd.Series(0.0, index=df.index)).fillna(0.0)
        df["tactical_evolution_score_diff"] = _r_tact - _b_tact

        return df

    def _build_estimators(self, X_tr, y_tr, sample_weight=None):
        from sklearn.pipeline import Pipeline as _Pipeline
        from sklearn.feature_selection import SelectPercentile, f_classif

        estimators = []
        sw = np.asarray(sample_weight) if sample_weight is not None else None

        # Class imbalance ratio — used by XGBoost (scale_pos_weight) and CatBoost
        n_pos = max(int(np.sum(y_tr == 1)), 1)
        n_neg = max(int(np.sum(y_tr == 0)), 1)
        spw = n_neg / n_pos  # >1 means red wins are minority (typical UFC dataset)
        print_metric("Class ratio (neg/pos):", f"{spw:.3f}")

        # LogisticRegression — SelectPercentile reduces 1200 decomposed features to
        # the top 40% (~480) by ANOVA F-score before the linear model sees them.
        # Linear models degrade badly in very high-dimensional noisy spaces.
        lr_pipe = _Pipeline([
            ("scaler", StandardScaler()),
            ("selector", SelectPercentile(f_classif, percentile=40)),
            ("lr", LogisticRegression(
                C=0.1, max_iter=1000, random_state=RANDOM_SEED,
                n_jobs=SAFE_N_JOBS, class_weight="balanced", solver="lbfgs",
            ))
        ])
        # LR Pipeline does not support sample_weight through Pipeline.fit in all
        # sklearn versions — fit without it; it uses class_weight="balanced" instead.
        lr_pipe.fit(X_tr, y_tr)
        estimators.append(("lr", lr_pipe))

        # XGBoost — tighter regularization defaults (Optuna overrides if available).
        # min_child_weight=15 ensures each leaf spans at least 15 fights' gradients
        # (on ~600 augmented training samples this prevents single-fight leaf nodes).
        # reg_lambda=4.0 adds stronger L2 shrinkage.
        if HAS_XGB:
            xgb_params = {
                "n_estimators": 400, "max_depth": 5, "learning_rate": 0.05,
                "subsample": 0.8, "colsample_bytree": 0.8,
                "min_child_weight": 15, "gamma": 1.5,
                "reg_alpha": 1.5, "reg_lambda": 4.0,
                "scale_pos_weight": spw,
                "random_state": RANDOM_SEED, "eval_metric": "logloss",
                "verbosity": 0, "n_jobs": SAFE_N_JOBS,
            }
            if hasattr(self, '_optuna_best_xgb_params') and self._optuna_best_xgb_params:
                xgb_params.update(self._optuna_best_xgb_params)
                xgb_params.setdefault("scale_pos_weight", spw)  # keep balance if Optuna didn't set it
            if self.gpu_info.get("xgb"):
                xgb_params["device"] = "cuda"
            clf = xgb.XGBClassifier(**xgb_params)
            clf.fit(X_tr, y_tr, sample_weight=sw)
            estimators.append(("xgb", clf))

        # LightGBM — Optuna-tuned params used if available, else conservative defaults
        if HAS_LGB:
            lgb_params = {
                "n_estimators": 500, "num_leaves": 31, "max_depth": -1,
                "learning_rate": 0.05, "subsample": 0.8,
                "colsample_bytree": 0.8, "min_child_samples": 30,
                "reg_alpha": 0.5, "reg_lambda": 1.0,
                "random_state": RANDOM_SEED, "verbose": -1,
                "n_jobs": SAFE_N_JOBS, "class_weight": "balanced",
            }
            if hasattr(self, '_optuna_best_lgb_params') and self._optuna_best_lgb_params:
                lgb_params.update(self._optuna_best_lgb_params)
            clf = lgb.LGBMClassifier(**lgb_params)
            clf.fit(X_tr, y_tr, sample_weight=sw)
            estimators.append(("lgb", clf))

        # Random Forest — constrained to prevent memorisation.
        # With 600 features and ~5700 augmented training samples, depth=12 and
        # min_samples_leaf=3 still yielded ~96% in-sample accuracy.
        # Increasing min_samples_leaf to 8 forces each leaf to cover at least
        # 8 samples, significantly reducing overfitting while keeping the
        # model deep enough to capture real interactions.
        rf = RandomForestClassifier(
            n_estimators=300, max_depth=10, min_samples_split=10,
            min_samples_leaf=8, max_features="sqrt",
            random_state=RANDOM_SEED, n_jobs=1, class_weight="balanced"
        )
        rf.fit(X_tr, y_tr, sample_weight=sw)
        estimators.append(("rf", rf))

        # MLP removed: its meta-learner coefficient was consistently negative,
        # meaning the stacking LR was betting against its predictions. MLPs
        # require careful tuning (learning rate, depth, regularization) on
        # tabular data and are not worth the noise in this ensemble without it.

        # CatBoost — more conservative regularization for small UFC dataset.
        # Original depth=6, iterations=500, l2=3 overfit on ~600 training fights;
        # depth=5, iterations=300, l2=8 prevents memorization.
        if HAS_CAT:
            cat_params = {
                "iterations": 300, "depth": 5, "learning_rate": 0.05,
                "l2_leaf_reg": 8, "bagging_temperature": 0.8,
                "random_strength": 1.5,
                "random_seed": RANDOM_SEED, "verbose": 0,
                "eval_metric": "Logloss",
                "auto_class_weights": "Balanced",
            }
            if self.gpu_info.get("cat"):
                cat_params["task_type"] = "GPU"
            clf = cb.CatBoostClassifier(**cat_params)
            clf.fit(X_tr, y_tr, sample_weight=sw)
            estimators.append(("cat", clf))

        return estimators

    def _train_method_clf(self, df_tr, df_val):
        """Train a 6-class winner+method predictor: Red_Decision, Red_KO/TKO, Red_Submission,
        Blue_Decision, Blue_KO/TKO, Blue_Submission."""
        def encode_winner_method(row):
            winner = row.get('winner', '')
            method = row.get('method', '')
            if winner not in ('Red', 'Blue') or pd.isna(method):
                return None
            return f"{winner}_{method}"

        df_tr = df_tr.copy()
        df_val = df_val.copy()
        df_tr["winner_method"] = df_tr.apply(encode_winner_method, axis=1)
        df_val["winner_method"] = df_val.apply(encode_winner_method, axis=1)

        # Drop rows where winner_method is None
        df_tr = df_tr[df_tr["winner_method"].notna()].copy()
        df_val = df_val[df_val["winner_method"].notna()].copy()

        # Store sorted class labels
        all_classes = sorted(df_tr["winner_method"].unique().tolist())
        self._method_classes = all_classes
        n_classes = len(all_classes)

        # Encode labels as integers
        class_to_int = {c: i for i, c in enumerate(all_classes)}
        df_tr["method_enc"] = df_tr["winner_method"].map(class_to_int)
        df_val["method_enc"] = df_val["winner_method"].map(class_to_int)

        feat_cols = [c for c in self.feature_cols if c in df_tr.columns]
        X_tr = df_tr[feat_cols].fillna(0).replace([np.inf, -np.inf], 0).values
        X_val = df_val[feat_cols].fillna(0).replace([np.inf, -np.inf], 0).values
        y_tr = df_tr["method_enc"].values
        y_val = df_val["method_enc"].values

        X_tr_sel = self.method_scaler.fit_transform(X_tr)
        X_val_sel = self.method_scaler.transform(X_val)

        # Phase 5: Method ensemble with stacking
        method_estimators = []

        if HAS_XGB:
            xgb_m_params = {
                "objective": "multi:softprob", "num_class": n_classes,
                "n_estimators": 500, "learning_rate": 0.05, "max_depth": 5,
                "eval_metric": "mlogloss", "random_state": 42,
                "verbosity": 0, "n_jobs": SAFE_N_JOBS,
            }
            if self.gpu_info.get("xgb"):
                xgb_m_params["device"] = "cuda"
            else:
                xgb_m_params["tree_method"] = "hist"
            method_estimators.append(("xgb_m", xgb.XGBClassifier(**xgb_m_params)))

        if HAS_LGB:
            lgb_m_params = {
                "objective": "multiclass", "num_class": n_classes,
                "n_estimators": 500, "learning_rate": 0.05, "max_depth": 5,
                "random_state": 42, "verbose": -1, "n_jobs": SAFE_N_JOBS,
            }
            method_estimators.append(("lgb_m", lgb.LGBMClassifier(**lgb_m_params)))

        if HAS_CAT:
            cat_m_params = {
                "loss_function": "MultiClass", "iterations": 500,
                "learning_rate": 0.05, "depth": 6, "random_seed": 42, "verbose": 0,
            }
            if self.gpu_info.get("cat"):
                cat_m_params["task_type"] = "GPU"
            method_estimators.append(("cat_m", cb.CatBoostClassifier(**cat_m_params)))

        from sklearn.pipeline import Pipeline as _Pipeline
        from sklearn.neural_network import MLPClassifier
        from sklearn.feature_selection import SelectPercentile, f_classif
        from sklearn.ensemble import RandomForestClassifier as _RFC
        method_estimators.append(("rf_m", _RFC(
            n_estimators=300, max_depth=10, n_jobs=1, random_state=42,
            class_weight="balanced"
        )))
        method_estimators.append(("mlp_m", _Pipeline([
            ("selector", SelectPercentile(f_classif, percentile=40)),
            ("mlp", MLPClassifier(
                hidden_layer_sizes=(256, 128, 64), max_iter=300,
                alpha=5e-4, early_stopping=True, random_state=42
            )),
        ])))

        t0_method = time.time()
        print_step("Training method ensemble (soft voting: XGB + LGB + CAT + RF + MLP)...")
        if len(method_estimators) > 1:
            method_vote = VotingClassifier(estimators=method_estimators, voting="soft", n_jobs=1)
            method_vote.fit(X_tr_sel, y_tr)
            self.method_clf = CalibratedClassifierCV(method_vote, method="isotonic", cv="prefit")
            self.method_clf.fit(X_val_sel, y_val)
        else:
            self.method_clf = method_estimators[0][1]
            self.method_clf.fit(X_tr_sel, y_tr)

        print_metric("Method ensemble time:", f"{time.time()-t0_method:.1f}s")
        method_acc = accuracy_score(y_val, self.method_clf.predict(X_val_sel))
        print_metric("Method (6-class) Val Accuracy:", f"{method_acc:.4f}")
        print_metric("Method classes:", str(self._method_classes))

        # Phase 8: Per-class metrics
        try:
            from sklearn.metrics import classification_report as _cr
            y_method_pred_diag = self.method_clf.predict(X_val_sel)
            print_section("METHOD MODEL PER-CLASS PERFORMANCE")
            print(_cr(y_val, y_method_pred_diag,
                      target_names=self._method_classes, zero_division=0))
        except Exception:
            pass

    # ── PREDICT UPCOMING FIGHTS ──────────────────────────────────────────────
    def predict_upcoming_fights(self, fights):
        """
        fights: list of dicts with keys:
          r_fighter, b_fighter, weight_class, gender, total_rounds
        Returns list of prediction dicts.
        """
        print_section("PREDICTING UPCOMING FIGHTS")
        self.predictions = []

        for fight in fights:
            r_name = fight.get("r_fighter", "")
            b_name = fight.get("b_fighter", "")
            weight_class = fight.get("weight_class", "")
            gender = fight.get("gender", "Men")
            total_rounds = int(fight.get("total_rounds", 3))

            # Exact name lookup — no fuzzy matching. Unknown names mean a debuting
            # fighter with no data, so skip the fight entirely.
            if r_name not in self.all_fighters:
                self._log(f"  Skipping {r_name} vs {b_name} — '{r_name}' not in records (debut?)")
                continue
            if b_name not in self.all_fighters:
                self._log(f"  Skipping {r_name} vs {b_name} — '{b_name}' not in records (debut?)")
                continue

            r_matched = r_name
            b_matched = b_name

            # Build feature vector from fighter histories
            result = self._build_fight_feature_vector(
                r_matched, b_matched, weight_class, gender, total_rounds
            )
            if result is None:
                self._log(f"  Skipping {r_name} vs {b_name} — could not build feature vector")
                continue

            feat_vec, raw_feat_vec = result

            # Scale then apply global selector (same pipeline as training)
            X = np.array([feat_vec])
            X_sel = self.scaler.transform(X)
            if hasattr(self, '_global_selector') and self._global_selector is not None:
                X_sel = self._global_selector.transform(X_sel)

            # Winner prediction
            win_proba = self.stacking_clf.predict_proba(X_sel)[0]
            classes = self.stacking_clf.classes_
            # classes may be [0, 1] where 1=Red wins
            r_idx = list(classes).index(1) if 1 in classes else 1
            b_idx = list(classes).index(0) if 0 in classes else 0
            r_win_prob = float(win_proba[r_idx])
            b_win_prob = float(win_proba[b_idx])

            # Method prediction (6-class) — uses raw (non-decomposed) features
            X_m_sel = self.method_scaler.transform(np.array([raw_feat_vec]))
            method_proba = self.method_clf.predict_proba(X_m_sel)[0]

            # Helper to safely get probability for a class label
            mc = getattr(self, '_method_classes', [])
            def _mp(label):
                try:
                    return float(method_proba[mc.index(label)]) if label in mc else 0.0
                except Exception:
                    return 0.0

            red_dec_prob  = _mp('Red_Decision')
            red_ko_prob   = _mp('Red_KO/TKO')
            red_sub_prob  = _mp('Red_Submission')
            blue_dec_prob = _mp('Blue_Decision')
            blue_ko_prob  = _mp('Blue_KO/TKO')
            blue_sub_prob = _mp('Blue_Submission')

            # Determine predicted winner
            winner_is_red = r_win_prob > b_win_prob
            winner_name = r_name if winner_is_red else b_name
            winner_conf = max(r_win_prob, b_win_prob)

            # Fight-level marginal method probs (sum over both winners)
            # Decision probability is split across Red_Decision + Blue_Decision,
            # so we must aggregate both sides before comparing against KO/Sub.
            total_dec = red_dec_prob + blue_dec_prob
            total_ko  = red_ko_prob  + blue_ko_prob
            total_sub = red_sub_prob + blue_sub_prob
            total_method_sum = total_dec + total_ko + total_sub
            if total_method_sum > 0:
                dec_p = total_dec / total_method_sum
                ko_p  = total_ko  / total_method_sum
                sub_p = total_sub / total_method_sum
            else:
                dec_p = ko_p = sub_p = 1/3

            # Rule-based method rates from each fighter's historical record.
            # These drive the Decision vs Finish split exclusively — the 6-class ML
            # model is too biased toward finishes to be trusted for that split.
            # ML is only used to differentiate KO/TKO vs Submission within finishes.
            def _fighter_method_rates(name):
                r_rows = self.df[self.df['r_fighter'] == name]
                b_rows = self.df[self.df['b_fighter'] == name]
                if len(r_rows) > 0 and len(b_rows) > 0:
                    row_r = r_rows.iloc[-1]
                    row_b = b_rows.iloc[-1]
                    dr = row_r.get('event_date', None)
                    db = row_b.get('event_date', None)
                    if dr is not None and db is not None:
                        row, pfx = (row_r, 'r') if dr >= db else (row_b, 'b')
                    else:
                        row, pfx = row_r, 'r'
                elif len(r_rows) > 0:
                    row, pfx = r_rows.iloc[-1], 'r'
                elif len(b_rows) > 0:
                    row, pfx = b_rows.iloc[-1], 'b'
                else:
                    return 0.30, 0.25, 0.45  # UFC base-rate defaults
                def _g(col): return float(row.get(col, 0) or 0)
                ko  = _g(f'{pfx}_ko_win_rate')
                sub = _g(f'{pfx}_sub_win_rate')
                dec = _g(f'{pfx}_decision_win_rate')
                tot = ko + sub + dec
                if tot > 0:
                    return ko/tot, sub/tot, dec/tot
                return 0.30, 0.25, 0.45

            r_ko_r, r_sub_r, r_dec_r = _fighter_method_rates(r_matched)
            b_ko_r, b_sub_r, b_dec_r = _fighter_method_rates(b_matched)
            rule_ko  = (r_ko_r  + b_ko_r)  / 2
            rule_sub = (r_sub_r + b_sub_r) / 2
            rule_dec = (r_dec_r + b_dec_r) / 2
            rule_tot = rule_ko + rule_sub + rule_dec
            if rule_tot > 0:
                rule_ko  /= rule_tot
                rule_sub /= rule_tot
                rule_dec /= rule_tot

            # Step 1: Decision vs Finish split — rule-based only.
            dec_p = rule_dec
            finish_p = 1.0 - dec_p

            # Step 2: KO vs Sub within finishes — blend ML signal with rule-based.
            # The ML model CAN distinguish KO vs Sub reasonably; it just can't be
            # trusted for the Decision vs Finish question.
            ml_finish = ko_p + sub_p
            if ml_finish > 0:
                ml_ko_share  = ko_p  / ml_finish
                ml_sub_share = sub_p / ml_finish
            else:
                ml_ko_share = ml_sub_share = 0.5

            rule_finish = rule_ko + rule_sub
            if rule_finish > 0:
                rule_ko_share  = rule_ko  / rule_finish
                rule_sub_share = rule_sub / rule_finish
            else:
                rule_ko_share = rule_sub_share = 0.5

            ko_share  = 0.5 * ml_ko_share  + 0.5 * rule_ko_share
            sub_share = 0.5 * ml_sub_share + 0.5 * rule_sub_share
            share_tot = ko_share + sub_share
            if share_tot > 0:
                ko_share  /= share_tot
                sub_share /= share_tot

            ko_p  = finish_p * ko_share
            sub_p = finish_p * sub_share

            # Winner-conditioned method prediction — use the predicted winner's own
            # historical rates for the Decision/Finish split, then blend ML signal
            # for KO vs Sub.  Previously this used fight-level (both-fighter average)
            # probs, so it would predict the most common method in the fight regardless
            # of who was picked to win.
            if winner_is_red:
                w_ko_r, w_sub_r, w_dec_r = r_ko_r, r_sub_r, r_dec_r
                ml_w_raw = np.array([red_dec_prob, red_ko_prob, red_sub_prob])
            else:
                w_ko_r, w_sub_r, w_dec_r = b_ko_r, b_sub_r, b_dec_r
                ml_w_raw = np.array([blue_dec_prob, blue_ko_prob, blue_sub_prob])

            # Normalize the ML winner-side probs
            ml_w_sum = ml_w_raw.sum()
            if ml_w_sum > 0:
                ml_w_raw = ml_w_raw / ml_w_sum
            ml_w_dec, ml_w_ko, ml_w_sub = ml_w_raw

            # Step 1: Decision vs Finish — winner's rule-based historical rate
            w_dec_p    = w_dec_r
            w_finish_p = 1.0 - w_dec_p

            # Step 2: KO vs Sub within finishes — blend winner's ML + rule-based
            w_finish_r = w_ko_r + w_sub_r
            w_ko_r_sh  = (w_ko_r  / w_finish_r) if w_finish_r > 0 else 0.5
            w_sub_r_sh = (w_sub_r / w_finish_r) if w_finish_r > 0 else 0.5

            ml_fin    = ml_w_ko + ml_w_sub
            ml_ko_sh  = (ml_w_ko  / ml_fin) if ml_fin > 0 else 0.5
            ml_sub_sh = (ml_w_sub / ml_fin) if ml_fin > 0 else 0.5

            w_ko_sh  = 0.5 * ml_ko_sh  + 0.5 * w_ko_r_sh
            w_sub_sh = 0.5 * ml_sub_sh + 0.5 * w_sub_r_sh
            w_sh_tot = w_ko_sh + w_sub_sh
            if w_sh_tot > 0:
                w_ko_sh  /= w_sh_tot
                w_sub_sh /= w_sh_tot

            winner_dec_p = float(w_dec_p)
            winner_ko_p  = float(w_finish_p * w_ko_sh)
            winner_sub_p = float(w_finish_p * w_sub_sh)

            method_names_arr = ['Decision', 'KO/TKO', 'Submission']
            method_arr = np.array([winner_dec_p, winner_ko_p, winner_sub_p])
            method_pred = method_names_arr[int(np.argmax(method_arr))]
            method_conf = float(np.max(method_arr))

            confidence = abs(r_win_prob - 0.5) * 2

            pred = {
                "r_fighter": r_name,
                "b_fighter": b_name,
                "weight_class": weight_class,
                "gender": gender,
                "total_rounds": total_rounds,
                "winner": winner_name,
                "winner_conf": winner_conf,
                "method": method_pred,
                "method_conf": method_conf,
                "winner_dec_p": winner_dec_p,
                "winner_ko_p": winner_ko_p,
                "winner_sub_p": winner_sub_p,
                # Fight-level marginal method probs
                "dec_p": dec_p,
                "ko_p": ko_p,
                "sub_p": sub_p,
                # Red and blue specific
                "r_win_prob": r_win_prob,
                "b_win_prob": b_win_prob,
                "confidence": confidence,
            }
            self.predictions.append(pred)

        return self.predictions

    def _compute_row_features(self, base):
        """
        Compute all derived features (Tiers 0-22) from a single row dict of r_/b_ base columns.
        Converts the dict to a 1-row DataFrame, runs _recompute_derived_features(), and
        returns a flat dict of feature_name -> float.
        Used by _build_fight_feature_vector() for prediction-time feature construction.
        """
        # Convert dict -> 1-row DataFrame
        df1 = pd.DataFrame([base])

        # Ensure event_date exists (needed for Tier 5 Z-score lookup)
        if "event_date" not in df1.columns:
            df1["event_date"] = pd.Timestamp("2025-01-01")

        # Run the shared vectorized feature computation
        df1 = self._recompute_derived_features(df1)

        # Extract all numeric features as a flat dict
        result = {}
        for col in df1.columns:
            if df1[col].dtype == object:
                continue
            try:
                val = df1[col].iloc[0]
                fval = float(val)
                result[col] = 0.0 if (math.isnan(fval) or math.isinf(fval)) else fval
            except (TypeError, ValueError):
                pass

        return result

    def _build_fight_feature_vector(self, r_name, b_name, weight_class, gender, total_rounds):
        """
        Build a single feature vector for a prospective fight.
        Phase 2: Uses proper corner-swap by literally swapping r_/b_ base columns
        and recomputing all derived features, then applying antisymmetric decomposition.
        """
        r_last = self._get_fighter_last_stats(r_name, "r")
        b_last = self._get_fighter_last_stats(b_name, "b")

        if r_last is None and b_last is None:
            return None

        if r_last is None:
            r_last = {}
        if b_last is None:
            b_last = {}

        # Build the base row dict with all r_/b_ columns
        base = {}
        base.update(r_last)
        base.update(b_last)
        base["weight_class"]  = weight_class
        base["gender"]        = gender
        base["total_rounds"]  = total_rounds
        base["total_rounds_num"] = float(total_rounds)
        base["is_title_enc"]  = 0.0
        base["gender_enc"]    = 1.0 if str(gender).lower() == "women" else 0.0

        # ELO and Glicko-2 for original corners
        r_elo = self.feature_engineer.elo_get(r_name)
        b_elo = self.feature_engineer.elo_get(b_name)
        rg    = self.feature_engineer.glicko2_get(r_name)
        bg    = self.feature_engineer.glicko2_get(b_name)
        base["r_elo_pre_fight"]  = r_elo
        base["b_elo_pre_fight"]  = b_elo
        base["r_glicko_pre_r"]   = rg[0]
        base["r_glicko_pre_rd"]  = rg[1]
        base["r_glicko_pre_vol"] = rg[2]
        base["b_glicko_pre_r"]   = bg[0]
        base["b_glicko_pre_rd"]  = bg[1]
        base["b_glicko_pre_vol"] = bg[2]

        # Common opponents and style cluster (same for both orientations)
        common = self.feature_engineer.get_common_opponent_features(r_name, b_name)
        rc = self.feature_engineer.get_fighter_cluster(r_name)
        bc_cluster = self.feature_engineer.get_fighter_cluster(b_name)
        mf = self.feature_engineer.get_style_matchup_features(rc, bc_cluster)
        base.update(common)
        base.update(mf)

        # Compute derived features for original orientation
        orig_feats = self._compute_row_features(base)
        # Also pass through raw base values for columns not covered by _compute_row_features
        for k, v in base.items():
            if k not in orig_feats:
                try:
                    orig_feats[k] = float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    orig_feats[k] = 0.0

        # Build swapped base: literally swap r_* <-> b_* column names
        swap_base = {}
        for k, v in base.items():
            if k.startswith("r_"):
                swap_base["b_" + k[2:]] = v
            elif k.startswith("b_"):
                swap_base["r_" + k[2:]] = v
            else:
                swap_base[k] = v

        # Recompute derived features for swapped orientation
        swap_feats = self._compute_row_features(swap_base)
        for k, v in swap_base.items():
            if k not in swap_feats:
                try:
                    swap_feats[k] = float(v) if v is not None else 0.0
                except (TypeError, ValueError):
                    swap_feats[k] = 0.0

        # Align to the trained feature column order
        ordered_cols = self.feature_cols
        orig_arr = np.array([orig_feats.get(col, 0.0) for col in ordered_cols], dtype=float)
        swap_arr = np.array([swap_feats.get(col, 0.0) for col in ordered_cols], dtype=float)
        orig_arr = np.nan_to_num(orig_arr)
        swap_arr = np.nan_to_num(swap_arr)

        # Antisymmetric decomposition: D = 0.5*(orig-swap), I = 0.5*(orig+swap)
        D   = 0.5 * (orig_arr - swap_arr)
        Inv = 0.5 * (orig_arr + swap_arr)
        feat_decomposed = np.concatenate([D, Inv])
        return feat_decomposed, orig_arr

    def _get_fighter_last_stats(self, fighter_name, corner):
        """Get the most recent feature snapshot from the dataframe for a given fighter."""
        r_rows = self.df[self.df["r_fighter"] == fighter_name]
        b_rows = self.df[self.df["b_fighter"] == fighter_name]

        stats = {}
        last_r = r_rows.iloc[-1] if len(r_rows) > 0 else None
        last_b = b_rows.iloc[-1] if len(b_rows) > 0 else None

        # Pick the most recent
        if last_r is not None and last_b is not None:
            date_r = last_r.get("event_date", pd.NaT)
            date_b = last_b.get("event_date", pd.NaT)
            if pd.notna(date_r) and pd.notna(date_b):
                row = last_r if date_r >= date_b else last_b
                src_corner = "r" if date_r >= date_b else "b"
            else:
                row = last_r
                src_corner = "r"
        elif last_r is not None:
            row = last_r
            src_corner = "r"
        elif last_b is not None:
            row = last_b
            src_corner = "b"
        else:
            return None

        # Map stats from src_corner -> corner prefix
        for col in self.df.columns:
            if col.startswith(f"{src_corner}_"):
                base = col[2:]
                target_col = f"{corner}_{base}"
                val = row.get(col, 0)
                try:
                    val = float(val)
                    if math.isnan(val):
                        val = 0.0
                except (TypeError, ValueError):
                    val = 0.0
                stats[target_col] = val

        # Also copy diff columns (they'll need to be recomputed anyway)
        return stats

    # ── EXPORT TO EXCEL ───────────────────────────────────────────────────────
    def export_predictions_to_excel(self, output_path):
        """Export predictions to Excel in a formatted table."""
        print_section("EXPORTING TO EXCEL")
        self._log(f"Writing to: {output_path}")

        if not self.predictions:
            self._log("No predictions to export.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "UFC Predictions"

        # Styles
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        header_font = Font(bold=True, size=11, color="000000")
        header_align = Alignment(horizontal="left", vertical="center")
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Match Model 3 exact column format
        headers = [
            "Red Fighter",   # col 1
            "Blue Fighter",  # col 2
            "Weight Class",  # col 3
            "Winner",        # col 4
            "Win%",          # col 5  <- percentage
            "Method",        # col 6
            "Method%",       # col 7  <- percentage
            "KO/TKO%",       # col 8  <- percentage
            "Submission%",   # col 9  <- percentage
            "Decision%",     # col 10 <- percentage
        ]
        # 1-based column indices that should be formatted as percentage
        pct_col_indices = {5, 7, 8, 9, 10}

        # Write header
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border

        # Write data
        for row_idx, pred in enumerate(self.predictions, 2):
            # Use winner-specific fields (new 6-class method format)
            winner_name = pred.get("winner") or (
                pred["r_fighter"] if pred.get("r_win_prob", 0.5) > pred.get("b_win_prob", 0.5)
                else pred["b_fighter"]
            )
            winner_conf  = pred.get("winner_conf",  max(pred.get("r_win_prob", 0.5), pred.get("b_win_prob", 0.5)))
            method_label = pred.get("method", "Decision")
            method_conf  = pred.get("method_conf",  0.0)
            ko_p         = pred.get("winner_ko_p",  pred.get("ko_p", 0.0))
            sub_p        = pred.get("winner_sub_p", pred.get("sub_p", 0.0))
            dec_p        = pred.get("winner_dec_p", pred.get("dec_p", 0.0))

            row_data = [
                pred["r_fighter"],  # Red Fighter
                pred["b_fighter"],  # Blue Fighter
                pred["weight_class"],  # Weight Class
                winner_name,        # Winner
                winner_conf,        # Win%   (0-1 decimal, formatted as %)
                method_label,       # Method (for predicted winner)
                method_conf,        # Method% (normalized for winner)
                ko_p,               # KO/TKO%  (winner-specific)
                sub_p,              # Submission%  (winner-specific)
                dec_p,              # Decision%  (winner-specific)
            ]

            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = border
                if col_idx in pct_col_indices:
                    cell.number_format = "0.00%"

        # Auto column widths
        for col_idx in range(1, len(headers) + 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        max_len = max(max_len, len(str(cell.value or "")))
                    except Exception:
                        pass
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(output_path)
        self._log(f"Saved predictions to {output_path}")
        print_metric("Fights exported:", len(self.predictions))


# ─────────────────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────────────────
class _GUIConsole:
    """Redirects sys.stdout / sys.stderr writes into a tkinter Text widget.
    Handles \\r (carriage-return) so the Optuna progress bar overwrites in place.

    Uses 'end-1c' to resolve the last real line number at overwrite time —
    avoids every tkinter mark/gravity ambiguity entirely.
    """

    def __init__(self, widget, root, real_stdout=None):
        self.widget       = widget
        self.root         = root
        self.real_stdout  = real_stdout if real_stdout is not None else _REAL_STDOUT

    def reset(self):
        pass  # no state to clear with this approach

    def write(self, text):
        if self.real_stdout is not None:
            try:
                self.real_stdout.write(text)
                self.real_stdout.flush()
            except UnicodeEncodeError:
                try:
                    safe = text.encode(
                        getattr(self.real_stdout, 'encoding', 'utf-8') or 'utf-8',
                        errors='replace'
                    ).decode(
                        getattr(self.real_stdout, 'encoding', 'utf-8') or 'utf-8',
                        errors='replace'
                    )
                    self.real_stdout.write(safe)
                    self.real_stdout.flush()
                except Exception:
                    pass
            except Exception:
                pass
        def _update():
            if '\r' in text:
                parts = text.split('\r')
                # Write any text that comes before the first \r normally
                if parts[0]:
                    self.widget.insert(tk.END, parts[0])
                # Each segment after a \r overwrites the current last line
                for seg in parts[1:]:
                    # "end-1c" is the last real character (avoids the mandatory
                    # trailing-newline ambiguity that "end" / "end linestart" have)
                    last_char_idx = self.widget.index("end-1c")
                    line_num      = last_char_idx.split('.')[0]
                    # Delete everything on that line (preserves the newline itself)
                    self.widget.delete(f"{line_num}.0", f"{line_num}.end")
                    if seg:
                        self.widget.insert(f"{line_num}.0", seg)
            else:
                self.widget.insert(tk.END, text)
            self.widget.see(tk.END)
        self.root.after(0, _update)

    def flush(self):
        pass


class UFCPredictorGUI:

    def __init__(self, root):
        self.root = root
        self.root.title("UFC FIGHT PREDICTOR")
        self.root.geometry("900x900")
        self.root.resizable(True, True)
        self.root.configure(bg="#1a1a2e")

        self.predictor = None
        self.data_path_var = tk.StringVar(value=DEFAULT_DATA_PATH)
        self.output_path_var = tk.StringVar(
            value=os.path.join(SCRIPT_DIR, "UFC_predictions.xlsx")
        )
        self.is_running = False

        self._build_ui()

    def _build_ui(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", font=("Helvetica", 10, "bold"), padding=6)
        style.configure("TLabel", background="#1a1a2e", foreground="#e0e0e0",
                        font=("Helvetica", 10))
        style.configure("TEntry", font=("Helvetica", 10))
        style.configure("TFrame", background="#1a1a2e")

        # ── Title ─────────────────────────────────────────────────────────
        title_frame = tk.Frame(self.root, bg="#16213e", pady=12)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame, text="UFC FIGHT PREDICTOR",
            font=("Helvetica", 20, "bold"), fg="#e94560", bg="#16213e"
        ).pack()

        # ── Main container ────────────────────────────────────────────────
        main = tk.Frame(self.root, bg="#1a1a2e", padx=16, pady=10)
        main.pack(fill="both", expand=True)

        # ── Data file row ─────────────────────────────────────────────────
        df_frame = tk.Frame(main, bg="#1a1a2e")
        df_frame.pack(fill="x", pady=4)
        tk.Label(df_frame, text="Data File:", width=12, anchor="w",
                 bg="#1a1a2e", fg="#e0e0e0", font=("Helvetica", 10)).pack(side="left")
        tk.Entry(df_frame, textvariable=self.data_path_var, font=("Helvetica", 10),
                 bg="#0f3460", fg="white", insertbackground="white",
                 relief="flat", width=60).pack(side="left", padx=4, fill="x", expand=True)
        tk.Button(df_frame, text="Browse", command=self._browse_data,
                  font=("Helvetica", 9, "bold"), bg="#e94560", fg="white",
                  relief="flat", padx=10, cursor="hand2").pack(side="left", padx=4)

        # ── Output file row ───────────────────────────────────────────────
        out_frame = tk.Frame(main, bg="#1a1a2e")
        out_frame.pack(fill="x", pady=4)
        tk.Label(out_frame, text="Output File:", width=12, anchor="w",
                 bg="#1a1a2e", fg="#e0e0e0", font=("Helvetica", 10)).pack(side="left")
        tk.Entry(out_frame, textvariable=self.output_path_var, font=("Helvetica", 10),
                 bg="#0f3460", fg="white", insertbackground="white",
                 relief="flat", width=60).pack(side="left", padx=4, fill="x", expand=True)
        tk.Button(out_frame, text="Browse", command=self._browse_output,
                  font=("Helvetica", 9, "bold"), bg="#e94560", fg="white",
                  relief="flat", padx=10, cursor="hand2").pack(side="left", padx=4)

        # ── Fight input label ─────────────────────────────────────────────
        lbl_frame = tk.Frame(main, bg="#1a1a2e")
        lbl_frame.pack(fill="x", pady=(8, 2))
        tk.Label(
            lbl_frame,
            text="Enter Fights  (one per line: Red Fighter, Blue Fighter, Weight Class, Gender, Rounds)",
            bg="#1a1a2e", fg="#a0c4ff", font=("Helvetica", 9, "italic")
        ).pack(anchor="w")

        # ── ScrolledText fight input ──────────────────────────────────────
        self.fight_input = scrolledtext.ScrolledText(
            main, height=11, font=("Courier New", 10),
            bg="#0f3460", fg="white", insertbackground="white",
            relief="flat", wrap="word"
        )
        self.fight_input.pack(fill="both", expand=True, pady=4)

        # ── Buttons row ───────────────────────────────────────────────────
        btn_frame = tk.Frame(main, bg="#1a1a2e")
        btn_frame.pack(fill="x", pady=6)

        tk.Button(btn_frame, text="Load Sample", command=self._load_sample,
                  font=("Helvetica", 10, "bold"), bg="#533483", fg="white",
                  relief="flat", padx=14, cursor="hand2").pack(side="left", padx=4)
        tk.Button(btn_frame, text="Clear", command=self._clear_input,
                  font=("Helvetica", 10, "bold"), bg="#444466", fg="white",
                  relief="flat", padx=14, cursor="hand2").pack(side="left", padx=4)

        self.run_btn = tk.Button(
            btn_frame, text="Generate Predictions",
            command=self._run_predictions,
            font=("Helvetica", 11, "bold"), bg="#e94560", fg="white",
            relief="flat", padx=20, cursor="hand2"
        )
        self.run_btn.pack(side="right", padx=4)

        # ── Console output panel ──────────────────────────────────────────
        con_label_frame = tk.Frame(main, bg="#1a1a2e")
        con_label_frame.pack(fill="x", pady=(8, 2))
        tk.Label(
            con_label_frame, text="Output",
            bg="#1a1a2e", fg="#a0c4ff", font=("Helvetica", 9, "italic")
        ).pack(side="left", anchor="w")
        tk.Button(
            con_label_frame, text="Clear",
            command=lambda: self.console_output.delete("1.0", tk.END),
            font=("Helvetica", 8), bg="#333355", fg="#a0c4ff",
            relief="flat", padx=6, cursor="hand2"
        ).pack(side="right")

        self.console_output = scrolledtext.ScrolledText(
            main, height=11, font=("Courier New", 9),
            bg="#0a0a1a", fg="#c8ffc8", insertbackground="white",
            relief="flat", wrap="word", state="normal"
        )
        self.console_output.pack(fill="both", expand=True, pady=(0, 4))
        self._console = _GUIConsole(self.console_output, self.root, real_stdout=sys.stdout)

    def _browse_data(self):
        path = filedialog.askopenfilename(
            title="Select Fight Data CSV",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir=SCRIPT_DIR
        )
        if path:
            self.data_path_var.set(path)

    def _browse_output(self):
        path = filedialog.asksaveasfilename(
            title="Save Predictions As",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=SCRIPT_DIR,
            initialfile="UFC_predictions.xlsx"
        )
        if path:
            self.output_path_var.set(path)

    def _load_sample(self):
        sample = (
            "Max Holloway,Dustin Poirier,Lightweight,Men,5\n"
            "Ilia Topuria,Charles Oliveira,Lightweight,Men,5\n"
            "Tatiana Suarez,Amanda Lemos,Strawweight,Women,3"
        )
        self.fight_input.delete("1.0", tk.END)
        self.fight_input.insert("1.0", sample)

    def _clear_input(self):
        self.fight_input.delete("1.0", tk.END)

    def _run_predictions(self):
        if self.is_running:
            return
        thread = threading.Thread(target=self._run_predictions_thread, daemon=True)
        thread.start()

    def _run_predictions_thread(self):
        self.is_running = True
        self.run_btn.config(state="disabled", text="Running...")
        _orig_stdout, _orig_stderr = sys.stdout, sys.stderr
        self._console.reset()
        sys.stdout = self._console
        sys.stderr = self._console
        try:
            data_path = self.data_path_var.get().strip()
            output_path = self.output_path_var.get().strip()

            if not os.path.isfile(data_path):
                messagebox.showerror("Error", f"Data file not found:\n{data_path}")
                return

            # Parse fights
            fights_text = self.fight_input.get("1.0", tk.END).strip()
            if not fights_text:
                messagebox.showerror("Error", "Please enter at least one fight.")
                return

            fights = self._parse_fights(fights_text)
            if not fights:
                return

            # Instantiate predictor
            self.predictor = UFCPredictor(data_path=data_path)

            self.predictor.load_data()
            self.predictor.fix_data_leakage()
            self.predictor.build_all_features()
            self.predictor.train()
            self.predictor.predict_upcoming_fights(fights)
            self.predictor.export_predictions_to_excel(output_path)

            messagebox.showinfo(
                "Complete",
                f"Predictions generated successfully!\n\nSaved to:\n{output_path}"
            )

        except Exception as e:
            tb = traceback.format_exc()
            print(f"Error: {e}")
            messagebox.showerror("Error", f"An error occurred:\n\n{e}\n\n{tb[-500:]}")
        finally:
            sys.stdout, sys.stderr = _orig_stdout, _orig_stderr
            self.is_running = False
            self.run_btn.config(state="normal", text="Generate Predictions")

    def _parse_fights(self, text):
        fights = []
        lines = [l.strip() for l in text.splitlines() if l.strip()]
        for i, line in enumerate(lines, 1):
            parts = [p.strip() for p in line.split(",")]
            if len(parts) != 5:
                messagebox.showerror(
                    "Parse Error",
                    f"Line {i}: Expected 5 comma-separated fields.\n"
                    f"Format: Red Fighter, Blue Fighter, Weight Class, Gender, Rounds\n"
                    f"Got: {line}"
                )
                return []
            r_fighter, b_fighter, weight_class, gender, rounds_str = parts
            try:
                total_rounds = int(rounds_str)
            except ValueError:
                messagebox.showerror(
                    "Parse Error",
                    f"Line {i}: Rounds must be an integer, got '{rounds_str}'"
                )
                return []
            fights.append({
                "r_fighter": r_fighter,
                "b_fighter": b_fighter,
                "weight_class": weight_class,
                "gender": gender,
                "total_rounds": total_rounds,
            })
        return fights


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    mp.freeze_support()

    print_section("UFC FIGHT PREDICTOR")
    print_step(f"Python: {sys.version.split()[0]}")
    print_step(f"Libraries: XGB={HAS_XGB}, LGB={HAS_LGB}, CAT={HAS_CAT}, Optuna={HAS_OPTUNA}")
    print_step(f"CPU cores: {mp.cpu_count()}  |  Safe n_jobs: {SAFE_N_JOBS}")
    print_step(f"Script dir: {SCRIPT_DIR}")
    print_step(f"Default data: {DEFAULT_DATA_PATH}")

    root = tk.Tk()
    app = UFCPredictorGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
